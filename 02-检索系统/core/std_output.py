# -*- coding: utf-8 -*-
"""标准范式输出核心 (core/std_output.py).

以 标准字段结构.json (从标准字段库 Excel 提取的父子级→标签规范) 为
**唯一输出范式**, 供 GUI「导入 MSDS → 输出内容」与 CLI 批量检索共用:

- 标准结构: 节 → 父子级组 → 标签 (含锚点/成分/通栏类型)
- 取值: 任意 ParseResult 按标准结构取值, 标准库有而文件没有 → '无数据'
- 多值扩列: 同标签多值自动生成 (2)(3)...; 成分槽按实际成分数展开
- 输出: render_std_matrix(results) → 列表结构 (供 xlsx/JSON 渲染)

标准结构 JSON 由 tools/extract_std_structure.py 从 Excel 生成;
若缺失, 自动从默认 Excel 路径提取.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from .schema import standard_name

ROOT = Path(__file__).resolve().parent.parent
STRUCT_JSON = ROOT / 'outputs' / '标准字段结构.json'
STD_XLSX = (r'F:\正式项目与模块化内容\Word 覆写模块\数据库\正式库'
            r'\标准字段数据库Excel\Excel.xlsx')

_SEQ_RE = re.compile(r'^\d+(?:\.\d+)*\s*')
_COMP_SLOT_RE = re.compile(r'^化学品名称（(\d+)）$')
_COMP_CAS_RE = re.compile(r'^CAS编号（(\d+)）$')
_COMP_CONC_RE = re.compile(r'^含量%（w/w）（(\d+)）$')

_STRUCT_CACHE: dict | None = None


def bare(name):
    return _SEQ_RE.sub('', name or '').strip()


# ------------------------------------------------------------------
# 标准结构加载
# ------------------------------------------------------------------

def _extract_from_excel() -> dict:
    """从默认 Excel 提取标准结构 (STRUCT_JSON 缺失时兜底)."""
    from openpyxl import load_workbook
    wb = load_workbook(STD_XLSX, data_only=True)
    ws = wb['Sheet1']
    ncol = ws.max_column
    cur_sec = None
    sec_of_col = {}
    for ci in range(1, ncol + 1):
        r1 = str(ws.cell(1, ci).value or '').strip()
        m = re.match(r'Section(\d+)', r1)
        if m:
            cur_sec = int(m.group(1))
        elif r1:
            m2 = re.match(r'(\d+)\.', r1)
            if m2:
                cur_sec = int(m2.group(1))
        sec_of_col[ci] = cur_sec
    first_of_group = {}
    for ci in range(1, ncol + 1):
        parent = str(ws.cell(2, ci).value or '').strip()
        sec = sec_of_col.get(ci)
        if sec is None or not parent:
            continue
        first_of_group.setdefault((sec, parent), ci)
    columns = []
    for ci in range(1, ncol + 1):
        parent = str(ws.cell(2, ci).value or '').strip()
        label = str(ws.cell(3, ci).value or '').strip()
        sec = sec_of_col.get(ci)
        if sec is None or not label:
            continue
        if _COMP_SLOT_RE.match(label) or _COMP_CAS_RE.match(label) \
                or _COMP_CONC_RE.match(label):
            typ = 'comp'
        elif parent == '通栏说明':
            typ = 'note'
        elif (bare(parent) == label
                and first_of_group.get((sec, parent)) == ci):
            typ = 'anchor'
        else:
            typ = 'field'
        columns.append({'sec': sec, 'parent': parent, 'label': label,
                        'type': typ})
    SEC_TITLES = {
        0: 'Section0 页眉页脚', 1: 'Section1 物料及供应商标识',
        2: 'Section2 危险性概述', 3: 'Section3 成分/组成资料',
        4: 'Section4 急救措施', 5: 'Section5 消防措施',
        6: 'Section6 意外泄漏措施', 7: 'Section7 操作和储存',
        8: 'Section8 接触控制/个人防护', 9: 'Section9 物理和化学特性',
        10: 'Section10 稳定性和反应性', 11: 'Section11 毒性资料',
        12: 'Section12 生态信息', 13: 'Section13 处理注意事项',
        14: 'Section14 运输信息', 15: 'Section15 法规信息',
        16: 'Section16 其他信息',
    }
    sections = []
    for sec in sorted({c['sec'] for c in columns}):
        groups = []
        cur_parent = None
        cur_fields = []
        cur_anchor = False
        cur_note = False

        def flush():
            nonlocal cur_parent, cur_fields, cur_anchor, cur_note
            if cur_parent is not None:
                groups.append({'parent': cur_parent, 'anchor': cur_anchor,
                               'note': cur_note, 'fields': cur_fields})
            cur_parent, cur_fields, cur_anchor, cur_note = None, [], False, False

        for col in columns:
            if col['sec'] != sec:
                continue
            parent = col['parent'] or '(节下)'
            typ = col['type']
            if typ == 'anchor':
                flush()
                cur_parent, cur_anchor, cur_fields = parent, True, []
            elif cur_parent is None:
                flush()
                cur_parent, cur_anchor = parent, False
                cur_note = (typ == 'note')
                cur_fields = [col['label']]
            elif parent == cur_parent:
                cur_fields.append(col['label'])
            else:
                flush()
                cur_parent, cur_anchor = parent, False
                cur_note = (typ == 'note')
                cur_fields = [col['label']]
        flush()
        sections.append({'number': sec,
                         'title': SEC_TITLES.get(sec, f'Section{sec}'),
                         'groups': groups})
    return {'version': '1.0', 'sections': sections, 'columns': columns}


def load_structure() -> dict:
    """加载标准结构 (优先 JSON, 兜底 Excel 提取)."""
    global _STRUCT_CACHE
    if _STRUCT_CACHE is not None:
        return _STRUCT_CACHE
    if STRUCT_JSON.exists():
        try:
            _STRUCT_CACHE = json.loads(STRUCT_JSON.read_text(encoding='utf-8'))
            return _STRUCT_CACHE
        except Exception:
            pass
    _STRUCT_CACHE = _extract_from_excel()
    return _STRUCT_CACHE


# ------------------------------------------------------------------
# 文件侧取值模型
# ------------------------------------------------------------------

def tpl_values(result) -> tuple[dict, dict, dict, dict]:
    """ParseResult → (精确表, 归并表, 成分数, notes).
    精确表 {节: {bare标签: [值]}} | 归并表 {节: {标准名: [值]}}
    """
    from .extract import build_hierarchy
    exact: dict[int, dict[str, list]] = {}
    merged: dict[int, dict[str, list]] = {}
    comps: dict[int, int] = {}
    notes: dict[int, list[str]] = {}

    def _put(sec, label, value):
        exact.setdefault(sec, {}).setdefault(bare(label), []).append(value)
        merged.setdefault(sec, {}).setdefault(
            standard_name(sec, label), []).append(value)

    for sn in build_hierarchy(result):
        for b in sn.big_titles:
            if b.kind == 'field' and b.value:
                _put(sn.number, b.title, b.value)
            for f in b.children:
                if f.kind == 'subtable':
                    for h, v in zip(f.sub_header,
                                    (f.sub_rows[0] if f.sub_rows else [])):
                        _put(sn.number, h, v)
                    continue
                if f.kind == 'note':
                    notes.setdefault(sn.number, []).append(f.value)
                    continue
                if not f.label:
                    continue
                _put(sn.number, f.label, f.value)
        for f in sn.direct_fields:
            if f.kind == 'component':
                comps[sn.number] = comps.get(sn.number, 0) + 1
                continue
            if f.kind == 'note':
                notes.setdefault(sn.number, []).append(f.value)
                continue
            if not f.label:
                continue
            _put(sn.number, f.label, f.value)
    return exact, merged, comps, notes


def std_merged_counts(columns: list[dict]) -> dict:
    """标准库列中每个标准名的出现次数 (判定归并是否唯一)."""
    from collections import Counter
    cnt = Counter()
    for c in columns:
        if c['type'] == 'field':
            cnt[(c['sec'], standard_name(c['sec'], c['label']))] += 1
    return cnt


# ------------------------------------------------------------------
# 单元格取值
# ------------------------------------------------------------------

def cell_value(result, sec, label, typ, multi_idx, std_cnt,
               exact, merged, comps, notes):
    """按标准结构取一个单元格值; None → 无数据. 锚点列恒空."""
    if typ == 'comp':
        m = (_COMP_SLOT_RE.match(label) or _COMP_CAS_RE.match(label)
             or _COMP_CONC_RE.match(label))
        idx = int(m.group(1)) - 1
        s3 = result.sections.get(3)
        if not s3 or idx >= len(s3.components):
            return None
        c = s3.components[idx]
        if '名称' in label:
            return c.name
        if 'CAS' in label:
            return c.cas
        return c.conc
    if typ == 'note':
        ns = notes.get(sec, [])
        if label == '通栏说明':
            return '\n'.join(ns) if ns else None
        m = re.match(r'^说明(\d+)$', label)
        if m and int(m.group(1)) <= len(ns):
            return ns[int(m.group(1)) - 1]
        if sec == 16 and label == '其他信息' and ns:
            return '\n'.join(ns)
        return None
    bl = bare(label)
    lst = exact.get(sec, {}).get(bl)
    if lst:
        return _pick(lst, multi_idx)
    std = standard_name(sec, label)
    if std_cnt.get((sec, std), 0) != 1:
        return None
    lst = merged.get(sec, {}).get(std)
    if not lst:
        return None
    return _pick(lst, multi_idx)


def _pick(lst, multi_idx):
    if multi_idx == 1:
        return lst[0]
    return lst[multi_idx - 1] if multi_idx - 1 < len(lst) else None


# ------------------------------------------------------------------
# 标准矩阵输出 (多文件)
# ------------------------------------------------------------------

def build_std_matrix(results: list) -> dict:
    """多文件标准矩阵.

    Returns:
        {columns: [(sec, parent, label, typ, anchor, multi_idx), ...],
         rows: [(型号, {列序号: 值|None}), ...],
         groups: [(start, end, 标题), ...]}
    """
    struct = load_structure()
    cols0 = [dict(c, anchor=(c['type'] == 'anchor'))
             for c in struct['columns']]
    std_cnt = std_merged_counts(cols0)

    # 多值扩列 + 成分超槽
    max_multi: dict[tuple, int] = {}
    max_comp = 5
    for c in cols0:
        if c['type'] == 'field' and not c['anchor']:
            max_multi.setdefault((c['sec'], bare(c['label'])), 1)
    tpl_cache = {}
    for r in results:
        exact, _m, comps, _n = tpl_values(r)
        tpl_cache[id(r)] = (exact, _m, comps, _n)
        for sec, m in exact.items():
            for bl, vlist in m.items():
                key = (sec, bl)
                if key in max_multi:
                    max_multi[key] = max(max_multi[key], len(vlist))
        for sec, n in comps.items():
            max_comp = max(max_comp, n)
    cols = []
    for c in cols0:
        cols.append((c['sec'], c['parent'], c['label'], c['type'],
                     c['anchor'], 1))
        if c['type'] != 'field' or c['anchor']:
            continue
        for i in range(2, max_multi.get((c['sec'], bare(c['label'])), 1) + 1):
            cols.append((c['sec'], c['parent'], c['label'], c['type'],
                         c['anchor'], i))
    for ci in range(6, max_comp + 1):
        for suffix, lab in (('名称', f'化学品名称（{ci}）'),
                            ('CAS', f'CAS编号（{ci}）'),
                            ('含量', f'含量%（w/w）（{ci}）')):
            cols.append((3, '3.2 成分', lab, 'comp', False, 1))

    # 分组 (行1 合并块)
    groups = []
    cur_title = None
    cur_start = None
    for idx, (sec, parent, label, typ, anchor, mi) in enumerate(cols):
        title = parent if parent else f'Section{sec}'
        if title != cur_title:
            if cur_start is not None:
                groups.append((cur_start, idx - 1, cur_title))
            cur_start = idx
            cur_title = title
    if cur_start is not None:
        groups.append((cur_start, len(cols) - 1, cur_title))

    # 数据行
    rows = []
    for r in results:
        exact, merged, comps, notes = tpl_cache[id(r)]
        model = _model_of(r)
        cells = []
        for sec, parent, label, typ, anchor, mi in cols:
            if anchor:
                cells.append(None)
            else:
                cells.append(cell_value(r, sec, label, typ, mi, std_cnt,
                                        exact, merged, comps, notes))
        rows.append((model, cells))
    return {'columns': cols, 'rows': rows, 'groups': groups}


def _model_of(result) -> str:
    sec1 = result.sections.get(1)
    if sec1:
        for row in sec1.iter_rows():
            if row.kind == 'field' and row.label == '产品名称' \
                    and row.value.strip():
                return row.value.strip()
    from pathlib import Path
    return Path(result.file_name).stem


# ------------------------------------------------------------------
# 单文件标准 JSON (GUI 导入 → 输出内容)
# ------------------------------------------------------------------

def render_std_json(result) -> dict:
    """单文件按标准结构输出 (供 GUI「导入 MSDS → 输出内容」).

    结构: {节: {父级组: [{标签, 值|无数据, 类型}...]}}
    与标准结构 sections 完全对应, 便于与标准库逐项比对.
    """
    struct = load_structure()
    cols0 = [dict(c, anchor=(c['type'] == 'anchor'))
             for c in struct['columns']]
    std_cnt = std_merged_counts(cols0)
    exact, merged, comps, notes = tpl_values(result)

    out = []
    for sec_def in struct['sections']:
        sec_no = sec_def['number']
        sec_out = {'number': sec_no, 'title': sec_def['title'],
                   'groups': []}
        for g in sec_def['groups']:
            g_out = {'parent': g['parent'], 'anchor': g['anchor'],
                     'note': g.get('note', False), 'fields': []}
            for label in g['fields']:
                # 找该标签对应标准列 (type)
                typ = 'field'
                for c in cols0:
                    if c['sec'] == sec_no and c['label'] == label \
                            and c['type'] != 'anchor':
                        typ = c['type']
                        break
                v = cell_value(result, sec_no, label, typ, 1, std_cnt,
                               exact, merged, comps, notes)
                g_out['fields'].append({
                    'label': label,
                    'type': typ,
                    'value': v if v is not None else '无数据',
                    'matched': v is not None,
                })
            sec_out['groups'].append(g_out)
        out.append(sec_out)
    return {'file': result.file_name, 'sections': out}
