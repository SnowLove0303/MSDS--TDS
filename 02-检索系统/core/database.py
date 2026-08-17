# -*- coding: utf-8 -*-
"""三表数据库模块: MSDS 标准化建库 (标准汇总表 + 原始表 + 字段映射表).

设计背景
--------
不同 MSDS 写法差异大 (标签/单位括号/父子级不一), 直接透视会列爆炸.
本模块按"三表"分层建库, 对应标准化流程的原始数据 / 标准数据 / 映射数据:

  表1 msds_summary (标准汇总总表):
      每个型号一行, 每标准字段一列 (经 Schema 归一化 + 单位剥离).
      最标准最统一, 是后续覆写/数据库的理想数据.

  表2 msds_raw (原始表):
      每个型号 x 每节 x 原始标签+字段 逐行 (长表), 保留读取的原始写法.
      未做任何归一化, 供审计原始来源 / 追溯差异.

  表3 msds_mapping (字段映射表):
      标准字段(老大) -> 各子级写法 (同义别名 / 单位变体 / 来源文件数).
      记录"A 标准化后是 A, 但 A 的子级包括 A 和 A(xx)", 供后续
      扩库 / 覆写 / 建列时查证映射关系.

用法 (命令行):
  python tools/build_database.py <入库目录> <输出xlsx>
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .docx_reader import read_msds
from .schema import (
    _ALIAS_LOOKUP,
    _build_alias_lookup,
    _strip_s9_unit,
    SECTION_SCHEMAS,
    standard_field_of,
    standard_fields,
    standard_name,
)

FONT = "Microsoft YaHei"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
MAP_STD_FILL = PatternFill("solid", fgColor="E2EFDA")
MODEL_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _sheet_header(ws, titles, widths):
    """写表头 (深蓝底白字) + 冻结首行 + 列宽."""
    for i, (t, w) in enumerate(zip(titles, widths), start=1):
        c = ws.cell(1, i, t)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _style_rows(ws, start, end, cols, bold_cols=()):
    """数据区统一边框/字体/换行."""
    for ri in range(start, end + 1):
        for ci in range(1, cols + 1):
            c = ws.cell(ri, ci)
            c.border = BORDER
            c.font = Font(name=FONT, size=9, bold=(ci in bold_cols))
            c.alignment = Alignment(vertical="top", wrap_text=True)


# ------------------------------------------------------------------
# 表2: 原始表 (长表, 未归一化)
# ------------------------------------------------------------------

def _raw_rows(result):
    """原始表行: [型号, 节, 类型, 序号, 原始标签, 原始字段] (未归一化)."""
    rows = []
    model = result.file_name.replace(".docx", "")
    for num in sorted(result.sections):
        sec = result.sections[num]
        for r in sec.iter_rows():
            if r.kind == "section":
                continue
            rows.append([model, num, r.kind, r.seq, r.label, r.value])
    return rows


# ------------------------------------------------------------------
# 表3: 字段映射表 (标准字段 -> 子级写法)
# ------------------------------------------------------------------

def _mapping_rows(results):
    """映射表: [节, 标准字段(老大), 子级写法, 类型, 出现次数].

    类型:
      - 标准名:   标准字段自身
      - 同义别名: Schema 中定义的 aliases
      - 单位变体: S9 自动检测剥离单位括号后的变体 (非显式别名)
    出现次数: 全库中该子级写法的 field 出现次数 (0 = 未在库中出现).
    """
    _build_alias_lookup()
    seen = Counter()
    for result in results:
        for num, sec in result.sections.items():
            for r in sec.iter_rows():
                if r.kind == "field" and r.label.strip():
                    seen[(num, r.label)] += 1

    rows = []
    all_fields = list(SECTION_SCHEMAS.items())
    for num in range(1, 17):
        flds = standard_fields(num)
        if num not in SECTION_SCHEMAS and flds:
            all_fields.append((num, flds))

    for num, flds in all_fields:
        for f in flds:
            rows.append([num, f.name, f.name, "标准名", seen.get((num, f.name), 0)])
            for alias in f.aliases:
                rows.append([num, f.name, alias, "同义别名",
                             seen.get((num, alias), 0)])
        # S9 单位变体自动检测: 全库出现的 S9 标签, 剥离单位后归并到标准名
        if num == 9:
            lookup = _ALIAS_LOOKUP.get(9, {})
            for (n, lbl), cnt in seen.items():
                if n != 9 or lbl in lookup:
                    continue
                stripped = _strip_s9_unit(lbl)
                if stripped and stripped != lbl and stripped in lookup:
                    rows.append([9, lookup[stripped], lbl, "单位变体", cnt])
    return rows


# ------------------------------------------------------------------
# 表1: 标准汇总总表 (透视, 每型号一行)
# ------------------------------------------------------------------

def _schema_field_names(num: int) -> set[str]:
    """该节 Schema 定义的标准字段名集合 (含 collapse 字段)."""
    return {f.name for f in standard_fields(num)}


def _summary_columns(results):
    """标准字段列集合 [(节, 标准字段名)]: Schema 顺序 + 数据补充."""
    cols = []
    seen = set()
    for num, flds in SECTION_SCHEMAS.items():
        for f in flds:
            if f.collapse or (num, f.name) in seen:
                continue
            seen.add((num, f.name))
            cols.append((num, f.name))
    # 未覆盖节 (S3/S16 等) 从实际数据补列; collapse 分组不产列.
    # 已定义 Schema 的节: 未命中的子标签 (S11 方法/物种/结果 等) 不补列,
    # 由 _summary_cell 归并到前一个标准字段值, 避免子标签列爆炸.
    for result in results:
        for num, sec in result.sections.items():
            std_names = _schema_field_names(num)
            for r in sec.iter_rows():
                if r.kind != "field":
                    continue
                std = standard_name(num, r.label)
                if not std or (num, std) in seen:
                    continue
                f = standard_field_of(num, r.label)
                if f is not None and f.collapse:
                    continue
                if std_names and std not in std_names:
                    continue          # 子标签归并, 不产列
                seen.add((num, std))
                cols.append((num, std))
    return cols


def _summary_cell(result, num, std):
    """取一个 ParseResult 某节某标准字段的值 (含子标签归并).

    规则: 命中的标准字段输出自身值; 紧邻其后的**未命中** field 行
    (子标签, 如 S11 毒理的方法/物种/结果、S2 的物理危害/皮肤刺激 等)
    以 '标签: 值' 格式归并到前一个标准字段值中 — 使父子级内容同列,
    避免成分列因子标签写法不同而爆炸.
    """
    sec = result.sections.get(num)
    if not sec:
        return ""
    std_names = _schema_field_names(num)
    parts: list[str] = []
    active = False   # 是否正处在 std 字段的覆盖区 (其后子标签归并到此)
    # 值混入标签: '有效成分/% / 30±1' → 值=30±1  (标签/单位 与 值 用 ' / ' 分隔)
    _label_value_re = re.compile(r"^(.*?)\s*/\s*([\d.\-±~%]+)$")
    for r in sec.iter_rows():
        if r.kind != "field":
            continue
        name = standard_name(num, r.label)
        if name in std_names:
            active = (name == std)
            if active:
                v = r.value.strip()
                if not v:
                    m2 = _label_value_re.match(r.label.strip())
                    if m2:
                        v = m2.group(2).strip()
                if v:
                    parts.append(v)
            continue
        # 未命中字段 → 归并到前一个标准字段
        if active and r.label.strip() and r.value.strip():
            parts.append(f"{r.label}: {r.value}")
    return "\n".join(parts)


# ------------------------------------------------------------------
# 主入口
# ------------------------------------------------------------------

def build_database(in_dir, out_path):
    """生成三表 Excel (标准汇总表 + 原始表 + 字段映射表).

    Returns: {files, failed, cols, rows, mappings, raw_rows, out}
    """
    in_dir = Path(in_dir)
    files = sorted(p for p in in_dir.glob("*.docx")
                   if not p.name.startswith("~$"))
    if not files:
        raise FileNotFoundError(f"目录中未找到 docx 文件: {in_dir}")

    ok_files = []
    results = []
    failed = []
    for p in files:
        try:
            results.append(read_msds(p))
            ok_files.append(p)
        except Exception as exc:
            failed.append(f"{p.name} ({exc.__class__.__name__})")
    if not ok_files:
        raise RuntimeError(f"目录下 {len(files)} 个 docx 全部读取失败:\n"
                           + "\n".join(failed))

    wb = Workbook()

    # ---- 表1 标准汇总总表 ----
    ws = wb.active
    ws.title = "1_标准汇总总表"
    cols = _summary_columns(results)
    c0 = ws.cell(1, 1, "型号")
    c0.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c0.fill = HEAD_FILL
    c0.alignment = Alignment(horizontal="center", vertical="center")
    c0.border = BORDER
    ws.column_dimensions["A"].width = 22
    for i, (num, std) in enumerate(cols, start=2):
        c = ws.cell(1, i, f"S{num}·{std}")
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = 18
    for ri, (p, r) in enumerate(zip(ok_files, results), start=2):
        model = p.name.replace(".docx", "")
        mc = ws.cell(ri, 1, model)
        mc.font = Font(name=FONT, size=9, bold=True)
        mc.fill = MODEL_FILL
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = BORDER
        for ci, (num, std) in enumerate(cols, start=2):
            v = _summary_cell(r, num, std)
            c = ws.cell(ri, ci, v if v else "无数据")
            c.font = Font(name=FONT, size=8)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[ri].height = 60
    ws.freeze_panes = "B2"

    # ---- 表2 原始表 (长表) ----
    ws2 = wb.create_sheet("2_原始表")
    _sheet_header(ws2, ["型号", "节", "类型", "序号", "原始标签", "原始字段"],
                  [24, 5, 7, 6, 34, 60])
    ri = 2
    for p, r in zip(ok_files, results):
        for row in _raw_rows(r):
            for ci, v in enumerate(row, start=1):
                ws2.cell(ri, ci, v)
            ri += 1
    _style_rows(ws2, 2, ri - 1, 6, bold_cols=(1,))

    # ---- 表3 字段映射表 ----
    ws3 = wb.create_sheet("3_字段映射表")
    _sheet_header(ws3, ["节", "标准字段(老大)", "子级写法", "类型", "出现次数"],
                  [6, 30, 40, 10, 10])
    mrows = _mapping_rows(results)
    for i, row in enumerate(mrows, start=2):
        for ci, v in enumerate(row, start=1):
            ws3.cell(i, ci, v)
        if ws3.cell(i, 4).value == "标准名":
            for ci in range(1, 6):
                ws3.cell(i, ci).fill = MAP_STD_FILL
    _style_rows(ws3, 2, 1 + len(mrows), 5, bold_cols=(2,))

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "files": len(ok_files),
        "cols": len(cols) + 1,
        "rows": len(ok_files),
        "mappings": len(mrows),
        "raw_rows": ri - 2,
        "failed": failed,
        "out": str(out_path),
    }

