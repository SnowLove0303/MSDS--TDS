# -*- coding: utf-8 -*-
"""Section 9 全量双轨检索、总表导出与分析统计脚本.

生成:
1. Section9_标准字段总表_全型号.xlsx (253 型号 × 23 标准字段)
2. Section9_原文字段明细总表_全量.xlsx (4925 行原文提取明细)
3. 控制台及 Markdown 详细分析报告数据
"""
import sqlite3
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.msds_db import _SKELETON, listed_section_rows, open_db

DB_PATH = r"F:\正式项目与模块化内容\冠志\MSDS\Word 覆写模块\数据库\正式库\Data Base\msds_standard.db"
OUT_DIR = Path(r"F:\正式项目与模块化内容\冠志\MSDS\Word 覆写模块\数据库\正式库\标准字段数据库Excel")

# S9 23 个标准字段列表及序号
S9_STANDARD_FIELDS = [
    ("9.1", "外观"),
    ("9.2", "嗅觉阈值"),
    ("9.3", "pH值"),
    ("9.4", "离子性"),
    ("9.5", "初沸点"),
    ("9.6", "闪点"),
    ("9.7", "蒸发速率"),
    ("9.8", "可燃性（固态、气态）"),
    ("9.9", "燃烧值"),
    ("9.10", "饱和蒸气压"),
    ("9.11", "相对蒸气密度"),
    ("9.12", "密度"),
    ("9.13", "水溶性"),
    ("9.14", "表面张力"),
    ("9.15", "辛醇/水分配系数对数值"),
    ("9.16", "自燃温度"),
    ("9.17", "引燃温度"),
    ("9.18", "分解温度"),
    ("9.19", "动力粘度"),
    ("9.20", "爆炸特性"),
    ("9.21", "粉尘爆炸级别"),
    ("9.22", "固体含量"),
    ("9.23", "其他信息"),
]

def style_sheet(ws, is_wide=True):
    """Excel 样式美化."""
    ws.freeze_panes = "C2" if is_wide else "A2"
    
    header_fill = PatternFill("solid", fgColor="1B365D")   # 深海蓝
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    alt_fill = PatternFill("solid", fgColor="F7FAFC")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        ws.row_dimensions[1].height = 28

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(row_idx, col_idx)
            c.font = data_font
            c.border = thin_border
            if is_even:
                c.fill = alt_fill
            if col_idx in (1, 3, 4, 7) and not is_wide:
                c.alignment = align_center
            else:
                c.alignment = align_left

    # 自适应列宽
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col[:100]:  # 采样前100行
            val = str(cell.value or '')
            # 中文算 2 字符
            val_len = sum(2 if ord(ch) > 127 else 1 for ch in val)
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)


def export_standard_table(conn):
    """1. 导出 Section 9 标准总表 (253 个型号 × 23 标准字段)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Section9 标准字段总表"

    headers = ["产品型号", "来源文件"] + [f"{seq} {name}" for seq, name in S9_STANDARD_FIELDS]
    ws.append(headers)

    models = conn.execute(
        "SELECT model_id, model, source_file FROM msds_model ORDER BY model"
    ).fetchall()

    for mid, model, source_file in models:
        fname = Path(source_file).name
        rows = listed_section_rows(conn, mid, 9)
        # 获取各标准字段的值
        val_map = {r.label: (r.value or "无数据") for r in rows if r.kind == "field"}
        row_data = [model, fname]
        for seq, name in S9_STANDARD_FIELDS:
            row_data.append(val_map.get(name, "无数据"))
        ws.append(row_data)

    style_sheet(ws, is_wide=True)
    out_path = OUT_DIR / "Section9_标准字段总表_全型号.xlsx"
    wb.save(out_path)
    print(f"✅ 标准总表已生成: {out_path} ({len(models)} 个型号)")
    return out_path


def export_raw_table(conn):
    """2. 导出 Section 9 原文字段明细总表 (4925 行原文记录)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Section9 原文字段明细"

    headers = [
        "明细ID", "产品型号", "来源文件", "原始序号",
        "原始标签", "归属标准字段", "是否标准字段", "原始字段内容"
    ]
    ws.append(headers)

    cur = conn.cursor()
    cur.execute("""
    SELECT id, model, source_file, raw_seq, raw_label, std_name, is_standard, raw_value
    FROM s9_raw_expression
    ORDER BY model, id
    """)
    rows = cur.fetchall()
    for r in rows:
        row_data = [
            r[0], r[1], Path(r[2]).name, r[3], r[4], r[5],
            "是" if r[6] == 1 else "否", r[7]
        ]
        ws.append(row_data)

    style_sheet(ws, is_wide=False)
    out_path = OUT_DIR / "Section9_原文字段明细总表_全量.xlsx"
    wb.save(out_path)
    print(f"✅ 原文字段明细总表已生成: {out_path} ({len(rows)} 行)")
    return out_path


def analyze_unclassified(conn):
    """分析未能被归纳进标准字段的 Section 9 清单."""
    cur = conn.cursor()
    cur.execute("""
    SELECT raw_label, COUNT(DISTINCT model) as m_cnt, GROUP_CONCAT(DISTINCT model) as models, GROUP_CONCAT(raw_value, ' | ') as vals
    FROM s9_raw_expression
    WHERE is_standard = 0
    GROUP BY raw_label
    ORDER BY m_cnt DESC, raw_label
    """)
    return cur.fetchall()


def analyze_standard_mappings(conn):
    """分析 23 个标准字段分别涵盖的不同表述."""
    cur = conn.cursor()
    cur.execute("""
    SELECT std_seq, std_name, raw_label, occurrences_count, models_sample, sample_values
    FROM s9_label_mapping
    ORDER BY CAST(SUBSTR(std_seq, 3) AS INTEGER), std_name, occurrences_count DESC
    """)
    return cur.fetchall()


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    conn = open_db(DB_PATH)
    export_standard_table(conn)
    export_raw_table(conn)

    print("\n" + "="*80)
    print("【未归入 Section 9 标准 23 字段的非标原文标签统计清单】")
    print("="*80)
    unclass = analyze_unclassified(conn)
    print(f"共发现 {len(unclass)} 种非标原文标签：\n")
    for r in unclass:
        lbl, m_cnt, models, vals = r
        m_list = models.split(",")
        m_sample = ", ".join(m_list[:5]) + ("..." if len(m_list) > 5 else "")
        vals_clean = vals.replace("\n", " ")[:100]
        print(f"• 标签: 「{lbl}」 | 涉及型号数: {m_cnt} | 型号示例: {m_sample} | 内容示例: {vals_clean}")
    print("="*80)
