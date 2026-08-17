# -*- coding: utf-8 -*-
"""生成标准库入库总表 (固定结构, 以 PEA-4139 模板为基准).

与 pivot_table (数据驱动动态列) 不同, 本工具生成**固定标准列结构**:
  - 列 = 本批 11 个文件 (10 产品 + PEA-4139 模板) 实际出现的标准字段
    (Schema 归一化, 排除 collapse 分组标签 / 引导段 note)
  - S3 成分表展开为 成分1名称/CAS/含量 ... 成分N名称/CAS/含量
  - 值未检索到 → 填「无数据」(区分确实没有)
  - 序号取该列多数型号使用的众数 (与模板一致)

用法:
  python tools/build_standard_table.py <入库目录> [模板docx] <输出xlsx>
"""
from __future__ import annotations

import re
import sys
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from core.docx_reader import read_msds
from core.s11 import (NOISE_LABELS, S11_MAJOR_FIELDS, S11_MAJOR_SEQ,
                      S11_MAJOR_SET, s11_group_rows, s11_is_subfield,
                      s11_sub_major, s11_value)
from core.schema import _strip_s9_unit, standard_field_of, standard_name

FONT = "Microsoft YaHei"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
TAG_FILL = PatternFill("solid", fgColor="EDEDED")
MODEL_FILL = PatternFill("solid", fgColor="F2F2F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF7E6")
NA_FILL = PatternFill("solid", fgColor="D9D9D9")   # 不适用 → 灰底
ND_FILL = PatternFill("solid", fgColor="FFF2CC")   # 无数据 → 黄底
IRR_FILL = PatternFill("solid", fgColor="F0E6FF")  # S9 无关 → 浅紫底
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 引导段/非字段标签 (note 性质, 不作为列)
_NON_FIELD_STD = {"类似产品的风险评估数据", "请参阅以下数据"}
# 永不输出的字段 (无实际用途, 如 S0 页码)
_EXCLUDE_STD = {(0, "页码")}


def _is_noise(num: int, label: str) -> bool:
    """噪声字段判定: 该 (节, 标准字段) 不产列."""
    if label in NOISE_LABELS:
        return True
    if num == 2 and label == "(总结句)":
        return True              # S2 无总结句 (OS-1330 的 note 是物质名误判)
    return False


def _model_of(result, file_name: str) -> str:
    """型号: 优先取 S1 产品名称, 回退文件名前缀."""
    sec = result.sections.get(1)
    if sec:
        for row in sec.iter_rows():
            if row.kind == "field" and row.label == "产品名称" and row.value.strip():
                return row.value.strip()
    return re.split(r"\s+", Path(file_name).stem)[0]


# ------------------------------------------------------------------
# S11 毒理大类归并
# ------------------------------------------------------------------
# 国标 GB/T 16483-2008 第11部分只有 11.1~11.10 十个毒理大类.
# 扩展结构 MSDS (EC-1801/OS-1330/HPU-7651/PA-3110) 把毒理子项
# (半数致死剂量/物种/方法/染毒途径等) 平铺成独立标签, 若逐个产列
# 会让 S11 膨胀到几十列且标签不明所以. 因此把这些子项**归并进所属
# 国标大类列** (按文档原始顺序归属最近的大类), 检索/总表均收敛到
# 国标 11 大类 + 总结句.
def _s11_groups_of(result) -> dict[str, list[str]]:
    """适配: SectionData → core.s11 归并 (取 S11 节 field 行的 (label, value))."""
    sec = result.sections.get(11)
    if not sec:
        return {}
    rows = ((row.label, row.value) for row in sec.iter_rows()
            if row.kind == "field")
    return s11_group_rows(rows)


def _s11_value(result, label: str) -> str:
    """S11 国标大类列取值: 该大类所有子项值按原文顺序合并 (换行)."""
    return s11_value(_s11_groups_of(result), label)


def collect_columns(results, tpl_idx: int = -1,
                    anchor: dict | None = None
                    ) -> list[tuple[int, list[tuple[str, str]]]]:
    """固定标准列结构: [(节号, [(标准字段名, 序号), ...])].

    anchor: 可选既有列结构 {节号: [(label, seq)...]}. 提供时:
      - 既有列按其 anchor 顺序**原样保留** (不重排, 保入库表列序稳定)
      - 本批文件中 anchor 里没有的新字段 → **追加到该节 anchor 列末尾**
        (按节内平均位置排序, 与"从上到下→从左到右"一致)
      - 不加 anchor 时 = 全量重算 (旧行为).

    规则 (无 anchor 时):
      - 仅收录本批文件实际出现的 field 标准字段 (排除 collapse 分组/引导段)
      - **列顺序 = 检索工具 build_matrix 同构**: 原始标签按其在 reader 节内
        的平均位置排序 (从上到下), 再按 Schema 归一化投影 (同名标准列合并,
        取首次出现位置) → 总表格"从左到右"与检索工具页面"从上到下"一一对应
        (不再按模板顺序+追加; 总结句按其真实位置落列, 不强制排节末)
      - S3 成分表展开为 成分N名称/CAS/含量 (最多 max_comp 个成分), 排字段后
      - 「其他信息」兜底字段 → 强制排该节最末
      - 序号: 以模板为准; 模板无该字段时取本批众数序号, 再无则留空.
    """
    max_comp = max((len(s.components) for r in results
                    for s in r.sections.values() if s.is_component_table), default=0)
    comp_sections = {num for r in results
                     for num, s in r.sections.items() if s.is_component_table}

    # 模板序号表: {节: {标准字段名: 序号}} (序号: 模板优先, 否则众数)
    tpl_seq: dict[int, dict[str, str]] = {}
    if results:
        tpl_r = results[tpl_idx]
        for num, sec in tpl_r.sections.items():
            m: dict[str, str] = {}
            for row in sec.iter_rows():
                if row.kind == "field" and row.label.strip():
                    s = standard_name(num, row.label)
                    if not s or s in _NON_FIELD_STD:
                        continue
                    f = standard_field_of(num, row.label)
                    if f is not None and f.collapse:
                        continue
                    if s and row.seq and s not in m:
                        m[s] = row.seq
            tpl_seq[num] = m

    def _major(seqs: dict[str, int]) -> str:
        return max(seqs, key=lambda s: (seqs[s], -len(s))) if seqs else ""

    # 结构离群文件 (按节: 该节字段结构与主流不同 → 参与位置平均会把列序
    # 拉偏), 排除在位置统计之外 (其字段仍进列集合, 按该文件节内位置追加
    # 节末). 模板永远参与 (权威骨架, 即使模板字段比主流多).
    sec_out = _section_outliers(results, tpl_idx)

    # 收集每节"原始标签"统计 (与检索工具 build_matrix 同构):
    #   raw: {节: {原始标签: {"any": 有值, "seqs": {seq: 次数}, "sum"/"cnt": 平均位置}}}
    #   原始标签 = reader 行 label (跨列通栏 / note 行 → "(总结句)")
    #   排除 collapse 分组标签 / 引导段 / 永不输出字段.
    #   离群文件只登记字段 (any/seqs), 不参与位置统计 (sum/cnt 为 0 → 追加节末).
    raw: dict[int, dict[str, dict]] = {}
    for ri, r in enumerate(results):
        for num, sec in r.sections.items():
            if num == 9:
                continue                     # S9 走开放字段逻辑 (_s9_columns)
            is_out = ri in sec_out.get(num, set())
            m = raw.setdefault(num, {})
            # S11 毒理: 列 = 国标大类 (s11_group_rows 归并, 与检索工具
            # build_matrix 完全同构) + 总结句. 大类位置固定国标顺序 →
            # 总表"从左到右" = 检索页面"从上到下". 子项 (LD50/物种/方法等)
            # 不产独立列 (OS-1330 等扩展结构逐子项平铺会让 S11 膨胀到
            # 几十列且标签不明所以).
            s11_groups = _s11_groups_of(r) if num == 11 else None
            for pos, row in enumerate(sec.iter_rows()):
                if row.kind == "section":
                    continue
                if s11_groups is not None and row.kind == "field":
                    if row.label.strip():
                        continue             # S11 子项/精确大类行 → 归并列统一产出
                    key = "(总结句)"          # label 空的 field → 通栏总结句
                    val = row.value
                elif row.kind == "field":
                    if not row.label.strip():
                        key = "(总结句)"      # 跨列/通栏总结句 → note 统一处理
                    else:
                        key = row.label.strip()
                    val = row.value
                elif row.kind == "note":
                    key = "(总结句)"
                    val = row.value
                else:
                    continue
                if key != "(总结句)":
                    f = standard_field_of(num, key)
                    if f is not None and f.collapse:
                        continue             # 分组标签不产列
                    s = standard_name(num, key)
                    if s in _NON_FIELD_STD or (num, s) in _EXCLUDE_STD or not s:
                        continue             # 引导段 / 永不输出 / 未识别
                    if _is_noise(num, s):
                        continue             # 噪声标签 (组分/组分名/接触途径)
                else:
                    s = "(总结句)"
                    if _is_noise(num, s):
                        continue             # 该节无总结句 (如 S2 note 是物质名)
                d = m.setdefault(key, {"any": False, "seqs": {}, "sum": 0, "cnt": 0,
                                       "osum": 0, "ocnt": 0})
                d["any"] |= bool(val.strip())
                if row.seq:
                    d["seqs"][row.seq] = d["seqs"].get(row.seq, 0) + 1
                if is_out:
                    d["osum"] += pos            # 离群文件位置 (节末内部排序)
                    d["ocnt"] += 1
                else:
                    d["sum"] += pos             # 主流文件位置
                    d["cnt"] += 1
            # S11 国标大类列: 按国标顺序产出 (与 build_matrix 一致), 序号 11.N
            # (模板无 11.8 等列时取众数 → 仍是国标序号), 位置在总结句之后按
            # 国标顺序递增 (基准 = 该节总行数, 保证大类列排所有真实行之后).
            if s11_groups:
                base = sum(1 for _ in sec.iter_rows())
                for i, major in enumerate(S11_MAJOR_FIELDS):
                    vals = s11_groups.get(major)
                    if not vals:
                        continue
                    d = m.setdefault(major, {"any": False, "seqs": {}, "sum": 0,
                                             "cnt": 0, "osum": 0, "ocnt": 0})
                    d["any"] = True
                    d["seqs"][S11_MAJOR_SEQ[major]] = \
                        d["seqs"].get(S11_MAJOR_SEQ[major], 0) + 1
                    if is_out:
                        d["osum"] += base + i
                        d["ocnt"] += 1
                    else:
                        d["sum"] += base + i
                        d["cnt"] += 1

    def _pos(d: dict) -> float:
        """列排序位置: 主流平均位置; 仅离群文件有的字段 → 追加节末区间
        (内部按离群文件节内位置排序, 与检索工具"从上到下"一致)."""
        if d["cnt"]:
            return d["sum"] / d["cnt"]
        if d["ocnt"]:
            return 10**9 - 1000 + d["osum"] / d["ocnt"]
        return 10**9

    out: list[tuple[int, list[tuple[str, str]]]] = []
    # anchor: 既有列集合 {num: set(label)} 用于判断"该字段是否已是既有列"
    anchor_labels: dict[int, set[str]] = {}
    if anchor:
        for num, acols in anchor.items():
            anchor_labels[num] = {l for l, _ in acols}
    for num in sorted(raw):
        m = raw[num]
        # 原始标签按节内平均位置升序 → 检索工具"从上到下" → 总表"从左到右"
        order = sorted(m, key=lambda k: (_pos(m[k]), k))
        # Schema 归一化投影: 同名标准列合并 (保留首次出现位置)
        seen: set[str] = set()
        cols: list[tuple[str, str]] = []
        for key in order:
            s = key if key == "(总结句)" else standard_name(num, key)
            if s in _NON_FIELD_STD or (num, s) in _EXCLUDE_STD or not s:
                continue
            if s in seen:
                continue
            seen.add(s)
            # 合并该标准名所有原始标签: any(有值) + seqs(序号众数)
            any_v = False
            seqs: dict[str, int] = {}
            for k, d in m.items():
                ks = k if k == "(总结句)" else standard_name(num, k)
                if ks == s:
                    any_v |= d["any"]
                    for sq, c in d["seqs"].items():
                        seqs[sq] = seqs.get(sq, 0) + c
            if s != "(总结句)" and not any_v and s not in tpl_seq.get(num, {}):
                continue                     # 全空列剔除 (模板标准字段保留)
            # 序号: 模板优先, 否则众数
            seq = tpl_seq.get(num, {}).get(s, "")
            if not seq:
                seq = _major(seqs)
            cols.append((s, seq))
        # S3 成分表展开: 成分N名称/CAS/含量 (排字段列之后, 按成分索引)
        if num in comp_sections and max_comp:
            for i in range(max_comp):
                for suffix in ("名称", "CAS", "含量"):
                    cols.append((f"成分{i+1}{suffix}", ""))
        # ---- anchor 模式: 既有列原序在前, 新样本独有字段追加节末 ----
        if anchor is not None and num in anchor:
            anchor_cur = anchor[num]
            # 既有列名集 (含成分列), 判断字段是否已存在
            anames = [l for l, _ in anchor_cur]
            aset = set(anames)
            # 既有列原序保留
            merged = list(anchor_cur)
            mseen = set(aset)
            # 新样本独有字段 (非既有的) 追加
            for l, sq in cols:
                if l not in mseen:
                    merged.append((l, sq))
                    mseen.add(l)
            cols = merged
        out.append((num, cols))

    # S9 放开字段: 名称(剥单位后)不一致即拆列, 保留特殊表达.
    #   列顺序 = 检索工具同构 (原始标签平均位置 → 归一化投影);
    #   「其他信息」强制排最末.
    #   列头去序号 (seq=''), 源序号信息写入辅助 sheet "S9序号映射",
    #   供推断引擎按源序号跟随覆写到 MSDS 模板.
    s9_cols = _s9_columns(results, tpl_idx)
    if anchor is not None and 9 in anchor:
        # anchor 模式: 既有 S9 列原序, 新样本独有 S9 字段追加节末
        anchor9 = [l for l, _ in anchor[9]]
        aset9 = set(anchor9)
        merged9 = list(anchor[9])
        for l, sq in s9_cols:
            if l not in aset9:
                merged9.append((l, sq))
                aset9.add(l)
        s9_cols = merged9
    if s9_cols:
        out.append((9, s9_cols))
    out.sort(key=lambda x: x[0])
    return out, max_comp


def _section_outliers(results, tpl_idx: int = -1) -> dict[int, set[int]]:
    """按节结构离群文件: {节号: set[文件索引]}, 模板永不排除.

    结构离群 = 该节字段结构与主流文件显著不同, 参与位置平均会把列序拉偏
    → 排序统计排除 (其字段仍进列集合, 按该文件节内位置追加节末).
    两维判定 (任一命中即该节离群, 与"行数"无关地捕捉字段结构差异):
      A. 行数离群: 该节 field/note 行数 > 中位行数×2
         (BL-8085 S11 49 行 vs 主流 7 行).
      B. 字段重叠率低: 主流共同字段集 = 排除行数离群后出现 ≥ 半数文件的
         原始标签; 某文件该节字段中属于共同字段集的比例 < 0.5 →
         字段结构与主流不同 (2-苯氧基乙醇 S9 旧结构: 14 字段仅 4 个与
         主流重叠, 而主流 11 份 S9 结构一致).
    """
    tpl_idx = tpl_idx if tpl_idx >= 0 else len(results) + tpl_idx
    # 阶段1: 行数统计 + 规则A
    row_out: dict[int, set[int]] = {}
    for num in range(17):
        counts = [sum(1 for row in r.sections.get(num).iter_rows()
                      if row.kind in ("field", "note"))
                  if r.sections.get(num) else 0
                  for r in results]
        pos = sorted(c for c in counts if c > 0)
        if not pos:
            continue
        med = pos[len(pos) // 2]
        if med <= 0:
            continue
        s = {i for i, c in enumerate(counts)
             if i != tpl_idx and c > med * 2}
        if s:
            row_out[num] = s
    # 阶段2: 主流共同字段集 (原始标签; 排除行数离群文件)
    common: dict[int, set[str]] = {}
    for num in range(17):
        cnt: dict[str, int] = {}
        n_main = 0
        for ri, r in enumerate(results):
            sec = r.sections.get(num)
            if not sec or ri in row_out.get(num, set()):
                continue
            n_main += 1
            for row in sec.iter_rows():
                if row.kind == "field" and row.label.strip():
                    raw = row.label.strip()
                    cnt[raw] = cnt.get(raw, 0) + 1
        thr = max(3, n_main // 2) if n_main else 0
        if thr:
            common[num] = {raw for raw, c in cnt.items() if c >= thr}
    # 阶段3: 规则B (字段重叠率)
    out: dict[int, set[int]] = {num: set(s) for num, s in row_out.items()}
    for num in range(17):
        common_set = common.get(num)
        if not common_set:
            continue
        cur = out.setdefault(num, set())
        for ri, r in enumerate(results):
            if ri == tpl_idx or ri in cur:
                continue
            sec = r.sections.get(num)
            if not sec:
                continue
            fs = {row.label.strip()
                  for row in sec.iter_rows()
                  if row.kind == "field" and row.label.strip()}
            if not fs:
                continue
            if len(fs & common_set) / len(fs) < 0.5:
                cur.add(ri)
    return {num: s for num, s in out.items() if s}


# S9 强制独立列 (语义/概念不同, 不因 Schema 别名归并):
#  自燃温度/引燃温度/着火点 是不同概念, 拆开各自成列.
_S9_FORCE_SPLIT = {"自燃温度", "引燃温度", "着火点", "自燃点"}

# pH 大小写变体 → 统一 'pH值' (pH/PH/ph/Ph + 可选"值"): 大小写差异归并
_PH_RE = re.compile(r"^pH\s*值?$", re.I)

# GB/T 16483-2008 第9部分条目归并 (标准原文「密度/相对密度」为同一条目):
#  密度 与 相对密度 合并为同一列, 各自单位/数值写法原样保留在值里.
_S9_GB_MERGE = {
    "密度": "密度/相对密度",
    "相对密度": "密度/相对密度",
}


def _s9_col_name(raw: str) -> str:
    """S9 原始标签 → 标准列名 (归一化).

    规则 (对齐用户意图 + GB/T 16483-2008):
      - 通用条件括号剥除 + 同义归并 (standard_name):
          pH值（1%水溶液）/PH值 → pH值    蒸汽压（20℃，Kpa）→饱和蒸气压
          蒸发速率/蒸发速度 → 蒸发速率     水中溶解度 → 水溶性
          动力粘度/粘度/25℃ → 粘度        比重 → 相对密度
      - 国标条目合并:                        密度/相对密度 → 密度/相对密度 (GB 原文)
      - 特殊限定括号保留 (独立列):          蒸发速度（醋酸丁酯-1）原样
      - 自燃/引燃/着火点 强制拆分:          引燃温度 → 引燃温度 (不并自燃温度)
    """
    from core.schema import _clean_label, standard_name
    lbl = _clean_label(raw)
    if not lbl:
        return ""
    if lbl in _S9_FORCE_SPLIT:
        return lbl
    m = _PH_RE.match(lbl)
    if m:
        lbl = "pH值"                  # 大小写差异归一化 (PH值/ph值/Ph值/PH → pH值)
    name = standard_name(9, lbl)
    return _S9_GB_MERGE.get(name, name)


def _s9_columns(results, tpl_idx: int = -1) -> list[tuple[str, str]]:
    """S9 列: [(标准列名, seq='')] (归一化后名称不一致即拆列).

    列顺序 = 模板 (国标 GB/T 16483-2008 第9部分, 9.1~9.23) 权威序为基准:
      - 模板 S9 字段按模板顺序排列 (国标条目序, 入库目标模板的阅读顺序)
      - 模板没有但主流文件出现的字段 (有效成分/百分比挥发性/着火点等)
        按其"后一个模板字段"插槽插入 (贴近源文档相对位置, 不拆散国标核心序)
      - 结构离群文件 (如 2-苯氧基乙醇 S9 旧结构) 特有字段 → 追加节末
      - 「其他信息」强制排最末. 列头去序号 → seq 恒为 ''.
    """
    from collections import Counter
    tpl_idx = tpl_idx if tpl_idx >= 0 else len(results) + tpl_idx
    s9_out = _section_outliers(results, tpl_idx).get(9, set())
    tpl_r = results[tpl_idx]
    # 模板基准 (国标序, 不含「其他信息」)
    base: list[str] = []
    sec = tpl_r.sections.get(9)
    if sec:
        for row in sec.iter_rows():
            if row.kind != "field" or not row.label.strip():
                continue
            n = _s9_col_name(row.label.strip())
            if n and n != "其他信息" and n not in base:
                base.append(n)
    base_set = set(base)
    # 非离群文件 S9 标准列序
    file_cols: list[list[str]] = []
    for ri, r in enumerate(results):
        if ri in s9_out or r.sections.get(9) is None:
            continue
        seq: list[str] = []
        for row in r.sections.get(9).iter_rows():
            if row.kind != "field" or not row.label.strip():
                continue
            n = _s9_col_name(row.label.strip())
            if n and n != "其他信息" and n not in seq:
                seq.append(n)
        if seq:
            file_cols.append(seq)
    # 非模板列 → 插槽 (插入"后一个模板字段"之前; 末尾无后续 → len(base))
    extras = set()
    for seq in file_cols:
        extras |= set(seq) - base_set
    freq = Counter(n for seq in file_cols for n in set(seq))
    slot_map: dict[str, int] = {}
    for n in extras:
        slots: list[int] = []
        for seq in file_cols:
            if n not in seq:
                continue
            i = seq.index(n)
            nxt_b = next((j for j in range(i + 1, len(seq)) if seq[j] in base_set), None)
            if nxt_b is None:
                slots.append(len(base))
            else:
                slots.append(base.index(seq[nxt_b]))
        slot_map[n] = Counter(slots).most_common(1)[0][0]
    # 组装: 模板基准 + 非模板字段按插槽从后往前插入
    cols = list(base)
    groups: dict[int, list[str]] = {}
    for n in extras:
        groups.setdefault(slot_map[n], []).append(n)
    for slot in sorted(groups, reverse=True):
        for n in sorted(groups[slot], key=lambda x: (-freq[x], x)):
            cols.insert(slot, n)
    # 离群文件特有字段: 追加节末 (按该文件节内位置排序, 归一化去重)
    seen = set(cols)
    out_sum: dict[str, int] = {}
    out_cnt: dict[str, int] = {}
    for ri, r in enumerate(results):
        if ri not in s9_out or r.sections.get(9) is None:
            continue
        for pos, row in enumerate(r.sections.get(9).iter_rows()):
            if row.kind != "field" or not row.label.strip():
                continue
            n = _s9_col_name(row.label.strip())
            if not n or n in seen:
                continue
            out_sum[n] = out_sum.get(n, 0) + pos
            out_cnt[n] = out_cnt.get(n, 0) + 1
    for n in sorted(out_cnt, key=lambda n: (out_sum[n] / out_cnt[n], n)):
        if n not in seen:
            cols.append(n)
            seen.add(n)
    cols = [n for n in cols if n != "其他信息"]   # 其他信息强制最末
    cols.append("其他信息")
    return [(n, "") for n in cols]


def _s9_field_meta(result) -> list[tuple]:
    """某型号 S9 字段元数据 (按源文档检索顺序).

    返回 [(标准列名, 源原始标签, 源seq, 状态, 值)...]
    状态: 有值 / 无数据 / 不适用 (不含"无关" — 无关 = 该型号没有此字段).
    """
    sec = result.sections.get(9)
    if not sec:
        return []
    out = []
    for row in sec.iter_rows():
        if row.kind != "field" or not row.label.strip():
            continue
        raw = row.label.strip()
        name = _s9_col_name(raw)
        v = (row.value or "").strip()
        if not v or v in _ND_VALUES:
            st, disp = "无数据", "无数据"
        elif v in _NA_VALUES:
            st, disp = "不适用", "不适用"
        else:
            st, disp = "有值", v
        out.append((name, raw, row.seq or "", st, disp))
    return out


def _s9_cell(result, label: str) -> str:
    """S9 单元格值: 四态 (有值显示实际值 / 无数据 / 不适用 / 无关).

    "无关" = 该型号原始 S9 里没有此字段 (与"无数据"区分).
    同字段多行时按 有值 > 不适用 > 无数据 优先取最高态.
    """
    sec = result.sections.get(9)
    if not sec:
        return "无关"
    best = ""
    best_pri = -1
    for row in sec.iter_rows():
        if row.kind != "field" or not row.label.strip():
            continue
        name = _s9_col_name(row.label.strip())
        if name != label:
            continue
        v = (row.value or "").strip()
        if not v or v in _ND_VALUES:
            pri, disp = 0, "无数据"
        elif v in _NA_VALUES:
            pri, disp = 1, "不适用"
        else:
            pri, disp = 2, v
        if pri > best_pri:
            best_pri, best = pri, disp
    return best if best_pri >= 0 else "无关"


def _value_of(result, num: int, label: str, max_comp: int) -> str:
    """取一个标准字段的值 (field 匹配 + 成分展开 + 总结句)."""
    if num == 9:
        return _s9_cell(result, label)
    sec = result.sections.get(num)
    if not sec:
        return ""
    # 成分列
    m = re.match(r"成分(\d+)(名称|CAS|含量)", label)
    if m:
        idx = int(m.group(1)) - 1
        if idx >= len(sec.components):
            return ""                     # 槽位不存在 → 外层填「无数据」
        c = sec.components[idx]
        val = {"名称": c.name, "CAS": c.cas, "含量": c.conc}[m.group(2)]
        if not val:
            return "——"                   # 成分存在但字段缺失 → 参考原文档输出 ——
        return val
    # 总结句
    if label == "(总结句)":
        vals = [row.value for row in sec.iter_rows()
                if row.kind == "note" or (row.kind == "field" and row.span
                                          and not row.label.strip())]
        # 过滤 "无数据资料" 等占位 note (子组标题下的占位行, 非总结句),
        # 避免 S12 总结句尾部堆积多个 "无数据资料".
        return "\n".join(v for v in vals
                         if v.strip() and v.strip() not in _ND_VALUES)
    # S11: 国标大类列 → 归并该大类所有子项值
    if num == 11 and label in S11_MAJOR_SET:
        return _s11_value(result, label)
    # 标准字段: 匹配任意同义写法
    #   multi 字段 (如 S2 防范说明 = 预防措施/事故响应/安全储存/废弃处置 的
    #   P 代码) → 汇总**所有**同义行值, 保证 P/H 开头条目 (P280/P305+P351+
    #   P338/P405/P501/Hxxx) 全部落入该列, 而非只取首个子类 (预防措施).
    #   非 multi → 取首个非空 (兼容旧行为).
    f = standard_field_of(num, label)
    vals: list[str] = []
    for row in sec.iter_rows():
        if not row.label.strip():
            continue
        if row.kind not in ("field", "subtable"):
            continue
        if standard_name(num, row.label) != label:
            continue
        if row.kind == "subtable":
            # 内嵌子表 (S8.2 生物限值等): 表头+数据行 → 多行文本.
            # 插到最前, 保证 "MSDS 有实际数据" 时优先展示, 而非标题行的
            # "无数据/无资料" 占位 (用户任务: MSDS 有信息检索结果就应有).
            st_rows = []
            for sr in row.sub_rows:
                pairs = [f"{h}: {v}" for h, v in zip(row.sub_header, sr)
                         if str(v).strip()]
                st_rows.append(" | ".join(pairs) if pairs else " | ".join(sr))
            if st_rows:
                vals.insert(0, "\n".join(st_rows))
            continue
        if row.value.strip():
            vals.append(row.value.strip())
    if not vals:
        return ""
    if f is not None and f.multi:
        return "\n".join(vals)
    return vals[0]


def _load_anchor_columns(xlsx_path) -> dict[int, list[tuple[str, str]]]:
    """从既有总表 xlsx 读取列结构 (第1行节标题, 第2行字段名).

    字段名可能内嵌序号前缀 ("11.1 急性毒性") → 拆为 (label, seq).
    返回 {节号: [(label, seq)...]}, 节号来自第1行标题前导数字.
    """
    import re
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["入库总表"]
    sec_order: list[int] = []
    cur_num: int | None = None
    for c in range(1, ws.max_column + 1):
        t = ws.cell(1, c).value
        if t:
            m = re.match(r"^(\d+)", str(t).strip())
            cur_num = int(m.group(1)) if m else (cur_num or 0)
        if cur_num is not None and cur_num not in sec_order:
            sec_order.append(cur_num)
    anchor: dict[int, list[tuple[str, str]]] = {}
    cur_num = None
    for c in range(2, ws.max_column + 1):
        t = ws.cell(1, c).value
        if t:
            m = re.match(r"^(\d+)", str(t).strip())
            cur_num = int(m.group(1)) if m else (cur_num or 0)
        f = ws.cell(2, c).value
        if f is None:
            continue
        # 拆 序号前缀: "11.1 急性毒性" → ("急性毒性", "11.1")
        seq = ""
        label = str(f).strip()
        mm = re.match(r"^(\d+\.\d+)\s+(.+)$", label)
        if mm:
            seq, label = mm.group(1), mm.group(2)
        # S11: 过滤旧表毒理子项列 (LD50/物种/方法等), 只保留国标大类+总结句
        # (v15 及更早把扩展结构子项平铺成独立列, 本次改为归并进大类列)
        if (cur_num or 0) == 11 and label != "(总结句)" and s11_is_subfield(label):
            continue
        anchor.setdefault(cur_num or 0, []).append((label, seq))
    # 按既有 sheet 节序返回
    ordered = {n: anchor[n] for n in sec_order if n in anchor}
    for n in sorted(anchor):
        if n not in ordered:
            ordered[n] = anchor[n]
    return ordered


def build_standard_table(in_dir, out_path, template=None,
                         exclude=(), base_xlsx=None) -> dict:
    """生成固定结构标准库入库总表.

    exclude: 排除文件名关键字元组 (如 ("BEK-750",)), 支持子串匹配.
    base_xlsx: 可选既有总表 xlsx 路径. 提供时以其列结构为锚定基准,
      新文件独有字段追加节末 (基于既有表补充, 不重排既有列).
    """
    in_dir = Path(in_dir)
    files = sorted(p for p in in_dir.glob("*.docx")
                   if not p.name.startswith("~$")
                   and not any(k in p.name for k in exclude))
    results = [read_msds(p) for p in files]
    ok_files = list(files)
    if template:
        tp = Path(template)
        results.append(read_msds(tp))
        ok_files.append(tp)

    anchor = _load_anchor_columns(base_xlsx) if base_xlsx else None
    columns, max_comp = collect_columns(results, anchor=anchor)
    wb = Workbook()
    ws = wb.active
    ws.title = "入库总表"

    # 表头两行
    col = 2
    spans = []
    for num, cols in columns:
        start = col
        for label, seq in cols:
            head = f"{seq} {label}".strip() if seq else label
            c = ws.cell(2, col, head)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = TAG_FILL
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = BORDER
            col += 1
        spans.append((start, col - 1, num))

    # 第一行节标题
    for start, end, num in spans:
        sec = results[0].sections.get(num)
        title = sec.full_title if sec else f"第{num}节"
        if num == 0:
            title = "0 页眉/页脚"
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = ws.cell(1, start, title)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        for cc in range(start, end + 1):
            ws.cell(1, cc).border = BORDER
        ws.row_dimensions[1].height = 28
    ws.cell(1, 1, ""); ws.cell(2, 1, "")
    ws.cell(1, 1).fill = HEAD_FILL
    ws.cell(2, 1).fill = TAG_FILL

    # 数据行
    for ri, (p, r) in enumerate(zip(ok_files, results), start=3):
        model = _model_of(r, p.name)
        mc = ws.cell(ri, 1, model)
        mc.font = Font(name=FONT, size=10, bold=True)
        mc.fill = MODEL_FILL
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = BORDER
        col = 2
        for num, cols in columns:
            for label, seq in cols:
                v = _value_of(r, num, label, max_comp)
                if not v:
                    v = "无数据"
                c = ws.cell(ri, col, v)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = BORDER
                if label == "(总结句)" and v != "无数据":
                    c.fill = NOTE_FILL
                # S9 开放字段四态标注: 无关=浅紫, 不适用=灰底, 无数据=黄底, 有值=白底
                elif num == 9:
                    if v == "无关":
                        c.fill = IRR_FILL
                    else:
                        st = _state_of(v)
                        if st == "不适用":
                            c.fill = NA_FILL
                        elif st == "无数据":
                            c.fill = ND_FILL
                col += 1
        ws.row_dimensions[ri].height = 60

    ws.column_dimensions["A"].width = 20
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "B3"

    # ---- 辅助 sheet: S9 序号映射 (供推断引擎按源序号跟随覆写 MSDS 模板) ----
    ws9 = wb.create_sheet("S9序号映射")
    hdr = ["型号", "字段(剥单位)", "源原始标签", "源seq", "状态", "值"]
    for i, h in enumerate(hdr, 1):
        c = ws9.cell(1, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for i, w in enumerate([20, 22, 30, 8, 8, 30], 1):
        ws9.column_dimensions[get_column_letter(i)].width = w
    ws9.freeze_panes = "A2"
    ri = 2
    for p, r in zip(ok_files, results):
        model = _model_of(r, p.name)
        for name, raw, seq, st, disp in _s9_field_meta(r):
            vals = [model, name, raw, seq, st, disp]
            for ci, v in enumerate(vals, 1):
                cc = ws9.cell(ri, ci, v)
                cc.font = Font(name=FONT, size=9)
                cc.alignment = Alignment(vertical="top", wrap_text=True)
                cc.border = BORDER
            ri += 1

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    total_cols = sum(len(c) for _, c in columns)
    return {"files": len(ok_files), "cols": total_cols + 1,
            "sections": len(columns), "out": str(out_path)}


# 值状态判定: 用于「不适用」分析
_NA_VALUES = {"不适用", "不适用。", "不适用/无数据", "NA", "n/a", "N/A"}
_ND_VALUES = {"无数据", "无数据。", "无数据资料", "无数据资料。", "无资料"}


def _state_of(v: str) -> str:
    """理化特性值状态: 有值 / 不适用(物理上不适用) / 无数据(未检索到)."""
    v = (v or "").strip()
    if not v:
        return "无数据"
    if v in _NA_VALUES or v.lower() in _NA_VALUES:
        return "不适用"
    if v in _ND_VALUES or v in _ND_VALUES:
        return "无数据"
    return "有值"


def analyze_s9_states(results) -> dict:
    """S9 不适用分析: {型号: {标准字段: 状态}}.

    状态: 有值 / 不适用 / 无数据 (用于判断哪些型号哪些理化特性不适用).
    """
    out = {}
    for r in results:
        sec = r.sections.get(9)
        if not sec:
            continue
        model = r.file_name.split(" ")[0].replace(".docx", "")
        states = {}
        for row in sec.iter_rows():
            if row.kind != "field" or not row.label.strip():
                continue
            s = standard_name(9, row.label)
            if s in _NON_FIELD_STD:
                continue
            if s not in states:
                states[s] = _state_of(row.value)
        out[model] = states
    return out


def query_s9(xlsx_path, model: str) -> list[tuple]:
    """从标准库入库总表检索某型号 S9 字段序列 (按列序从左到右, 从"外观"开始).

    返回 [(字段名(剥单位), 源seq, 状态, 值, 源原始标签)...],
    只含 有值 / 无数据 / 不适用 (不含"无关" = 该型号没有的字段).
    供推断引擎/后续覆写按"源序号"定位跟随覆写 MSDS 模板.

    列顺序以主表 S9 列头 (去序号字段名) 从左到右为准.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if "S9序号映射" not in wb.sheetnames:
        raise ValueError("表格缺少 S9序号映射 sheet (需由新版 build_standard_table 生成)")
    ws9 = wb["S9序号映射"]
    meta = {}
    for r in range(2, ws9.max_row + 1):
        if ws9.cell(r, 1).value == model:
            name = ws9.cell(r, 2).value
            meta[name] = (ws9.cell(r, 4).value or "",    # 源seq
                          ws9.cell(r, 5).value,           # 状态
                          ws9.cell(r, 6).value,           # 值
                          ws9.cell(r, 3).value)           # 源原始标签
    ws = wb["入库总表"]
    ordered = [(c, h) for c in range(2, ws.max_column + 1)
               if (h := ws.cell(2, c).value) in meta]
    return [(h, *meta[h]) for _, h in ordered]


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    in_dir = sys.argv[1]
    out = sys.argv[2]
    tpl = sys.argv[3] if len(sys.argv) > 3 else None
    # 第4个参数起为排除文件关键字 (如 BEK-750); 前缀 @base.xlsx 指定既有总表
    excl = []
    base = None
    for a in sys.argv[4:]:
        if a.startswith("@") and a.endswith(".xlsx"):
            base = a[1:]
        else:
            excl.append(a)
    print(build_standard_table(in_dir, out, tpl, exclude=tuple(excl), base_xlsx=base))

