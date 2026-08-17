# -*- coding: utf-8 -*-
"""标准范式批量检索输出 (build_std_matrix).

以 标准字段结构.json (从标准字段库 Excel 提取的父子级→标签规范) 为
**唯一输出范式**, 复用 core/std_output.py 核心 (与 GUI「标准范式输出」同源).

规则:
  - 所有检索结果必须按标准库的 (父级 → 标签) 列结构输出
  - 标准库有、文件未检出 → 自动标注「无数据」(黄色底)
  - 同一父级标签下多个结果 → 自动扩列 (其他危险(2)(3)...)
  - 成分槽: 标准库预留 5 槽, 文件成分超过 5 个 → 动态追加列
  - 锚点列 (父级占位) 留空

用法:
  python build_std_matrix.py <标准库xlsx> <docx或目录...> -o 输出.xlsx
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.docx_reader import read_msds
from core.std_output import build_std_matrix as _build

FONT = '微软雅黑'
HEAD_FILL = PatternFill('solid', fgColor='1F4E79')
SEC_FILL = PatternFill('solid', fgColor='2E3D5C')
TAG_FILL = PatternFill('solid', fgColor='EDEDED')
MODEL_FILL = PatternFill('solid', fgColor='F2F2F2')
ND_FILL = PatternFill('solid', fgColor='FFF2CC')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def render_xlsx(matrix: dict, out_path: Path, failed: list[str]) -> int:
    """标准矩阵 → xlsx. 返回列数(含型号)."""
    cols = matrix['columns']
    rows = matrix['rows']
    groups = matrix['groups']

    wb = Workbook()
    ws = wb.active
    ws.title = '标准范式检索输出'

    # 行1 父级分组 + 行2 标签
    for idx, (sec, parent, label, typ, anchor, mi) in enumerate(cols):
        hdr = label if mi == 1 else f'{label}({mi})'
        c = ws.cell(2, idx + 2, hdr)
        c.font = Font(name=FONT, size=9, bold=True)
        c.fill = TAG_FILL
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[ws.cell(2, idx + 2).column_letter].width = 13

    for s, e, t in groups:
        if e > s:
            ws.merge_cells(start_row=1, start_column=s + 2,
                           end_row=1, end_column=e + 2)
        c = ws.cell(1, s + 2, t)
        c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        c.fill = SEC_FILL if str(t).startswith('Section') else HEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        for cc in range(s + 2, e + 3):
            ws.cell(1, cc).border = BORDER

    ws.cell(1, 1, '型号')
    ws.cell(1, 1).fill = SEC_FILL
    ws.cell(1, 1).font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    ws.cell(2, 1, '')
    ws.cell(2, 1).fill = TAG_FILL
    ws.column_dimensions['A'].width = 24

    for ri, (model, cells) in enumerate(rows, start=3):
        mc = ws.cell(ri, 1, model)
        mc.font = Font(name=FONT, size=9, bold=True)
        mc.fill = MODEL_FILL
        mc.alignment = Alignment(horizontal='center', vertical='center')
        mc.border = BORDER
        for ci, v in enumerate(cells):
            c = ws.cell(ri, ci + 2, v if v is not None else '无数据')
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
            if v is None:
                c.fill = ND_FILL
        ws.row_dimensions[ri].height = 40

    ws.freeze_panes = 'C3'
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(cols) + 1


def main():
    args = [a for a in sys.argv[1:]]
    if len(args) < 3 or '-o' not in args:
        print(__doc__)
        print('用法: python build_std_matrix.py <标准库xlsx> <docx或目录...> -o 输出.xlsx')
        sys.exit(1)
    oi = args.index('-o')
    _std_xlsx = args[0]
    out_path = Path(args[oi + 1])
    inputs = [a for a in args[1:oi]]

    files = []
    for raw in inputs:
        p = Path(raw)
        if p.is_dir():
            files.extend(sorted(f for f in p.rglob('*.docx')
                                if not f.name.startswith('~$')))
        elif p.exists() and p.suffix.lower() == '.docx':
            files.append(p)
    files = sorted(set(files), key=lambda x: str(x).lower())
    if not files:
        print('❌ 未找到 docx 文件')
        sys.exit(2)

    results = []
    failed = []
    for f in files:
        try:
            results.append(read_msds(f))
        except Exception as exc:
            failed.append(f'{f.name}: {exc}')
    if not results:
        print('❌ 全部读取失败')
        sys.exit(2)

    matrix = _build(results)
    n_cols = render_xlsx(matrix, out_path, failed)
    print(f'✅ 已生成: {out_path}')
    print(f'   文件 {len(results)} | 列 {n_cols} | 分组 {len(matrix["groups"])}')
    if failed:
        print('   跳过失败:')
        for x in failed:
            print(f'     - {x}')


if __name__ == '__main__':
    main()
