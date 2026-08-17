# -*- coding: utf-8 -*-
"""MSDS 批量读取入口 — 快速批量读取指定 MSDS 文件/目录/通配符.

复用 结构读取/core 的成熟解析 (read_msds + 三级树/分层提取),
提供统一批处理 CLI: 汇总统计 / 节过滤 / 关键词检索 / JSON·TSV·文本导出 /
JSON 报告 (含读取状态、异常、缺失节) / token 完整性校验.

用法 (直接运行, 不依赖当前工作目录; 使用绝对路径定位 core 模块):
  python batch_read.py <docx或目录或通配符...> [选项]

  位置参数: 文件路径 / 目录 (递归扫描 *.docx) / 通配符 (如 "F:\\...\\*.docx")

  批量化检索能力 (跨文档关键词检索):
    --query 词              关键词检索 (全部文件统一; 多词用空格分隔)
    --scope label|value|all|section   检索范围 (默认 all;
                            label=字段标签, value=内容, all=全字段,
                            section=节号精确匹配如 --query 9 --scope section)
    --any                  多关键词 OR 匹配 (默认 AND: 全部词都须命中)
    --name-filter 子串1,子串2   按文件名子串过滤待处理文件 (逗号分隔 OR,
                            大小写不敏感; 实现"只检索某批次型号如 BL,OS")
    --hits                 文件命中清单模式: 只输出每个命中文件的
                            "命中条目数 + 文件名", 0 命中文件不列出,
                            末尾合计命中文件数/命中条目数 (批量审计"谁命中")
    --show-empty           检索时保留 0 命中的文件块 (默认过滤, 减噪音)

  选项:
    --sections 1,3,9        仅提取指定节 (逗号分隔)
    --json | --tsv | --matrix   输出格式 (默认 text 三级树)
                            --matrix 宽表矩阵: 行=一个文档, 列=字段;
                            列顺序=(节号, reader 节内上下顺序);
                            列头=Section{n} {序号范围} {标题}, 同标签
                            不同序号自动合并且显示范围 (9.4~9.5 初沸点);
                            无标题用 Section{n}特殊字段; 成分拆 名称/CAS/含量;
                            与 reader 三列表显示一一对应
    --states               仅 --matrix: 每字段列后加"值状态"三态列
                            (有值 / 无数据=原文有字段但空 / 无此字段=原文没有)
    --comp-cols            成分分列输出 (默认输出模式): S3 成分从"每条成分
                            一行整格拼接"改为"每文档一行, 成分1|CAS1|含量1|
                            成分2|CAS2|含量2 交替平铺"; 文本/TSV/JSON 均适用;
                            方便数据库入库. GUI 显示与核心解析不变
    --out 文件              导出到文件 (默认 stdout; TSV 带 BOM, Excel 可开;
                            --matrix 且 .xlsx 时直接写 Excel 工作簿)
    --summary               末尾打印汇总统计表
    --report 文件           生成 JSON 报告 (每文件状态/异常/缺失节/节覆盖;
                             默认不含条目全文, 加 --with-entries 才包含)
    --with-entries          报告里包含提取条目全文 (默认剥离, 控文件体积)
    --verify                token 级完整性校验 (慢; 逐文件与原文逐 token 比对)
    --fail-fast             首个读取失败立即退出 (返回码 1)
    --skip-empty            跳过读取结果无任何内容的文件
    --verbose               打印每文件处理进度 (推荐批量化使用)
    --quiet                 只输出汇总/报告, 不输出逐文件内容

  返回码: 0 = 全部成功 | 1 = 存在失败/异常文件 | 2 = 参数错误

  示例:
    python batch_read.py "F:\\...\\PU-1034 msds_CN 国彩.docx" --verbose
    python batch_read.py "F:\\数据库\\MSDS\\中文" --sections 3 --tsv --out s3.tsv
    python batch_read.py "F:\\数据库\\MSDS\\英文\\*.docx" --query "供应商" --scope label
    python batch_read.py "F:\\数据库\\MSDS" --report scan.json --summary --quiet
    python batch_read.py "F:\\数据库\\MSDS" --verify --report integrity.json
    python batch_read.py "F:\\数据库\\MSDS\\中文" --matrix --out 对比表.xlsx
    python batch_read.py "F:\\数据库\\MSDS\\中文" --sections 3 --matrix --out 成分矩阵.tsv
    python batch_read.py "F:\\数据库\\MSDS\\中文" --matrix --states --out 对比表_带状态.xlsx
    python batch_read.py "F:\\数据库\\MSDS\\中文" --sections 3 --comp-cols --tsv --out 成分分列.tsv
    python batch_read.py "F:\\数据库\\MSDS" --comp-cols --json --out 成分分列.json
    # ---- 批量化检索增强 ----
    python batch_read.py "F:\\数据库\\MSDS" --query "二丙二醇" --hits          # 命中文件清单
    python batch_read.py "F:\\数据库\\MSDS" --query "供应商" --scope label \
        --name-filter BL,OS --tsv --out bl_os_供应商.tsv                        # 只查 BL/OS 型号
    python batch_read.py "F:\\数据库\\MSDS" --query "危险 警示" --any --tsv      # 多词 OR
    python batch_read.py "F:\\数据库\\MSDS" --query "9.4" --scope section --tsv  # 按节号检索
"""
from __future__ import annotations

import glob
import json
import re
import sys
import time
from pathlib import Path

# ============================================================
# 定位结构读取根目录 (绝对路径, 不依赖当前工作目录)
#   本文件位于: 结构读取/批量化读取/batch_read.py
#   根目录为:  结构读取/
# ============================================================
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

try:
    from core.docx_reader import read_msds
    from core.extract import build_hierarchy, flatten_nodes
    from core.s11 import S11_MAJOR_FIELDS, s11_group_rows
except ImportError as exc:  # 核心模块缺失 → 明确报错
    print(f"❌ 无法导入 MSDS 核心模块: {exc}", file=sys.stderr)
    print(f"   请确认目录结构: {_ROOT} 下应有 core/ (含 docx_reader.py / extract.py)", file=sys.stderr)
    sys.exit(2)


# S11 国标大类 → 固定序号 (11.1~11.10), 检索列头与入库总表 "11.N 大类名" 呼应
_S11_MAJOR_SEQ = {m: f"11.{i}" for i, m in enumerate(S11_MAJOR_FIELDS, 1)}
_S11_RANK = {m: i for i, m in enumerate(S11_MAJOR_FIELDS)}


# ============================================================
# 输入展开: 文件 / 目录(递归) / 通配符 → 文件列表
# ============================================================

def expand_inputs(args: list[str]) -> tuple[list[Path], list[str]]:
    """把位置参数展开为 docx 文件列表. 返回 (文件列表, 无法匹配的输入).

    - 已存在文件 → 直接收 (自动跳过 Word 锁文件 ~$*)
    - 已存在目录 → rglob 递归扫描 *.docx
    - 含通配符 → glob 展开 (匹配文件)
    - 不存在且无通配符 → 计入无法匹配列表
    """
    files: list[Path] = []
    unmatched: list[str] = []
    for raw in args:
        p = Path(raw)
        if p.exists() and p.is_file():
            if p.suffix.lower() == ".docx" and not p.name.startswith("~$"):
                files.append(p.resolve())
            continue
        if p.exists() and p.is_dir():
            files.extend(
                f.resolve() for f in p.rglob("*.docx")
                if not f.name.startswith("~$")
            )
            continue
        if any(ch in raw for ch in "*?[]"):
            hits = [Path(h).resolve() for h in glob.glob(raw)
                    if h.lower().endswith(".docx") and not Path(h).name.startswith("~$")]
            if hits:
                files.extend(hits)
            else:
                unmatched.append(raw)
            continue
        unmatched.append(raw)
    # 去重 (同文件可能被目录+通配符同时命中), 保持稳定排序
    seen: set[Path] = set()
    dedup: list[Path] = []
    for f in sorted(files, key=lambda x: str(x).lower()):
        if f not in seen:
            seen.add(f)
            dedup.append(f)
    return dedup, unmatched


# ============================================================
# token 级完整性校验 (--verify)
# ============================================================

_TOKEN_RE = re.compile(r"[一-鿿]{3,}|[A-Za-z]{3,}|\d{2,}")

def _tokens(text: str) -> set[str]:
    return set(_TOKEN_RE.findall(text or ""))

def _parse_text(result) -> str:
    """把 ParseResult 全文 (节标题/字段/行/成分) 拼回, 供 token 比对."""
    parts = [result.header, result.footer]
    for sec in result.sections.values():
        parts.extend([sec.full_title, sec.title])
        for f in sec.fields:
            parts.extend([f.label, f.value])
        parts.extend(sec.lines)
        parts.append(sec.component_header)
        for c in sec.components:
            parts.extend([c.name, c.cas, c.conc])
        # 内嵌子表 (S8.2 生物限值等): 表头+数据行 也纳入全文,
        # 避免 --verify 把子表内容误报为 source 缺失
        for st in sec.sub_tables:
            parts.extend([st.title, st.seq])
            parts.extend(st.header)
            for r in st.rows:
                parts.extend(r)
    return "\n".join(parts)


# ============================================================
# 批量化检索增强 (跨文档关键词检索)
#   - AND(默认)/OR 多词匹配
#   - scope=section 精确节号 (原 core 为包含匹配, 易误命中 9/19/90)
#   - 归一化与 core.extract._norm 语义一致 (不依赖 core 私有函数)
# ============================================================

_NORM_STRIP_RE = re.compile(r"[\s：:，,。；;（）()\-]")

def _norm(s: str) -> str:
    """归一化: 去空白/冒号/中英文标点, 小写 (与 core.extract._norm 同语义)."""
    return _NORM_STRIP_RE.sub("", (s or "")).lower()


def _any_or_all(flags, any_match: bool) -> bool:
    """AND(默认) / OR 组合."""
    flags = list(flags)
    return any(flags) if any_match else all(flags)


def search_entries(entries, query: str, scope: str = "all",
                   any_match: bool = False):
    """增强版跨文档检索 (batch_read 层, 不改 core).

    - query 空格分隔多词; any_match=False → AND (全词命中), True → OR (任一词)
    - scope: label / value / all / section (section=精确节号, 支持 "9" 或 "s9")
    - 返回新列表, 不影响原 entries
    - S0 页码字段不参与检索 (页脚动态页码无检索价值)
    """
    terms = [_norm(t) for t in query.split() if t.strip()]
    if not terms:
        return [e for e in entries
                if not (e.section == 0 and (e.label or "").strip() == "页码")]
    out = []
    for e in entries:
        if e.section == 0 and (e.label or "").strip() == "页码":
            continue
        if scope == "section":
            # 精确节号: 任一查询词剥 s 前缀后等于条目节号
            hit = any(_norm(str(e.section)) == _norm(t.lstrip("sS"))
                      for t in terms)
        elif scope == "label":
            hay = _norm(e.full_label())
            hit = _any_or_all((_norm(t) in hay for t in terms), any_match)
        elif scope == "value":
            hay = _norm(e.value)
            hit = _any_or_all((_norm(t) in hay for t in terms), any_match)
        else:  # "all": 标签 + 小标题 + 大标题 + 内容
            hay = (_norm(e.full_label()) + " " + _norm(e.sub_title) + " "
                   + _norm(e.big_title) + " " + _norm(e.value))
            hit = _any_or_all((_norm(t) in hay for t in terms), any_match)
        if hit:
            out.append(e)
    return out


# ============================================================
# 单文件处理
# ============================================================

def process_file(path: Path, sections: set[int] | None, query: str | None,
                 scope: str, verify: bool, any_match: bool = False) -> dict:
    """读取并提取单文件. 返回记录 dict (含状态/条目/异常/校验)."""
    from docx import Document  # 惰性导入 (仅 --verify 需要)

    record: dict = {"path": str(path), "name": path.name, "status": "ok"}
    try:
        r = read_msds(path)
    except Exception as exc:
        record.update({"status": "error", "error": f"{type(exc).__name__}: {exc}"})
        return record

    record["summary"] = r.summary()
    record["sections"] = {
        str(n): {"title": s.full_title,
                 "fields": len(s.fields), "lines": len(s.lines),
                 "components": len(s.components),
                 "is_component": s.is_component_table}
        for n, s in r.sections.items()
    }
    record["anomalies"] = [
        {"level": a.level, "section": a.section,
         "message": a.message, "detail": a.detail}
        for a in r.anomalies
    ]
    record["missing_sections"] = sorted(
        n for n in range(1, 17) if n not in r.sections
    )

    # 三级树 → 扁平条目 (供输出/检索)
    nodes = build_hierarchy(r)
    if sections:
        nodes = [n for n in nodes if n.number in sections]
    entries = flatten_nodes(nodes)
    if query:
        entries = search_entries(entries, query, scope, any_match)
    record["entries"] = [e.to_dict() for e in entries]
    record["entries_count"] = len(entries)
    # 命中状态: 检索时 = 是否命中至少一条; 未检索时所有 ok 文件均视为命中
    record["hit"] = (not query) or len(entries) > 0

    # 成分结构化: 有序列表 (供 --matrix / --comp-cols 拆列)
    #   component_list 保持 reader 顺序; components 为 name→(cas,conc) 映射
    sec3 = r.sections.get(3)
    record["component_list"] = [
        {"name": c.name, "cas": c.cas, "conc": c.conc}
        for c in (sec3.components if sec3 else [])
    ]
    record["components"] = {
        c.name: {"cas": c.cas, "conc": c.conc}
        for c in (sec3.components if sec3 else [])
    }

    # 完整性校验
    if verify:
        try:
            doc = Document(str(path))
            src = set()
            for tb in doc.tables:
                for row in tb.rows:
                    seen: set[int] = set()
                    for cell in row.cells:
                        if id(cell._tc) in seen:
                            continue
                        seen.add(id(cell._tc))
                        src |= _tokens(cell.text)
            miss = src - _tokens(_parse_text(r))
            record["verify"] = {"missing_count": len(miss),
                                "missing": sorted(miss)[:20]}
        except Exception as exc:
            record["verify"] = {"missing_count": -1,
                                "missing": [], "error": str(exc)}
    return record


# ============================================================
# 输出
# ============================================================

def _entries_from_record(rec: dict):
    """记录里的 entries dict → ExtractedField 列表 (供 render)."""
    from core.extract import ExtractedField
    out = []
    for d in rec.get("entries", []):
        e = ExtractedField(
            section=d.get("section", 0), big_title=d.get("big_title", ""),
            sub_title=d.get("sub_title", ""), label=d.get("label", ""),
            value=d.get("value", ""), seq=d.get("seq", ""),
            kind=d.get("kind", "field"), editable=d.get("editable", True),
        )
        out.append(e)
    return out


def render_text_simple(records: list[dict], verbose: bool, query: str | None = None,
                       show_empty: bool = False) -> str:
    """文本输出 (扁平分层, 每文件一个块).

    query 非空时默认过滤 0 命中文件 (减噪音); --show-empty 保留.
    """
    from core.extract import render_text as _render_text
    buf: list[str] = []
    for rec in records:
        if rec["status"] == "error":
            buf.append(f"⚠️ 读取失败 {rec['name']}: {rec['error']}")
            continue
        if query and not show_empty and rec["entries_count"] == 0:
            continue
        entries = _entries_from_record(rec)
        head = f"===== {rec['name']} | {len(entries)} 条 ====="
        if rec["anomalies"]:
            head += " [异常: "
            head += "; ".join(f"S{a['section']} {a['message']}" for a in rec["anomalies"])
            head += "]"
        buf.append(head)
        for a in rec["anomalies"]:
            mark = "⚠️" if a["level"] == "warn" else "❌"
            buf.append(f"  [{mark}] S{a['section']}: {a['message']} {a['detail']}")
        if rec.get("verify") and rec["verify"].get("missing_count", 0) > 0:
            buf.append(f"  [verify] token 遗漏 {rec['verify']['missing_count']} 个: {rec['verify']['missing']}")
        buf.append(_render_text(entries))
    return "\n".join(buf)


def render_tsv_all(records: list[dict]) -> str:
    """TSV 输出: 文件名 | 节 | 大标题 | 小标题 | 标签 | 字段."""
    rows = ["文件名\t节\t大标题\t小标题\t标签\t字段"]
    for rec in records:
        if rec["status"] != "ok":
            continue
        for e in _entries_from_record(rec):
            rows.append("\t".join([
                rec["name"], str(e.section), e.big_title, e.sub_title,
                e.full_label(), (e.value or "").replace("\t", " ").replace("\n", " "),
            ]))
    return "\n".join(rows)


def render_json_all(records: list[dict]) -> str:
    """JSON 输出: {文件名: 条目列表}."""
    payload = {rec["name"]: rec.get("entries", []) for rec in records
               if rec["status"] == "ok"}
    return json.dumps(payload, ensure_ascii=False, indent=2)


def render_hits(records: list[dict], with_query: bool = False) -> str:
    """文件命中清单 (--hits): 每命中文件一行 `命中条目数\t文件名`.

    - with_query: 有检索时 0 命中文件不列出; 未检索时列出全部 ok 文件
    - 按命中条目数降序排列, 末尾合计命中文件数/命中条目数
    """
    rows = []
    for rec in records:
        if rec["status"] != "ok":
            continue
        if with_query and rec["entries_count"] == 0:
            continue
        rows.append((rec["entries_count"], rec["name"]))
    rows.sort(key=lambda x: (-x[0], x[1].lower()))
    total_ok = sum(1 for r in records if r["status"] == "ok")
    total_hits = sum(n for n, _ in rows)
    buf = ["命中条目数\t文件名"]
    buf += [f"{n}\t{name}" for n, name in rows]
    buf.append(f"命中文件 {len(rows)} / {total_ok} | 命中条目 {total_hits}")
    return "\n".join(buf)


# ============================================================
# 成分分列输出 (--comp-cols): 每文档一行, 成分1|CAS1|含量1|成分2|CAS2|含量2
#   交替平铺. 只动输出层, 不改 core 解析 / GUI 显示.
# ============================================================

def _max_components(records: list[dict]) -> int:
    """所有 OK 记录里最多的成分数."""
    return max((len(r.get("component_list", [])) for r in records
                if r["status"] == "ok"), default=0)


def render_compcols_tsv(records: list[dict]) -> str:
    """成分分列 TSV: 文件 | 成分1 | CAS1 | 含量1 | 成分2 | CAS2 | 含量2 | ..."""
    n = _max_components(records)
    head = ["文件"]
    for i in range(1, n + 1):
        head += [f"成分{i}", f"CAS{i}", f"含量{i}"]
    rows = ["\t".join(head)]
    for rec in records:
        if rec["status"] != "ok":
            continue
        clist = rec.get("component_list", [])
        cells = [rec["name"]]
        for i in range(n):
            if i < len(clist):
                c = clist[i]
                cells += [c["name"], c["cas"], c["conc"]]
            else:
                cells += ["", "", ""]
        rows.append("\t".join(x.replace("\t", " ").replace("\n", "⏎") for x in cells))
    return "\n".join(rows)


def render_compcols_json(records: list[dict]) -> str:
    """成分分列 JSON: {文件名: [{name, cas, conc}, ...]}."""
    payload = {
        rec["name"]: rec.get("component_list", [])
        for rec in records if rec["status"] == "ok"
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def render_compcols_text(records: list[dict]) -> str:
    """成分分列文本 (每文件块内成分逐行分列)."""
    buf: list[str] = []
    for rec in records:
        if rec["status"] == "error":
            buf.append(f"⚠️ 读取失败 {rec['name']}: {rec['error']}")
            continue
        clist = rec.get("component_list", [])
        buf.append(f"===== {rec['name']} | 成分 {len(clist)} 个 =====")
        for i, c in enumerate(clist, 1):
            buf.append(f"  {i}. {c['name']}  [CAS: {c['cas']} | 含量: {c['conc']}]")
    return "\n".join(buf)


# ============================================================
# 宽表矩阵 (--matrix): 行 = 一个文档, 列 = 字段, 列头 = 字段标识
#   与 reader 显示同构 — 列头就是 reader 每行字段的"节 + 序号 + 标题":
#     - 有序号字段:  Section{n} {序号范围} {label}  如 Section9 9.4~9.5 初沸点
#     - 无序号字段:  Section{n} {label}            如 Section2 GHS分类
#     - 无标题文本:  Section{n} 特殊字段           如 Section5 特殊字段
#     - 成分:        Section3 成分{i}-名称/CAS/含量 (按成分索引拆三列)
#   列顺序 = (节号升序, 节内 reader 上下解析顺序), 与三列表呈现一致.
#   同 (节, 标签) 不同序号 → 合并一列, 列头显示序号范围 (消除序号漂移列分裂).
#   --states: 每字段列后跟"值状态"三态列 (有值/无数据/无此字段).
# ============================================================

def _seq_sort_key(seq: str) -> tuple:
    """'9.4' → (9,4); '9.10' → (9,10); '' → 排最后."""
    try:
        a, _, b = seq.partition(".")
        return (int(a or 0), int(b) if b.isdigit() else 10**9)
    except ValueError:
        return (10**9, 10**9)


def _seq_adjacent(a: str, b: str) -> bool:
    """两序号连续? 同主号且次号差 1 (9.4→9.5; 9.9→9.10)."""
    if not a or not b:
        return False
    aa, _, ab = a.partition(".")
    ba, _, bb = b.partition(".")
    if not aa.isdigit() or not ba.isdigit() or aa != ba:
        return False
    return (ab.isdigit() and bb.isdigit() and int(bb) == int(ab) + 1)


def _seq_range(seqs: set[str]) -> str:
    """合并列序号范围: '9.4' | '9.4~9.5' | '9.4~9.5, 9.17' (空 seq 忽略)."""
    seqs = sorted((s for s in seqs if s), key=_seq_sort_key)
    if not seqs:
        return ""
    parts: list[str] = []
    seg = [seqs[0]]
    for s in seqs[1:]:
        if _seq_adjacent(seg[-1], s):
            seg.append(s)
        else:
            parts.append(f"{seg[0]}~{seg[-1]}" if len(seg) > 1 else seg[0])
            seg = [s]
    parts.append(f"{seg[0]}~{seg[-1]}" if len(seg) > 1 else seg[0])
    return ", ".join(parts)


def build_matrix(records: list[dict], states: bool = False
                 ) -> tuple[list[str], list[tuple], list[tuple]]:
    """构建宽表. 返回 (列头列表, 列key列表, [(文件名, {列key: 值})]).

    列顺序 = (节号升序, 节内 reader 上下解析顺序), 与三列表呈现一致.
    列头:
      有序号字段 Section{n} {序号范围} {label} (同标签不同序号合并, 显示范围)
      无序号字段 Section{n} {label} | 无标题文本 Section{n} 特殊字段 |
      成分 Section3 成分{i}-名称/CAS/含量.
    states=True: 每字段列后跟状态列 (有值/无数据/无此字段).
    """
    from core.extract import ExtractedField
    cols: dict[tuple, dict] = {}      # key -> {"seqs"|None, "header"}
    rows: list[tuple] = []            # (name, {key: value})
    present: list[dict] = []          # (name, {key: 字段是否存在})
    pos_sum: dict[tuple, int] = {}    # key -> Σ 节内 reader 位置
    pos_cnt: dict[tuple, int] = {}    # key -> 出现文档数

    for rec in records:
        if rec["status"] != "ok":
            continue
        comps = rec.get("components", {})
        row: dict[tuple, str] = {}
        pr: dict[tuple, bool] = {}
        special: dict[int, int] = {}
        comp_idx: dict[int, int] = {}
        sec_counter: dict[int, int] = {}   # 每文档每节: 当前 reader 位置

        def _next_pos(sec: int) -> int:
            c = sec_counter.get(sec, 0)
            sec_counter[sec] = c + 1
            return c

        # S11 毒理: 按文档原始顺序收集 (label, value) → 国标大类归并
        # (复用 core.s11, 与入库总表 build_standard_table 取值完全一致;
        #  避免检索页面把 LD50/物种/方法等子项平铺成几十列不明所以的列)
        s11_rows = [(d.get("label", ""), d.get("value", ""))
                    for d in rec.get("entries", [])
                    if d.get("section") == 11 and d.get("kind") == "field"
                    and d.get("label")]
        s11_groups = s11_group_rows(s11_rows) if s11_rows else {}
        s11_pos: int | None = None   # 本文档 S11 首个字段的 reader 位置

        for d in rec.get("entries", []):
            e = ExtractedField(
                section=d.get("section", 0), big_title=d.get("big_title", ""),
                sub_title=d.get("sub_title", ""), label=d.get("label", ""),
                value=d.get("value", ""), seq=d.get("seq", ""),
                kind=d.get("kind", "field"), editable=d.get("editable", True),
            )
            if e.kind == "component":
                idx = comp_idx.get(e.section, 1)
                comp_idx[e.section] = idx + 1
                info = comps.get(e.label, {})
                pos = _next_pos(e.section)
                for field, val in (("名称", e.label),
                                   ("CAS", info.get("cas", "")),
                                   ("含量", info.get("conc", ""))):
                    key = (e.section, "c", idx, field)
                    if key not in cols:
                        cols[key] = {
                            "seqs": None,
                            "header": f"Section{e.section} 成分{idx}-{field}"}
                    pos_sum[key] = pos_sum.get(key, 0) + pos
                    pos_cnt[key] = pos_cnt.get(key, 0) + 1
                    pr[key] = True
                    if val:
                        row[key] = val
                continue
            # S11 毒理字段: 子项已在 s11_groups 归并, 不产独立列; 仅推进位置
            if e.section == 11 and e.label:
                if s11_pos is None:
                    s11_pos = _next_pos(11)
                else:
                    _next_pos(11)
                continue
            if e.label:
                key = (e.section, "f", e.label)
                # S2 防范说明 multi 归并: 预防措施/事故响应/安全储存/废弃处置
                # 等同义子类的 P 代码合并到 "防范说明" 一列 (schema 设计意图),
                # 保证 P/H 开头条目 (P280/P305+P351+P338/P405/P501/Hxxx) 在检索
                # 矩阵中完整可见, 而非只保留首个子类 (预防措施) 或分散成多列.
                if e.section == 2 and e.label != "防范说明":
                    from core.schema import standard_field_of
                    f = standard_field_of(2, e.label)
                    if f is not None and f.multi and f.name == "防范说明":
                        key = (e.section, "f", "防范说明")
                if key not in cols:
                    cols[key] = {"seqs": set(), "header": None}
                cols[key]["seqs"].add(e.seq)
                pos_sum[key] = pos_sum.get(key, 0) + _next_pos(e.section)
                pos_cnt[key] = pos_cnt.get(key, 0) + 1
                pr[key] = True
                if e.value:
                    if key in row:
                        row[key] = row[key] + "\n" + e.value
                    else:
                        row[key] = e.value
            else:
                idx = special.get(e.section, 1)
                special[e.section] = idx + 1
                key = (e.section, "s", idx)
                header = f"Section{e.section} 特殊字段"
                if idx > 1:
                    header += f" ({idx})"
                if key not in cols:
                    cols[key] = {"seqs": None, "header": header}
                pos_sum[key] = pos_sum.get(key, 0) + _next_pos(e.section)
                pos_cnt[key] = pos_cnt.get(key, 0) + 1
                pr[key] = True
                if e.value:
                    row[key] = e.value

        # S11 国标大类列: 按国标顺序建列, 值 = 该大类子项归并 (换行).
        # 列头 "Section11 11.N 大类名" 与入库总表 "11.N 大类名" 一一对应.
        if s11_groups:
            base = s11_pos if s11_pos is not None else 0
            for major in S11_MAJOR_FIELDS:
                vals = s11_groups.get(major)
                if not vals:
                    continue
                key = (11, "f", "S11:" + major)
                if key not in cols:
                    cols[key] = {"seqs": None,
                                 "header": f"Section11 {_S11_MAJOR_SEQ[major]} {major}"}
                pos_sum[key] = pos_sum.get(key, 0) + base
                pos_cnt[key] = pos_cnt.get(key, 0) + 1
                pr[key] = True
                row[key] = "\n".join(vals)
        rows.append((rec["name"], row))
        present.append(pr)

    # 字段列头定稿 (需要全部 seqs 才能算序号范围)
    for key, col in cols.items():
        if col["seqs"] is not None:
            head = f"Section{key[0]}"
            rng = _seq_range(col["seqs"])
            if rng:
                head += " " + rng
            head += " " + key[2]
            col["header"] = head

    # 列顺序 = (节号, 节内 reader 上下顺序). 取该标签在所有文档中
    # 该节内位置的**平均值** → 不在首文档的字段 (如 离子性) 也按其真实
    # 相对位置排, 不受"某个文档缺字段"影响; 平均位置相同则按插入序 (稳定).
    # S11 特殊: 总结句排最前, 10 大类别按国标固定顺序 (11.1→11.10),
    # 不随文档子项平铺乱序.
    def _col_key(k: tuple):
        if k[0] != 11:
            return (k[0], pos_sum[k] / pos_cnt[k])
        if k[1] == "s":
            return (11, 0, -1)                        # S11 总结句
        if k[1] == "f" and k[2].startswith("S11:"):
            return (11, 0, _S11_RANK.get(k[2][4:], 99))   # 大类按国标顺序
        return (11, pos_sum[k] / pos_cnt[k])

    ordered = sorted(cols, key=_col_key)
    keys: list[tuple] = []
    columns: list[str] = []
    for k in ordered:
        keys.append(k)
        columns.append(cols[k]["header"])
        if states:
            keys.append(k + ("state",))
            columns.append(cols[k]["header"] + "·状态")

    final_rows: list[tuple] = []
    for (name, row), pr in zip(rows, present):
        cells = dict(row)
        if states:
            for k in ordered:
                sk = k + ("state",)
                cells[sk] = ("有值" if k in row
                             else ("无数据" if k in pr else "无此字段"))
        final_rows.append((name, cells))
    return columns, keys, final_rows


def render_matrix_tsv(records: list[dict], states: bool = False) -> str:
    """矩阵 TSV (带 BOM, Excel 可直接开; 换行折成 ⏎ 保一行一文档)."""
    columns, keys, rows = build_matrix(records, states)
    lines = ["\t".join(["文件"] + columns)]
    for name, row in rows:
        cells = [name]
        for k in keys:
            v = row.get(k, "") or ""
            cells.append(v.replace("\t", " ").replace("\n", "⏎"))
        lines.append("\t".join(cells))
    return "\n".join(lines)


# 三态着色 (配合 --states 的值状态列)
_STATE_FILLS = {
    "有值":      "E2EFDA",   # 浅绿
    "无数据":    "F2F2F2",   # 浅灰
    "无此字段":  "FCE4EC",   # 浅红
}
_STATE_FONT_COLORS = {
    "有值":      "1F6F33",
    "无数据":    "808080",
    "无此字段":  "C62828",
}
_STATE_FONT = {  # 与状态列值一致 → 是否判定为状态列
    "有值", "无数据", "无此字段",
}


def _auto_width(texts: list[str], min_w: float = 9, max_w: float = 42) -> float:
    """按内容估算列宽: 中文≈2, ASCII≈1, 取最长行."""
    w = min_w
    for t in texts:
        if not t:
            continue
        n = sum(2 if ord(ch) > 0x2E80 else 1 for ch in t)
        n += t.count("\n") * 2   # 每换行增加一行高 → 略放宽
        w = max(w, n)
    return min(w + 2, max_w)


def write_matrix_xlsx(records: list[dict], out: str, states: bool = False) -> int:
    """矩阵写 .xlsx (美化: 节分组表头/三态着色/斑马纹/细边框/列宽自适应).

    - 表头: 深蓝白字加粗; 状态列表头 (·状态) 用紫色区分
    - 值状态三态: 有值=浅绿, 无数据=浅灰, 无此字段=浅红 (+ 同色字体)
    - 数据: 细边框 + 斑马纹行 (交替底色, 状态列覆盖斑马纹)
    - 列宽: 按内容自适应 (中文按 2 宽估算), 冻结首行 + 自动筛选
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("❌ --matrix 导出 xlsx 需要 openpyxl (pip install openpyxl)",
              file=sys.stderr)
        return -1
    columns, keys, rows = build_matrix(records, states)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MSDS矩阵"
    ws.append(["文件"] + columns)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    state_head_fill = PatternFill("solid", fgColor="6B4FA0")

    # 表头 (状态列用紫色)
    for i, cell in enumerate(ws[1], 1):
        is_state_col = i > 1 and columns[i - 2].endswith("·状态")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.fill = state_head_fill if is_state_col else head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center")
        cell.border = border

    # 数据行
    for r, (name, row) in enumerate(rows, start=2):
        zebra = "F8FAFD" if (r % 2 == 0) else None
        ws.append([name] + [row.get(k, "") or "" for k in keys])
        for i, cell in enumerate(ws[r], 1):
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="left")
            cell.font = Font(size=10)
            if i == 1:
                cell.font = Font(size=10, bold=True, color="1F4E79")
                if zebra:
                    cell.fill = PatternFill("solid", fgColor=zebra)
                continue
            v = cell.value
            if v in _STATE_FONT and states:
                # 三态着色
                cell.fill = PatternFill("solid", fgColor=_STATE_FILLS[v])
                cell.font = Font(size=10, color=_STATE_FONT_COLORS[v],
                                 bold=(v == "无数据"))
            elif zebra:
                cell.fill = PatternFill("solid", fgColor=zebra)

    # 列宽自适应 (按每列最长内容)
    ws.column_dimensions["A"].width = _auto_width(
        [r[0] for r in rows], min_w=22, max_w=30)
    for j, k in enumerate(keys, start=2):
        col_letter = openpyxl.utils.get_column_letter(j)
        texts = [(row.get(k, "") or "") for _, row in rows]
        ws.column_dimensions[col_letter].width = _auto_width(
            texts + [columns[j - 2]], min_w=9, max_w=40)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out)
    return len(rows)


def render_summary(records: list[dict], inputs: list[str],
                   query: str | None = None, any_match: bool = False) -> str:
    """汇总统计表. query 非空时额外显示检索命中统计."""
    ok = [r for r in records if r["status"] == "ok"]
    err = [r for r in records if r["status"] == "error"]
    total_entries = sum(r["entries_count"] for r in ok)
    total_fields = sum(r["summary"]["fields"] for r in ok)
    total_comps = sum(r["summary"]["components"] for r in ok)
    total_anom = sum(r["summary"]["anomalies"] for r in ok)
    if query:
        hit_files = [r for r in ok if r["entries_count"] > 0]
        hit_entries = total_entries
    missing_files = [r for r in ok if r["missing_sections"]]
    anom_by_section: dict[int, int] = {}
    for r in ok:
        for a in r["anomalies"]:
            anom_by_section[a["section"]] = anom_by_section.get(a["section"], 0) + 1
    verified = [r for r in ok if r.get("verify") and r["verify"].get("missing_count", 0) > 0]

    buf = ["", "─" * 58, "📊 批量读取汇总", "─" * 58]
    buf.append(f"  输入: {len(inputs)} 个 (文件/目录/通配符)")
    buf.append(f"  文件: 共 {len(records)} | ✅ 成功 {len(ok)} | ❌ 失败 {len(err)}")
    if query:
        mode = "OR" if any_match else "AND"
        buf.append(f"  🔍 检索 '{query}' ({mode}): 命中文件 {len(hit_files)} / {len(ok)}"
                   f" | 命中条目 {hit_entries}")
    if ok:
        buf.append(f"  字段: {total_fields} | 成分: {total_comps} | 提取条目: {total_entries}")
        buf.append(f"  异常: {total_anom} 项")
        if anom_by_section:
            buf.append("  异常分布 (按节): " + ", ".join(
                f"S{n}×{c}" for n, c in sorted(anom_by_section.items())))
        if missing_files:
            buf.append(f"  ⚠️ 缺失节: {len(missing_files)} 个文件")
            for r in missing_files[:8]:
                buf.append(f"      {r['name']}: 缺 S" + ",".join(map(str, r["missing_sections"])))
            if len(missing_files) > 8:
                buf.append(f"      ... 等 {len(missing_files)} 个")
        if verified:
            buf.append(f"  ⚠️ token 遗漏: {len(verified)} 个文件")
            for r in verified[:5]:
                buf.append(f"      {r['name']}: {r['verify']['missing_count']} 个")
            if len(verified) > 5:
                buf.append(f"      ... 等 {len(verified)} 个")
    if err:
        buf.append("  失败清单:")
        for r in err[:10]:
            buf.append(f"      {r['name']}: {r['error']}")
        if len(err) > 10:
            buf.append(f"      ... 等 {len(err)} 个")
    buf.append("─" * 58)
    return "\n".join(buf)


# ============================================================
# CLI
# ============================================================

def _parse_args(argv: list[str]) -> tuple | None:
    """解析参数. 返回 (inputs, opts) 或 None (参数错误)."""
    inputs: list[str] = []
    opts = {
        "sections": None, "query": None, "scope": "all",
        "fmt": "text", "out": None, "summary": False, "report": None,
        "verify": False, "fail_fast": False, "skip_empty": False,
        "verbose": False, "quiet": False, "with_entries": False,
        "states": False, "comp_cols": False,
        "any_match": False, "hits": False, "show_empty": False,
        "name_filter": None,
    }
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("--sections", "-s") and i + 1 < len(argv):
            opts["sections"] = {int(x) for x in argv[i + 1].split(",")
                                if x.strip().isdigit()}
            i += 2
        elif a in ("--query", "-q") and i + 1 < len(argv):
            opts["query"] = argv[i + 1]; i += 2
        elif a == "--scope" and i + 1 < len(argv):
            opts["scope"] = argv[i + 1]; i += 2
        elif a == "--any":
            opts["any_match"] = True; i += 1
        elif a == "--hits":
            opts["hits"] = True; i += 1
        elif a == "--show-empty":
            opts["show_empty"] = True; i += 1
        elif a == "--name-filter":
            # 值可逗号分隔 (BL,OS) 或连续多值 (BL OS); 直到下一个 - 开头参数.
            # 兼容 PowerShell 把 "BL-8125,BL-8127" 展平为两个独立参数.
            subs: list[str] = []
            j = i + 1
            while j < len(argv) and not argv[j].startswith("-"):
                subs.extend(x.strip() for x in argv[j].split(",") if x.strip())
                j += 1
            if subs:
                opts["name_filter"] = subs
            i = j
        elif a in ("--json", "-j"):
            opts["fmt"] = "json"; i += 1
        elif a in ("--tsv", "-t"):
            opts["fmt"] = "tsv"; i += 1
        elif a in ("--matrix", "-m"):
            opts["fmt"] = "matrix"; i += 1
        elif a == "--states":
            opts["states"] = True; i += 1
        elif a in ("--comp-cols", "-c"):
            opts["comp_cols"] = True; i += 1
        elif a in ("--text",):
            opts["fmt"] = "text"; i += 1
        elif a in ("--out", "-o") and i + 1 < len(argv):
            opts["out"] = argv[i + 1]; i += 2
        elif a == "--summary":
            opts["summary"] = True; i += 1
        elif a == "--report" and i + 1 < len(argv):
            opts["report"] = argv[i + 1]; i += 2
        elif a == "--with-entries":
            opts["with_entries"] = True; i += 1
        elif a == "--verify":
            opts["verify"] = True; i += 1
        elif a == "--fail-fast":
            opts["fail_fast"] = True; i += 1
        elif a == "--skip-empty":
            opts["skip_empty"] = True; i += 1
        elif a in ("--verbose", "-v"):
            opts["verbose"] = True; i += 1
        elif a == "--quiet":
            opts["quiet"] = True; i += 1
        elif a in ("-h", "--help"):
            return "help"
        else:
            inputs.append(a); i += 1
    return inputs, opts


_USAGE = __doc__.strip() + "\n"


def main(argv: list[str] | None = None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    parsed = _parse_args(argv)
    if parsed is None or parsed == "help":
        print(_USAGE)
        return 2 if parsed is None else 0
    inputs, opts = parsed

    # stdout UTF-8 (Windows 控制台中文不乱码)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not inputs:
        print(_USAGE)
        return 2

    files, unmatched = expand_inputs(inputs)
    for u in unmatched:
        print(f"⚠️ 无法匹配输入: {u}", file=sys.stderr)
    if not files:
        print("❌ 未找到任何 .docx 文件", file=sys.stderr)
        return 2 if unmatched else 2

    # --name-filter: 按文件名子串过滤 (逗号分隔 OR, 大小写不敏感)
    if opts["name_filter"]:
        subs = opts["name_filter"]
        files = [f for f in files
                 if any(s.lower() in f.name.lower() for s in subs)]
        if not files:
            print(f"❌ --name-filter {','.join(subs)} 过滤后无匹配文件",
                  file=sys.stderr)
            return 2

    if opts["verbose"] and not opts["quiet"]:
        print(f"📁 待处理 {len(files)} 个文件 (来自 {len(inputs)} 个输入)")
        if opts["name_filter"]:
            print(f"   文件名过滤: {','.join(opts['name_filter'])}")
        if opts["sections"]:
            print(f"   节过滤: {','.join(map(str, sorted(opts['sections'])))}")
        if opts["query"]:
            mode = "OR" if opts["any_match"] else "AND"
            print(f"   检索: '{opts['query']}' (scope={opts['scope']}, {mode})")

    records: list[dict] = []
    t0 = time.time()
    for idx, f in enumerate(files, 1):
        if opts["verbose"] and not opts["quiet"]:
            print(f"  [{idx}/{len(files)}] {f.name} ...", end=" ", flush=True)
        rec = process_file(f, opts["sections"], opts["query"],
                           opts["scope"], opts["verify"],
                           any_match=opts["any_match"])
        if rec["status"] == "error":
            if opts["verbose"] and not opts["quiet"]:
                print(f"❌ {rec['error']}")
            if opts["fail_fast"]:
                print(f"❌ --fail-fast 终止: {f} ({rec['error']})", file=sys.stderr)
                return 1
        elif opts["skip_empty"] and rec["entries_count"] == 0 and not rec["summary"]["sections"]:
            if opts["verbose"] and not opts["quiet"]:
                print("(跳过, 无内容)")
            continue
        else:
            if opts["verbose"] and not opts["quiet"]:
                print(f"✅ {rec['summary']['sections']}节 / {rec['summary']['fields']}字段 / {rec['summary']['components']}成分 / {rec['summary']['anomalies']}异常")
        records.append(rec)
    elapsed = time.time() - t0

    # ---- 输出 (--out 写文件不受 --quiet 抑制; quiet 只静默逐文件正文) ----
    if opts["hits"]:
        # --hits 文件命中清单: 聚合视图, 不受 --tsv/--json/--matrix/--comp-cols
        # 影响; 未检索时列出全部 ok 文件
        body = render_hits(records, bool(opts["query"]))
    elif opts["fmt"] == "json":
        body = (render_compcols_json(records) if opts["comp_cols"]
                else render_json_all(records))
    elif opts["fmt"] == "tsv":
        body = (render_compcols_tsv(records) if opts["comp_cols"]
                else render_tsv_all(records))
    elif opts["fmt"] == "matrix":
        body = render_matrix_tsv(records, opts["states"])
    else:
        body = (render_compcols_text(records) if opts["comp_cols"]
                else render_text_simple(records, opts["verbose"],
                                        opts["query"], opts["show_empty"]))
    if opts["out"]:
        n_ok = sum(1 for r in records if r["status"] == "ok")
        if opts["fmt"] == "matrix" and Path(opts["out"]).suffix.lower() == ".xlsx":
            n_rows = write_matrix_xlsx(records, opts["out"], opts["states"])
            columns, _, _ = build_matrix(records, opts["states"])
            print(f"✅ 已导出 {n_rows} 行(文档) × {len(columns) + 1} 列 → {opts['out']} ({elapsed:.1f}s)")
        else:
            enc = "utf-8-sig" if opts["fmt"] in ("tsv", "matrix") else "utf-8"
            Path(opts["out"]).write_text(body, encoding=enc)
            if opts["hits"]:
                print(f"✅ 已导出文件命中清单 → {opts['out']} ({elapsed:.1f}s)")
            else:
                n_entries = sum(r["entries_count"] for r in records)
                print(f"✅ 已导出 {n_entries} 条 / {n_ok} 文件 → {opts['out']} ({elapsed:.1f}s)")
    elif not opts["quiet"] or opts["hits"]:
        print(body)

    # ---- 汇总 (quiet 模式下仍打印, 因汇总不是逐文件内容) ----
    if opts["summary"]:
        print(render_summary(records, inputs, opts["query"], opts["any_match"]))

    # ---- 报告 (JSON, 含每文件状态/异常/缺失节) ----
    n_err = sum(1 for r in records if r["status"] == "error")
    if opts["report"]:
        # 默认报告不含 entries 全文 (全库可含数万条 → 数十 MB);
        # 加 --with-entries 才序列化条目 (供下游程序消费).
        report_records = records
        if not opts["with_entries"]:
            report_records = []
            for rec in records:
                slim = {k: v for k, v in rec.items() if k != "entries"}
                report_records.append(slim)
        report = {
            "tool": "MSDS batch_read",
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "elapsed_sec": round(elapsed, 2),
            "input_specs": inputs,
            "unmatched_inputs": unmatched,
            "total_files": len(records),
            "ok_files": sum(1 for r in records if r["status"] == "ok"),
            "error_files": n_err,
            "missing_inputs": len(unmatched),
            "with_entries": bool(opts["with_entries"]),
            "files": report_records,
        }
        Path(opts["report"]).write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"📄 报告已生成 → {opts['report']}")

    # 返回码: 存在读取失败 或 有输入无法匹配 → 1 (供批处理脚本判断)
    return 1 if (n_err or unmatched) else 0


if __name__ == "__main__":
    sys.exit(main())