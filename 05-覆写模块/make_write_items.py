#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用户交互覆写：Excel 表单 → 写入项 JSON（17 节契约）
====================================================
用户在测试库 Excel 里填写 Section 1/3/9，本脚本把填写内容转成写入项 JSON：
  - 忠实反映：填什么就覆写什么（不做人为截断/新增）；
  - 自动派生 S0 页眉页脚（产品名称/产品型号 从 S1 提取）；
  - 顶层补 keep_structure / empty_policy 契约。

用法:
  python make_write_items.py                 # 忠实模式（默认，用户交互）
  python make_write_items.py --demo          # 演示模式（人为构造差异覆盖 CRUD）
"""
import argparse
import json
import re
from pathlib import Path

EXCEL = Path(r"F:\正式项目与模块化内容\Word 覆写模块\数据库\测试库\PEA-4139 MSDS表单.xlsx")
OUT = Path(__file__).parent / "write_items_interactive.json"

FIELD_ALIASES = {
    "嗅觉阀值": "嗅觉阈值",
}

def seq_key(seq):
    if not seq:
        return None
    parts = str(seq).split(".")
    nums = []
    for p in parts:
        if p.isdigit():
            nums.append(int(p))
        else:
            break
    return tuple(nums)

def norm_label(s):
    if not s:
        return ""
    s = str(s).replace(" ", "").replace("　", "").replace("\xa0", "")
    return FIELD_ALIASES.get(s, s)

def norm_field(text):
    """提取 (seq, label)。'9.1 外观：' -> ('9.1','外观')；'中文名称：' -> ('','中文名称')。"""
    if not text:
        return "", ""
    s = str(text).replace("　", " ").replace("\xa0", " ").strip()
    m = re.match(r"^(\d+(?:\.\d+)*)\s*(.*)$", s)
    if m and m.group(2):
        seq, rest = m.group(1), m.group(2).strip()
    else:
        seq, rest = "", s
    rest = re.sub(r"[：:，,。；;]+$", "", rest).strip()
    rest = rest.replace(" ", "")
    return seq, norm_label(rest)

def derive_s0(s1_items):
    """从 S1 派生 S0 页眉页脚字段（用户交互时由系统自动生成）。
    产品名称/产品型号：优先 S1 产品名称；为空则从中文名称提取型号
    （如 '水性羟基聚酯-丙烯酸分散体 PEA-4139' → 'PEA-4139'）。
    Version/公司名称/修订日期 用户交互不填，保留模板原值。
    """
    prod = next((it["value"] for it in s1_items if it["label"] == "产品名称"), "").strip()
    cn = next((it["value"] for it in s1_items if it["label"] == "中文名称"), "").strip()
    if not prod:
        m = re.search(r"[A-Za-z]{1,6}-?\d{2,}", cn)
        prod = m.group(0) if m else ""
    if not prod:
        return []
    return [{"seq": "", "label": "产品名称", "value": prod},
            {"seq": "", "label": "产品型号", "value": prod}]

def main():
    ap = argparse.ArgumentParser(description="用户交互覆写：Excel 表单 → 写入项 JSON")
    ap.add_argument("--demo", action="store_true",
                    help="演示模式：人为构造差异（改/删/增）；默认忠实反映 Excel")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def t(v):
        return str(v).strip() if v is not None else ""

    # ---- 按 section 分组（与引擎同一判定）----
    blocks = {}
    cur = None
    for r in rows:
        a = t(r[0]) if len(r) > 0 else ""
        b = t(r[1]) if len(r) > 1 else ""
        c = t(r[2]) if len(r) > 2 else ""
        m = re.match(r"^(\d+)\.\s*[^\d\s]", a)
        if m:
            cur = int(m.group(1))
            blocks.setdefault(cur, []).append((a, b, c, True))
        elif cur is not None:
            blocks.setdefault(cur, []).append((a, b, c, False))

    # ---- S1：忠实全量（用户填了哪些字段就带哪些）----
    s1 = []
    for a, b, c, is_title in blocks.get(1, []):
        if is_title or not b:
            continue
        _, label = norm_field(a)
        if label:
            s1.append({"seq": "", "label": label, "value": b})

    # ---- S3：产品类型 + 全部成分（忠实）----
    s3 = {"产品类型": "", "components": []}
    in_comp = False
    comp_header_seen = False
    for a, b, c, is_title in blocks.get(3, []):
        seq, label = norm_field(a)
        if is_title:
            continue
        if label == "产品类型":
            s3["产品类型"] = b or ""
            continue
        if not b and not c and label == "成分":
            in_comp = True
            continue
        if in_comp:
            if not comp_header_seen:
                comp_header_seen = True
                continue
            if a or b or c:
                s3["components"].append({"name": a, "cas": b, "conc": c})

    # ---- S9：全部有值字段（忠实）----
    s9 = []
    for a, b, c, is_title in blocks.get(9, []):
        if is_title or not b:
            continue
        seq, label = norm_field(a)
        if label:
            s9.append({"seq": seq, "label": label, "value": b})
    s9.sort(key=lambda x: seq_key(x["seq"]))

    # ---- S0：自动派生页眉页脚 ----
    s0 = derive_s0(s1)

    # ---- 演示模式：人为构造差异覆盖 CRUD ----
    if args.demo:
        s1 = [it for it in s1 if it["label"] in
              ("中文名称", "化学品分类", "产品使用建议和使用限制")]
        s3["components"] = s3["components"][:2]
        s9 = [it for it in s9 if seq_key(it["seq"]) <= (9, 19)]
        s9.append({"seq": "9.23", "label": "折射率", "value": "1.45"})
        s9.sort(key=lambda x: seq_key(x["seq"]))
        s0 = []                        # 演示模式不派生 S0

    write_items = {"sections": {"0": s0, "1": s1, "3": s3, "9": s9},
                   "keep_structure": [2],
                   "empty_policy": "warn"}
    OUT.write_text(json.dumps(write_items, ensure_ascii=False, indent=2), encoding="utf-8")
    print("写入项已生成:", OUT)
    print("S0 字段数:", len(s0), s0)
    print("S1 字段数:", len(s1), [it["label"] for it in s1])
    print("S3 产品类型:", s3["产品类型"], "成分:", len(s3["components"]))
    print("S9 字段数:", len(s9), "首个:", s9[0]["label"], "末尾:", s9[-1]["label"], s9[-1]["value"])

if __name__ == "__main__":
    main()

