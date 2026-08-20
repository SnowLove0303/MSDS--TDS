# -*- coding: utf-8 -*-
"""MSDS SQLite 标准字段库 (围绕「父子级 + 标签」结构表的数据库实现).

把 MSDS 结构读取结果落成可查询的关系库, 四张表:

  schema_field  标准字段字典 (列定义来源)
      节 Schema 的标准字段: 节号 / 标准序号 / 标准字段名 / kind /
      collapse / multi / 同义别名(JSON) / 是否入宽表. 由 core/schema.py
      的 SECTION_SCHEMAS + _BASE_SCHEMAS 固化导出.

  msds_model    型号主表
      中文库每个型号严格一行: 型号 / 来源文件 / 来源类型(docx|xlsx) /
      sha256 / 统计摘要 / 导入时间. model_id 是明细与宽表的外键.

  msds_field    明细长表 (父子级全保留, 可追溯)
      每行一个 序号|标签|字段: model_id / 节 / 序号seq / 原始标签 /
      标准字段名(standard_name 归一化) / 值 / kind(field|sub|note|
      component|subtable) / editable / 行序. sub 行即"大标题"父级,
      按 (model_id, section, row_index) 即可还原 节→大标题→标签→字段.

  msds_wide     Schema 标准字段宽表 (主查询)
      每型号一行, 每个标准字段一列 (列名 S{节}__{标准名}, 节前缀防跨节
      重名; S0 页码列剔除, 与透视总表一致). 值 = standard_result 归并
      (同义合并 + 折叠空父级 + 多值 \n 连接), 无值 → NULL (展示层转
      「无数据」).

数据来源双轨 (source 字段区分):
  docx  由 read_msds 直接解析入库 (明细/宽表同源)
  xlsx  由「导出当前MSDS/模板」透视总表 (行1节/行2标签/行3值) 反灌

用法 (命令行):
  python tools/build_msds_db.py <db> <docx或目录...> [--from-xlsx 透视表.xlsx]
"""
from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import sys
from pathlib import Path

from .schema import (is_guide_line, standard_field_of, standard_fields,
                     standard_name, standard_result)

# 宽表剔除的字段 (2026-08-17 用户确认: 补齐 S0 页码 → 不再剔除)
_WIDE_EXCLUDE: set[str] = set()

# 列名中的特殊字符 (如 "氟化橡胶 –FKM" / "丁腈橡胶 – NBR") 原样保留为标识符,
# SQLite 支持 Unicode 标识符; 查询时用双引号引用列名.

_DDL = """
CREATE TABLE IF NOT EXISTS schema_field (
    id          INTEGER PRIMARY KEY,
    section     INTEGER NOT NULL,          -- 节号 0..16
    seq         TEXT DEFAULT '',           -- 标准序号 (如 "8.1"; 空=不固定)
    name        TEXT NOT NULL,             -- 标准字段名
    kind        TEXT DEFAULT 'field',      -- field | sub | note | comp
    collapse    INTEGER DEFAULT 0,         -- 空父级分组标签 → 折叠
    multi       INTEGER DEFAULT 0,         -- 可多行值 (列表字段)
    in_wide     INTEGER DEFAULT 1,         -- 是否生成宽表列 (页码=0)
    aliases     TEXT DEFAULT '[]',         -- 同义别名 JSON
    UNIQUE(section, name)
);

CREATE TABLE IF NOT EXISTS msds_model (
    model_id        INTEGER PRIMARY KEY AUTOINCREMENT,
    model           TEXT NOT NULL,         -- 型号 (如 PEA-4139)
    source          TEXT NOT NULL,         -- docx | xlsx
    source_file     TEXT NOT NULL,         -- 来源文件路径
    sha256          TEXT DEFAULT '',
    header          TEXT DEFAULT '',
    footer          TEXT DEFAULT '',
    sections_count  INTEGER DEFAULT 0,
    tables_count    INTEGER DEFAULT 0,
    fields_count    INTEGER DEFAULT 0,
    components_count INTEGER DEFAULT 0,
    anomalies_count INTEGER DEFAULT 0,
    language       TEXT NOT NULL DEFAULT 'zh-CN',
    skeleton_version TEXT NOT NULL DEFAULT '17-section-9.37',
    created_at      TEXT DEFAULT (datetime('now', 'localtime')),
    UNIQUE(model)
);

CREATE TABLE IF NOT EXISTS msds_field (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    model_id    INTEGER NOT NULL REFERENCES msds_model(model_id),
    section     INTEGER NOT NULL,          -- 节号
    seq         TEXT DEFAULT '',           -- 序号 (如 "8.1")
    label       TEXT DEFAULT '',           -- 原始标签 (去序号)
    std_name    TEXT DEFAULT '',           -- 标准字段名 (standard_name 归一化)
    value       TEXT DEFAULT '',           -- 字段内容
    kind        TEXT DEFAULT 'field',      -- field | sub | note | component | subtable
    editable    INTEGER DEFAULT 1,
    row_index   INTEGER DEFAULT 0,         -- 节内行序 (父子级顺序)
    sub_header  TEXT DEFAULT '[]',         -- kind=subtable: 表头列 JSON (如 ["组分名称","标准来源",...])
    sub_rows    TEXT DEFAULT '[]',         -- kind=subtable: 数据行 JSON ([[...],[...]])
    note_type   TEXT DEFAULT ''            -- kind=note 的类型化槽位: 产品说明|成分参考引导段|法规定义|欧盟废弃细则|指引段|法规条目|免责声明|(空=普通总结句)
);

CREATE TABLE IF NOT EXISTS msds_wide (
    model_id    INTEGER PRIMARY KEY REFERENCES msds_model(model_id)
    /* + 每标准字段一列: S{节}__{标准名}, 由 init_db 动态追加 */
);

CREATE INDEX IF NOT EXISTS idx_field_model ON msds_field(model_id);
CREATE INDEX IF NOT EXISTS idx_field_section ON msds_field(section, label);
CREATE INDEX IF NOT EXISTS idx_field_std ON msds_field(std_name);

CREATE TABLE IF NOT EXISTS field_mapping (
    mapping_id       INTEGER PRIMARY KEY AUTOINCREMENT,
    section          INTEGER NOT NULL,
    standard_name    TEXT NOT NULL,
    raw_label        TEXT NOT NULL,
    mapping_type     TEXT NOT NULL, -- 标准名|可选字段|单位变体|待归类
    occurrence_count INTEGER NOT NULL DEFAULT 0,
    status           TEXT NOT NULL DEFAULT 'active', -- active|pending|rejected
    first_seen_model TEXT DEFAULT '',
    UNIQUE(section, raw_label)
);

CREATE TABLE IF NOT EXISTS msds_unmapped (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    model_id    INTEGER NOT NULL REFERENCES msds_model(model_id),
    section     INTEGER NOT NULL,
    seq         TEXT DEFAULT '',
    raw_label   TEXT NOT NULL,
    raw_value   TEXT DEFAULT '',
    reason      TEXT NOT NULL DEFAULT '未匹配标准骨架',
    status      TEXT NOT NULL DEFAULT 'pending',
    created_at  TEXT DEFAULT (datetime('now', 'localtime')),
    UNIQUE(model_id, section, seq, raw_label, raw_value)
);

-- 兼容旧查询的 Section 9 映射视图表；数据由 refresh_s9_mapping 统一生成。
CREATE TABLE IF NOT EXISTS s9_label_mapping (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    std_name TEXT NOT NULL,
    std_seq TEXT DEFAULT '',
    raw_label TEXT NOT NULL,
    occurrences_count INTEGER DEFAULT 0,
    models_sample TEXT DEFAULT '',
    sample_values TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now', 'localtime')),
    UNIQUE(std_name, raw_label)
);

CREATE TABLE IF NOT EXISTS s9_raw_expression (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    model TEXT NOT NULL,
    source_file TEXT DEFAULT '',
    raw_seq TEXT DEFAULT '',
    raw_label TEXT DEFAULT '',
    std_name TEXT DEFAULT '',
    is_standard INTEGER DEFAULT 0,
    raw_value TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now', 'localtime'))
);
"""


# ------------------------------------------------------------------
# 建库
# ------------------------------------------------------------------

def wide_columns() -> list[tuple[int, str, str]]:
    """宽表列定义 [(节, 列名, 标准字段名), ...].

    遍历全部 17 节 (0..16): standard_fields(num) 自动取 SECTION_SCHEMAS
    或 _BASE_SCHEMAS; collapse 分组与剔除字段不产列.
    """
    cols: list[tuple[int, str, str]] = []
    for num in range(0, 17):
        for f in standard_fields(num):
            if f.collapse or f.name in _WIDE_EXCLUDE:
                continue
            cols.append((num, f"S{num}__{f.name}", f.name))
    return cols


def init_db(conn: sqlite3.Connection) -> list[tuple[int, str, str]]:
    """全量重建: 确保框架表存在 → DROP 四表 → 建表 + 灌 schema_field 字典.

    用于首次建库与 --init 重置; 调用即清空 msds_model/msds_field/msds_wide.
    """
    conn.executescript(_DDL)          # 首次: 框架表存在
    for t in ("msds_wide", "msds_unmapped", "field_mapping",
              "s9_raw_expression", "s9_label_mapping",
              "msds_field", "msds_model", "schema_field"):
        conn.execute(f"DROP TABLE IF EXISTS {t}")
    conn.executescript(_DDL)          # 重建全部 (含空 msds_wide)
    conn.execute("DROP TABLE IF EXISTS msds_wide")
    cols = wide_columns()
    wide_sql = ["model_id INTEGER PRIMARY KEY REFERENCES msds_model(model_id)"]
    for num, col, name in cols:
        wide_sql.append(f'"{col}" TEXT')
    conn.execute("CREATE TABLE msds_wide (" + ", ".join(wide_sql) + ")")

    cur = conn.cursor()
    for num in range(0, 17):
        for field_index, f in enumerate(standard_fields(num), 1):
            # Section 9 的规范序号是骨架的一部分，不能依赖源文件序号；
            # 原始文件的 9.x 序号只作为事实层信息保存。
            schema_seq = f"9.{field_index}" if num == 9 else f.seq
            cur.execute(
                "INSERT OR REPLACE INTO schema_field"
                " (section, seq, name, kind, collapse, multi, in_wide, aliases)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (num, schema_seq, f.name, f.kind,
                 1 if f.collapse else 0, 1 if f.multi else 0,
                 0 if f.name in _WIDE_EXCLUDE else 1,
                 json.dumps(list(f.aliases), ensure_ascii=False)))
            # 标准名和可选原始表达进入正式映射库；标准骨架名称永远取 f.name。
            cur.execute(
                "INSERT OR IGNORE INTO field_mapping"
                " (section, standard_name, raw_label, mapping_type, status)"
                " VALUES (?,?,?,?,?)",
                (num, f.name, f.name, "标准名", "active"))
            for alias in f.aliases:
                cur.execute(
                    "INSERT OR IGNORE INTO field_mapping"
                    " (section, standard_name, raw_label, mapping_type, status)"
                    " VALUES (?,?,?,?,?)",
                    (num, f.name, alias, "可选字段", "active"))
    conn.commit()
    return cols


def refresh_s9_mapping(conn: sqlite3.Connection) -> None:
    """从正式标准映射和原始明细刷新 S9 兼容映射表.

    正式来源是 field_mapping；两个 s9_* 表只保留兼容查询，不再作为
    骨架或标准字段定义来源。
    """
    conn.execute("DELETE FROM s9_label_mapping")
    conn.execute("DELETE FROM s9_raw_expression")
    conn.execute(
        "INSERT INTO s9_label_mapping"
        " (std_name, std_seq, raw_label, occurrences_count, models_sample, sample_values)"
        " SELECT fm.standard_name, sf.seq, fm.raw_label, fm.occurrence_count, '', ''"
        " FROM field_mapping fm LEFT JOIN schema_field sf"
        " ON sf.section=9 AND sf.name=fm.standard_name"
        " WHERE fm.section=9 AND fm.status='active' AND fm.standard_name<>''")
    conn.execute(
        "INSERT INTO s9_raw_expression"
        " (model, source_file, raw_seq, raw_label, std_name, is_standard, raw_value)"
        " SELECT m.model, m.source_file, f.seq, f.label, f.std_name,"
        " CASE WHEN f.label=f.std_name THEN 1 ELSE 0 END, f.value"
        " FROM msds_field f JOIN msds_model m ON m.model_id=f.model_id"
        " WHERE f.section=9 AND f.kind='field' AND f.label<>''")
    conn.commit()


# ------------------------------------------------------------------
# 入库: docx 源
# ------------------------------------------------------------------

def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def classify_note(section: int, value: str) -> str:
    """总结句类型化槽位 (用户确认的共用结构): 按节 + 特征词分类.

    - S11/S12: 引导段 (is_guide_line) → 成分参考引导段; 否则 → 产品说明
    - S13:     含 欧盟/EWC → 欧盟废弃细则; 否则 → 法规定义
    - S15:     引导段 → 指引段; 否则 → 法规条目
    - S16:     免责声明
    - 其他节: 空 (普通总结句, 不强制分类)
    """
    v = (value or "").strip()
    if section in (11, 12):
        return "成分参考引导段" if is_guide_line(v) else "产品说明"
    if section == 13:
        return "欧盟废弃细则" if ("欧盟" in v or "EWC" in v) else "法规定义"
    if section == 15:
        # 指引段: 引导句 (is_guide_line 或 以 物质/相关/本/此/按/根/以下 开头);
        # 否则为法规条目 (《xx法》/GB 编号/条例 等)
        if is_guide_line(v) or re.match(r"^(物质|相关|本|此|按|根|以下)", v):
            return "指引段"
        return "法规条目"
    if section == 16:
        return "免责声明"
    return ""


def _subtable_text(title: str, header: list[str],
                   rows: list[list[str]]) -> str:
    """子表 (S8.2 生物限值等) → 多行文本 (表头: 值 拼接, 同 extract 语义)."""
    lines: list[str] = []
    for r in rows:
        if header:
            pairs = [f"{h}: {v}" for h, v in zip(header, r) if str(v).strip()]
            lines.append(" | ".join(pairs) if pairs else " | ".join(r))
        else:
            lines.append(" | ".join(str(v) for v in r))
    return "\n".join(lines)


def _wide_row_values(result, cols: list[tuple[int, str, str]]) -> dict[str, str]:
    """宽表取值: 与明细入库同一归一化管线 (_db_std_name + 清单过滤).

    S11/S12 子项归并到大类、S15 任意法规名归并到「法规条目」、清单外字段丢弃 —
    保证宽表与 msds_field 明细完全一致 (此前 standard_result 未做归并导致
    S15 法规条目等宽表列取不到值).
    """
    from collections import defaultdict
    std: dict[int, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    cur_major: dict[int, str] = {}
    for num, sec in sorted(result.sections.items()):
        for row in sec.iter_rows():
            if row.kind != "field":
                continue
            s = _db_std_name(num, row.label, cur_major)
            if s and s in _listed_field_names(num) and (row.value or "").strip():
                std[num][s].append(row.value)

    out: dict[str, str] = {}
    for num, col, name in cols:
        vals = std.get(num, {}).get(name)
        if vals:
            v = "\n".join(x for x in vals if x and x.strip()).strip()
            if v:
                out[col] = v
    # 子表 (如 S8 生物限值) → 归入对应标准字段列 (无子表时列保持 NULL)
    for num, sec in sorted(result.sections.items()):
        for row in sec.iter_rows():
            if row.kind == "subtable" and row.label:
                std_name = standard_name(num, row.label)
                col = f"S{num}__{std_name}"
                if col in {c[1] for c in cols} and col not in out:
                    out[col] = _subtable_text(row.label, row.sub_header,
                                              row.sub_rows)
    return out


# S12 生态大类别 (飞书清单): 子项归并上下文
_S12_MAJOR = frozenset({"生态毒性", "持久性和降解性", "其他不利的影响"})


def _db_std_name(num: int, label: str, cur_major: dict[int, str]) -> str:
    """入库标准名归一化 (保证所有 field 都有统一标准名, 检索按此对齐).

    - 常规: schema.standard_name (同义别名/单位剥离)
    - S11: 子项 (方法/物种/NOAEL 等) → 当前国标大类 (11.1~11.10 上下文);
      无大类上下文用 label 语义兜底 (s11_sub_major)
    - S12: 子项 → 当前生态大类 (12.1~12.3 上下文)
    - S15: 未命中 Schema 的 field (任意法规名) → "法规条目" 统一
    """
    std = standard_name(num, label)
    if num == 11:
        from .s11 import S11_MAJOR_SET, s11_sub_major
        if label in S11_MAJOR_SET or std in S11_MAJOR_SET:
            cur_major[11] = std if std in S11_MAJOR_SET else label
            return cur_major[11]
        return cur_major.get(11) or s11_sub_major(label) or std
    if num == 12:
        if label in _S12_MAJOR:
            cur_major[12] = label
            return label
        return cur_major.get(12) or std
    if num == 15:
        s15_names = {f.name for f in standard_fields(15)}
        if std not in s15_names:
            return "法规条目"
    return std


# 每节清单字段名集合 (缓存): 清单外 field 行丢弃 (结构冻结)
_FIELD_NAME_CACHE: dict[int, frozenset[str]] = {}


def _listed_field_names(num: int) -> frozenset[str]:
    if num not in _FIELD_NAME_CACHE:
        _FIELD_NAME_CACHE[num] = frozenset(f.name for f in standard_fields(num))
    return _FIELD_NAME_CACHE[num]


def insert_docx(conn: sqlite3.Connection, path: str | Path,
                cols: list[tuple[int, str, str]]) -> int:
    """中文 docx → 单型号标准记录 + 原始映射审计 + 宽表.

    型号是中文库的业务唯一键；重新导入同型号时更新这一条记录，绝不
    生成第二条同型号数据。清单外字段不再丢弃，而是进入 msds_unmapped。
    """
    from .docx_reader import read_msds
    path = Path(path)
    result = read_msds(str(path))

    model = _model_of(result, path.name)

    cur = conn.cursor()
    cur.execute(
        "INSERT INTO msds_model (model, source, source_file, sha256, header, footer,"
        " sections_count, tables_count, fields_count, components_count, anomalies_count, language, skeleton_version)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)"
        " ON CONFLICT(model) DO UPDATE SET"
        " model=excluded.model, sha256=excluded.sha256, header=excluded.header,"
        " footer=excluded.footer, sections_count=excluded.sections_count,"
        " tables_count=excluded.tables_count, fields_count=excluded.fields_count,"
        " components_count=excluded.components_count,"
        " anomalies_count=excluded.anomalies_count, source=excluded.source,"
        " source_file=excluded.source_file, language=excluded.language,"
        " skeleton_version=excluded.skeleton_version"
        " RETURNING model_id",
        (model, "docx", str(path), _sha256(path), result.header, result.footer,
         result.sections_count, result.tables_count,
         sum(len(s.fields) for s in result.sections.values()),
         sum(len(s.components) for s in result.sections.values()),
         len(result.anomalies), "zh-CN", "17-section-9.37"))
    model_id = cur.fetchone()[0]

    conn.execute("DELETE FROM msds_field WHERE model_id=?", (model_id,))
    conn.execute("DELETE FROM msds_unmapped WHERE model_id=?", (model_id,))

    cur_major: dict[int, str] = {}
    for num, sec in sorted(result.sections.items()):
        for idx, row in enumerate(sec.iter_rows()):
            if row.kind == "section":
                continue
            if row.kind == "subtable":
                # 子表 (S8.2 生物限值等): 表头/数据行 JSON 入库, value=多行文本
                conn.execute(
                    "INSERT INTO msds_field (model_id, section, seq, label, std_name,"
                    " value, kind, editable, row_index, sub_header, sub_rows)"
                    " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (model_id, num, row.seq, row.label,
                     standard_name(num, row.label),
                     _subtable_text(row.label, row.sub_header, row.sub_rows),
                     "subtable", 0, idx,
                     json.dumps(list(row.sub_header), ensure_ascii=False),
                     json.dumps([list(r) for r in row.sub_rows], ensure_ascii=False)))
                continue
            kind = "component" if (sec.is_component_table and row.kind == "field"
                                   and not row.label) else row.kind
            ntype = classify_note(num, row.value) if kind == "note" else ""
            std = (_db_std_name(num, row.label, cur_major) if row.kind == "field"
                   else standard_name(num, row.label))
            # 结构冻结: 清单外 field 不得进入标准字段表，但必须进入待归类池，
            # 保留原始标签/值，供后续人工确认映射；绝不扩充标准骨架。
            if row.kind == "field" and std and std not in _listed_field_names(num):
                conn.execute(
                    "INSERT OR IGNORE INTO msds_unmapped"
                    " (model_id, section, seq, raw_label, raw_value, reason)"
                    " VALUES (?,?,?,?,?,?)",
                    (model_id, num, row.seq, row.label, row.value,
                     "未匹配标准骨架字段"))
                conn.execute(
                    "INSERT INTO field_mapping"
                    " (section, standard_name, raw_label, mapping_type, occurrence_count, status, first_seen_model)"
                    " VALUES (?,?,?,?,1,'pending',?)"
                    " ON CONFLICT(section, raw_label) DO UPDATE SET"
                    " occurrence_count=occurrence_count+1, status='pending'",
                    (num, "", row.label, "待归类", model))
                continue
            if row.kind == "field" and row.label.strip():
                conn.execute(
                    "UPDATE field_mapping SET occurrence_count=occurrence_count+1"
                    " WHERE section=? AND raw_label=?",
                    (num, row.label))
            conn.execute(
                "INSERT INTO msds_field"
                " (model_id, section, seq, label, std_name, value, kind, editable, row_index, note_type)"
                " VALUES (?,?,?,?,?,?,?,?,?,?)",
                (model_id, num, row.seq, row.label, std, row.value, kind,
                 1 if row.editable else 0, idx, ntype))
    # 成分行 (S3 components) 以 kind=component 单独入库
    for num, sec in sorted(result.sections.items()):
        if not sec.is_component_table:
            continue
        base = sum(1 for r in sec.iter_rows() if r.kind != "section")
        for ci, c in enumerate(sec.components):
            from .s3_component_std import standardize_component_name
            std_c = standardize_component_name(c.name, c.cas)
            conn.execute(
                "INSERT INTO msds_field"
                " (model_id, section, seq, label, std_name, value, kind, editable, row_index)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (model_id, num, "3", c.name, std_c,
                 f"CAS: {c.cas} | 含量: {c.conc}",
                 "component", 1 if c.editable else 0, base + ci))

    wide = _wide_row_values(result, cols)
    _upsert_wide(conn, model_id, wide)
    conn.commit()
    return model_id


def _model_of(result, file_name: str) -> str:
    """型号回退链: S1 产品名称 → S0 页眉产品名称 → 文件名前缀.

    模板类 docx 的 S1 产品名称常为空占位, 而页眉 (S0) 有实际型号
    (如 "产品名称: PEA-4139"), 逐级回退保证 docx/xlsx 同型号一致.
    """
    for num in (1, 0):
        sec = result.sections.get(num)
        if sec:
            for row in sec.iter_rows():
                if row.kind == "field" and row.label == "产品名称" \
                        and row.value.strip():
                    return row.value.strip()
    import re
    return re.split(r"\s+", Path(file_name).stem)[0]


def _upsert_wide(conn: sqlite3.Connection, model_id: int,
                 values: dict[str, str]) -> None:
    """写入宽表一行 (已存在则整行替换)."""
    cols = list(values.keys())
    if not cols:
        conn.execute("INSERT OR REPLACE INTO msds_wide (model_id) VALUES (?)",
                     (model_id,))
        return
    quoted = [f'"{c}"' for c in cols]
    conn.execute(
        f"INSERT INTO msds_wide (model_id, {', '.join(quoted)})"
        f" VALUES (?{', ?' * len(cols)})"
        f" ON CONFLICT(model_id) DO UPDATE SET "
        + ", ".join(f'"{c}"=excluded."{c}"' for c in cols),
        [model_id] + [values[c] for c in cols])


# ------------------------------------------------------------------
# 入库: xlsx 透视总表源 (行1 节 / 行2 序号+标签 / 行3 值)
# ------------------------------------------------------------------

def insert_pivot_xlsx(conn: sqlite3.Connection, path: str | Path,
                      cols: list[tuple[int, str, str]]) -> int:
    """透视总表 xlsx → 型号主表 + 明细 + 宽表. 返回 model_id."""
    from openpyxl import load_workbook
    from .structure import split_seq
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws.max_row < 3:
        raise ValueError(f"透视表行数不足 (需行1节/行2标签/行3数据): {path}")

    # 行1: 节 → 列范围; 行2: 序号+标签; 行3: 值 (A 列 = 型号)
    ncol = ws.max_column
    sec_of_col: dict[int, int | None] = {}
    cur_sec: int | None = None
    for ci in range(2, ncol + 1):
        r1 = str(ws.cell(1, ci).value or "").strip()
        # 节标题两种形态: "0 页眉/页脚" (无点号) / "1.物料及供应商标识" (带点号)
        m = re_match(r"^(\d+)", r1)
        if m:
            cur_sec = int(m.group(1))
        sec_of_col[ci] = cur_sec

    model = str(ws.cell(3, 1).value or Path(path).stem).strip() or Path(path).stem
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO msds_model (model, source, source_file)"
        " VALUES (?,?,?)"
        " ON CONFLICT(model) DO UPDATE SET source=excluded.source, source_file=excluded.source_file"
        " RETURNING model_id",
        (model, "xlsx", str(path)))
    model_id = cur.fetchone()[0]
    conn.execute("DELETE FROM msds_field WHERE model_id=?", (model_id,))
    conn.execute("DELETE FROM msds_unmapped WHERE model_id=?", (model_id,))

    ridx = 0
    comps: dict[int, dict[str, str]] = {}   # 成分槽: {idx: {name/cas/conc}}
    _comp_slot_re = re.compile(r"^成分(\d+)(名称|CAS|含量)$")
    for ci in range(2, ncol + 1):
        sec = sec_of_col.get(ci)
        label_raw = str(ws.cell(2, ci).value or "").strip()
        val = ws.cell(3, ci).value
        val = "" if val is None else str(val).strip()
        if sec is None or not label_raw:
            continue
        seq, label = split_seq(label_raw)
        if label_raw == "(总结句)":
            # 透视表 (总结句) 列可能合并多句 → 按行拆分, 每句一个类型化槽位
            for ln in [x.strip() for x in val.split("\n") if x.strip()]:
                conn.execute(
                    "INSERT INTO msds_field"
                    " (model_id, section, seq, label, std_name, value, kind, editable, row_index, note_type)"
                    " VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (model_id, sec, seq, label, "", ln, "note", 1, ridx,
                     classify_note(sec, ln)))
                ridx += 1
            continue
        m = _comp_slot_re.match(label)
        if m:
            # 成分槽列 (成分N名称/CAS/含量) → 组合为 component 行 (S3 结构)
            comps.setdefault(int(m.group(1)), {})[m.group(2)] = val
            continue
        std = standard_name(sec, label)
        if std and std not in _listed_field_names(sec):
            conn.execute(
                "INSERT OR IGNORE INTO msds_unmapped"
                " (model_id, section, seq, raw_label, raw_value, reason) VALUES (?,?,?,?,?,?)",
                (model_id, sec, seq, label, val, "未匹配标准骨架字段"))
            conn.execute(
                "INSERT INTO field_mapping"
                " (section, standard_name, raw_label, mapping_type, occurrence_count, status, first_seen_model)"
                " VALUES (?,?,?,?,1,'pending',?)"
                " ON CONFLICT(section, raw_label) DO UPDATE SET occurrence_count=occurrence_count+1, status='pending'",
                (sec, "", label, "待归类", model))
            continue
        if label:
            conn.execute(
                "UPDATE field_mapping SET occurrence_count=occurrence_count+1 WHERE section=? AND raw_label=?",
                (sec, label))
        conn.execute(
            "INSERT INTO msds_field"
            " (model_id, section, seq, label, std_name, value, kind, editable, row_index)"
            " VALUES (?,?,?,?,?,?,?,?,?)",
            (model_id, sec, seq, label, std, val, "field",
             1, ridx))
        ridx += 1

    # 成分槽 (S3): 组合后的 component 行入库
    for idx in sorted(comps):
        c = comps[idx]
        name = (c.get("名称") or "").strip()
        if not name:
            continue
        conn.execute(
            "INSERT INTO msds_field"
            " (model_id, section, seq, label, std_name, value, kind, editable, row_index)"
            " VALUES (?,?,?,?,?,?,?,?,?)",
            (model_id, 3, "3.2", name, standard_name(3, name),
             f"CAS: {c.get('CAS', '')} | 含量: {c.get('含量', '')}",
             "component", 1, ridx))
        ridx += 1

    # 宽表: 透视表值按标准字段归并 (同标签多列合并到同列, 取首个非空)
    wide: dict[str, str] = {}
    for ci in range(2, ncol + 1):
        sec = sec_of_col.get(ci)
        label_raw = str(ws.cell(2, ci).value or "").strip()
        val = ws.cell(3, ci).value
        val = "" if val is None else str(val).strip()
        if sec is None or not label_raw or label_raw == "(总结句)":
            continue
        std = standard_name(sec, label_raw)
        if label_raw:
            conn.execute(
                "UPDATE field_mapping SET occurrence_count=occurrence_count+1 WHERE section=? AND raw_label=?",
                (sec, label_raw))
        col = f"S{sec}__{std}"
        if std and col in {c[1] for c in cols} and val and col not in wide:
            wide[col] = val
    _upsert_wide(conn, model_id, wide)
    conn.commit()
    return model_id


def re_match(pattern: str, text: str):
    import re
    return re.match(pattern, text)


# ------------------------------------------------------------------
# 查询辅助
# ------------------------------------------------------------------

def list_models(conn: sqlite3.Connection) -> list[tuple]:
    """全部型号 (型号, 来源, 来源文件, 明细行数, 导入时间)."""
    return conn.execute(
        "SELECT m.model, m.source, m.source_file,"
        " (SELECT COUNT(*) FROM msds_field f WHERE f.model_id=m.model_id),"
        " m.created_at FROM msds_model m ORDER BY m.model_id").fetchall()


def search(conn: sqlite3.Connection, query: str, limit: int = 50) -> list[tuple]:
    """关键词检索 (型号/序号/标签/标准字段/值/槽位/Schema别名 LIKE), 返回 (型号, 节, 序号, 标签, 值)."""
    q = (query or "").strip()
    if not q:
        return []
    like = f"%{q}%"
    clause = (
        "(m.model LIKE ? OR f.seq LIKE ? OR f.label LIKE ? OR f.std_name LIKE ?"
        " OR f.value LIKE ? OR f.note_type LIKE ? OR f.sub_header LIKE ? OR f.sub_rows LIKE ?"
        " OR f.std_name IN (SELECT s.name FROM schema_field s WHERE s.section=f.section AND (s.name LIKE ? OR s.aliases LIKE ?)))"
    )
    return conn.execute(
        "SELECT m.model, f.section, f.seq,"
        " CASE WHEN f.section=3 AND f.kind='component' AND f.std_name!='' THEN f.std_name ELSE f.label END,"
        " substr(f.value,1,120)"
        " FROM msds_field f JOIN msds_model m ON m.model_id=f.model_id"
        f" WHERE {clause}"
        " ORDER BY m.model, f.section, f.row_index LIMIT ?",
        (like, like, like, like, like, like, like, like, like, like, limit)).fetchall()


def wide_row(conn: sqlite3.Connection, model_id: int) -> dict[str, str]:
    """宽表一行 → {列名: 值} (跳过 NULL)."""
    row = conn.execute("SELECT * FROM msds_wide WHERE model_id=?",
                       (model_id,)).fetchone()
    if row is None:
        return {}
    keys = [d[0] for d in conn.execute("SELECT * FROM msds_wide LIMIT 0").description]
    return {k: v for k, v in zip(keys, row) if v is not None and k != "model_id"}


def open_db(path: str | Path) -> sqlite3.Connection:
    """打开数据库; 仅当库文件不存在 (或 schema_field 空) 时初始化.

    已存在的库不重建, 避免 open 即摧毁 msds_wide 数据; 强制重建
    用 tools/build_msds_db.py --init (或先删除 db 文件).
    """
    conn = sqlite3.connect(str(path))
    try:
        has = conn.execute("SELECT 1 FROM schema_field LIMIT 1").fetchone()
    except sqlite3.Error:
        has = None
    if has is None:
        init_db(conn)
    return conn


# ==================================================================
# 数据库检索 API (PRD 第 3 节: 按型号唯一索引 / 关键词检索)
# ==================================================================

# 节标题 (规范纯标题, 不带前缀数字; 序号由 sn.number / sec 统一拼接)
SEC_TITLES: dict[int, str] = {
    0: "页眉/页脚", 1: "物料及供应商标识", 2: "危险性概述",
    3: "成分/组成资料", 4: "急救措施", 5: "消防措施",
    6: "意外泄漏措施", 7: "操作和储存", 8: "接触控制/个人防护",
    9: "物理和化学特性", 10: "稳定性和反应性", 11: "毒性资料",
    12: "生态信息", 13: "处理注意事项", 14: "运输信息",
    15: "法规信息", 16: "其他信息",
}


def sec_full_title(num: int) -> str:
    """返回规范全标题 (如 '1. 物料及供应商标识' / '0. 页眉/页脚')."""
    t = SEC_TITLES.get(num, f"第{num}节")
    return f"{num}. {t}"


def find_models(conn: sqlite3.Connection, query: str) -> list[tuple]:
    """按型号检索 (唯一索引): 精确 + 模糊 (包含匹配), 返回
    (model_id, model, source, source_file, fields_count, created_at)."""
    q = (query or "").strip()
    if not q:
        return []
    like = f"%{q}%"
    return conn.execute(
        "SELECT m.model_id, m.model, m.source, m.source_file,"
        " (SELECT COUNT(*) FROM msds_field f WHERE f.model_id=m.model_id),"
        " m.created_at FROM msds_model m"
        " WHERE m.model = ? OR m.model LIKE ?"
        " ORDER BY CASE WHEN m.model = ? THEN 0 ELSE 1 END, m.model",
        (q, like, q)).fetchall()


def model_detail(conn: sqlite3.Connection, model_id: int) -> dict:
    """型号元信息 (状态栏/CLI 头部展示)."""
    m = conn.execute(
        "SELECT model, source, source_file, sha256, header, footer,"
        " sections_count, tables_count, fields_count, components_count,"
        " anomalies_count, created_at FROM msds_model WHERE model_id=?",
        (model_id,)).fetchone()
    if m is None:
        return {}
    keys = ("model", "source", "source_file", "sha256", "header", "footer",
            "sections_count", "tables_count", "fields_count",
            "components_count", "anomalies_count", "created_at")
    return dict(zip(keys, m))


def _section_rows(conn: sqlite3.Connection, model_id: int,
                  sections: set[int] | None = None) -> list[tuple]:
    """明细行 (按 节 → 行序), 可选节过滤."""
    sql = ("SELECT section, seq, label, std_name, value, kind, editable,"
           " row_index, note_type, sub_header, sub_rows"
           " FROM msds_field WHERE model_id=?")
    args: list = [model_id]
    if sections:
        sql += " AND section IN (%s)" % ",".join("?" * len(sections))
        args.extend(sorted(sections))
    sql += " ORDER BY section, row_index"
    return conn.execute(sql, args).fetchall()


def model_tree_nodes(conn: sqlite3.Connection, model_id: int,
                     sections: set[int] | None = None) -> list:
    """明细 → 三级父子级树 (SectionNode), 与 core.extract.build_hierarchy
    同结构同顺序: 序号行 = 独立父级, 无序号标签挂最近父级, S11 国标大类归并."""
    from .extract import BigTitleNode, FieldNode, SectionNode
    from .s11 import S11_MAJOR_FIELDS, S11_MAJOR_SEQ, s11_group_rows

    rows = _section_rows(conn, model_id, sections)
    nodes: list[SectionNode] = []
    cur_num: int | None = None
    sn: SectionNode | None = None
    cur: BigTitleNode | None = None
    field_rows: list[tuple] = []

    def flush_s11() -> None:
        """S11 字段行 → 国标大类归并 (11.1~11.10), 与 docx 树一致."""
        nonlocal field_rows
        if cur_num != 11 or not field_rows or sn is None:
            field_rows = []
            return
        groups = s11_group_rows([(r[2], r[3]) for r in field_rows
                                 if r[2] and r[3]])
        for major in S11_MAJOR_FIELDS:
            vals = groups.get(major)
            if not vals:
                continue
            sn.big_titles.append(BigTitleNode(
                seq=S11_MAJOR_SEQ[major], title=major, value="\n".join(vals),
                kind="field", editable=True))
        field_rows = []

    for (sec, seq, label, std, value, kind, editable, ridx,
         note_type, sub_header, sub_rows) in rows:
        if sec != cur_num:
            flush_s11()
            cur_num = sec
            sn = SectionNode(number=sec, title=SEC_TITLES.get(sec, f"第{sec}节"),
                             full_title=sec_full_title(sec))
            nodes.append(sn)
            cur = None
        if kind == "sub":
            flush_s11()
            cur = BigTitleNode(seq=seq, title=label, kind="sub",
                               editable=bool(editable), index=ridx)
            sn.big_titles.append(cur)
        elif kind == "field":
            if sec == 11:
                field_rows.append((ridx, seq, label, value))
                continue
            if seq:
                cur = BigTitleNode(seq=seq, title=label, value=value,
                                   kind="field", editable=bool(editable),
                                   index=ridx)
                sn.big_titles.append(cur)
            elif cur is not None:
                cur.children.append(FieldNode(label=label, value=value,
                                              kind="field",
                                              editable=bool(editable),
                                              index=ridx))
            else:
                sn.direct_fields.append(FieldNode(label=label, value=value,
                                                  kind="field",
                                                  editable=bool(editable),
                                                  index=ridx))
        elif kind == "note":
            fn = FieldNode(label=note_type or "", value=value, kind="note",
                           editable=bool(editable), index=ridx)
            if cur is not None:
                cur.children.append(fn)
            else:
                sn.direct_fields.append(fn)
        elif kind == "component":
            from .s3_component_std import standardize_component_name
            cas = ""
            for part in (value or "").split(" | "):
                if part.startswith("CAS:"):
                    cas = part[4:].strip()
            std_c = standardize_component_name(label, cas)
            fn = FieldNode(label=std_c, value=value, kind="component",
                           editable=bool(editable), index=ridx)
            if cur is not None:
                cur.children.append(fn)
            else:
                sn.direct_fields.append(fn)
        elif kind == "subtable":
            import json as _json
            fn = FieldNode(label=label, value=value, kind="subtable",
                           editable=bool(editable), index=ridx,
                           sub_header=_json.loads(sub_header or "[]"),
                           sub_rows=_json.loads(sub_rows or "[]"))
            if cur is not None:
                cur.children.append(fn)
            else:
                sn.direct_fields.append(fn)
    flush_s11()
    return nodes


def listed_tree_nodes(conn: sqlite3.Connection, model_id: int,
                      sections: set[int] | None = None,
                      s9_active_only: bool = False,
                      active_only: bool = False) -> list:
    """按清单骨架构建三级树 (CLI --model 渲染): 结构与 PEA-4139 模板参照一致.

    s9_active_only=True: Section 9 仅输出有真实数据的项 (排除 无数据/不适用 占位).
    """
    from .extract import BigTitleNode, FieldNode, SectionNode
    nodes: list[SectionNode] = []
    for num in sorted(_SKELETON):
        if sections and num not in sections:
            continue
        sn = SectionNode(number=num, title=SEC_TITLES.get(num, f"第{num}节"),
                         full_title=sec_full_title(num))
        cur: BigTitleNode | None = None
        for row in listed_section_rows(conn, model_id, num,
                                      s9_active_only=s9_active_only,
                                      active_only=active_only):
            if row.kind == "sub":
                cur = BigTitleNode(seq=row.seq, title=row.label, kind="sub")
                sn.big_titles.append(cur)
            elif row.kind == "field":
                if row.seq:
                    cur = BigTitleNode(seq=row.seq, title=row.label,
                                       value=row.value, kind="field")
                    sn.big_titles.append(cur)
                elif cur is not None:
                    cur.children.append(FieldNode(label=row.label,
                                                  value=row.value, kind="field"))
                else:
                    sn.direct_fields.append(FieldNode(label=row.label,
                                                      value=row.value, kind="field"))
            elif row.kind == "note":
                fn = FieldNode(label=row.label, value=row.value, kind="note")
                if cur is not None:
                    cur.children.append(fn)
                else:
                    sn.direct_fields.append(fn)
            elif row.kind == "subtable":
                fn = FieldNode(label=row.label, value=row.value, kind="subtable",
                               sub_header=row.sub_header, sub_rows=row.sub_rows)
                if cur is not None:
                    cur.children.append(fn)
                else:
                    sn.direct_fields.append(fn)
        nodes.append(sn)
    return nodes


def render_model_tree(conn: sqlite3.Connection, model_id: int,
                      sections: set[int] | None = None,
                      s9_active_only: bool = False,
                      active_only: bool = False) -> str:
    """三级父子级树文本 (清单骨架, 同 main.py --extract 的 render_tree 格式)."""
    from .extract import render_tree
    return render_tree(listed_tree_nodes(conn, model_id, sections,
                                         s9_active_only=s9_active_only,
                                         active_only=active_only))


def render_model_json(conn: sqlite3.Connection, model_id: int,
                      sections: set[int] | None = None,
                      s9_active_only: bool = False,
                      active_only: bool = False) -> str:
    """嵌套 JSON (清单骨架, 同 render_tree_json)."""
    from .extract import render_tree_json
    return render_tree_json(listed_tree_nodes(conn, model_id, sections,
                                              s9_active_only=s9_active_only,
                                              active_only=active_only))


def render_model_tsv(conn: sqlite3.Connection, model_id: int,
                     sections: set[int] | None = None,
                     s9_active_only: bool = False,
                     active_only: bool = False) -> str:
    """扁平 TSV: 文件|节|大标题|小标题|标签|字段 (清单骨架)."""
    from .extract import flatten_nodes
    model = conn.execute("SELECT model FROM msds_model WHERE model_id=?",
                         (model_id,)).fetchone()
    fname = model[0] if model else f"model{model_id}"
    rows = ["文件\t节\t大标题\t小标题\t标签\t字段"]
    for e in flatten_nodes(listed_tree_nodes(conn, model_id, sections,
                                            s9_active_only=s9_active_only,
                                            active_only=active_only)):
        rows.append("\t".join([
            fname, str(e.section), e.big_title, e.sub_title,
            e.full_label(), e.value.replace("\t", " ").replace("\n", " ")]))
    return "\n".join(rows)


def model_search(conn: sqlite3.Connection, query: str,
                 sections: set[int] | None = None, limit: int = 200) -> list[tuple]:
    """关键词检索库内数据: 匹配 型号/序号/标签/标准字段名/字段内容/槽位/Schema别名 (空格多词 AND).

    返回 [(型号, model_id, 节, 序号, 标签, 值摘要), ...] (按型号→节→行序).
    """
    terms = [t.strip() for t in (query or "").split() if t.strip()]
    if not terms:
        return []
    clause = (
        "(m.model LIKE ? OR f.seq LIKE ? OR f.label LIKE ? OR f.std_name LIKE ?"
        " OR f.value LIKE ? OR f.note_type LIKE ? OR f.sub_header LIKE ? OR f.sub_rows LIKE ?"
        " OR f.std_name IN (SELECT s.name FROM schema_field s WHERE s.section=f.section AND (s.name LIKE ? OR s.aliases LIKE ?)))"
    )
    cond = " AND ".join(clause for _ in terms)
    sql = ("SELECT m.model, m.model_id, f.section, f.seq, f.label,"
           " substr(f.value,1,120), f.kind, f.std_name"
           " FROM msds_field f JOIN msds_model m ON m.model_id=f.model_id"
           f" WHERE {cond}")
    args: list = []
    for t in terms:
        p = f"%{t}%"
        args.extend([p] * 10)
    if sections:
        sql += " AND f.section IN (%s)" % ",".join("?" * len(sections))
        args.extend(sorted(sections))
    sql += (" ORDER BY m.model, f.section, f.row_index LIMIT ?")
    args.append(limit)
    return conn.execute(sql, args).fetchall()


def model_section_rows(conn: sqlite3.Connection, model_id: int,
                       num: int) -> list:
    """一节明细 → SectionRow 列表 (GUI SectionView.show_rows 直接渲染).

    - S11: field 行按国标大类归并 (11.1~11.10), note 行保留 — 与 docx 视图一致
    - subtable 行携带 sub_header/sub_rows (JSON 还原)
    - component 行以 field 行呈现 (label=成分名, value=原明细)
    """
    from .s11 import S11_MAJOR_FIELDS, S11_MAJOR_SEQ, s11_group_rows
    from .structure import SectionRow

    rows = _section_rows(conn, model_id, {num})
    out: list[SectionRow] = []
    if num == 11:
        fields = [(r[2], r[3]) for r in rows if r[5] == "field" and r[2]]
        groups = s11_group_rows(fields) if fields else {}
        for major in S11_MAJOR_FIELDS:
            vals = groups.get(major)
            if vals:
                out.append(SectionRow(kind="field", seq=S11_MAJOR_SEQ[major],
                                      label=major, value="\n".join(vals),
                                      editable=True))
        for r in rows:
            if r[5] == "note":
                out.append(SectionRow(kind="note", label=r[8] or "",
                                      value=r[3], editable=bool(r[6]), span=True,
                                      index=r[7]))
        return out
    for (sec, seq, label, std, value, kind, editable, ridx,
         note_type, sub_header, sub_rows) in rows:
        if kind == "sub":
            out.append(SectionRow(kind="sub", seq=seq, label=label,
                                  editable=bool(editable), index=ridx))
        elif kind == "field":
            out.append(SectionRow(kind="field", seq=seq, label=label, value=value,
                                  editable=bool(editable), index=ridx))
        elif kind == "note":
            out.append(SectionRow(kind="note", label=note_type or "",
                                  value=value, editable=bool(editable),
                                  span=True, index=ridx))
        elif kind == "component":
            out.append(SectionRow(kind="field", label=label, value=value,
                                  editable=bool(editable), index=ridx))
        elif kind == "subtable":
            out.append(SectionRow(kind="subtable", label=label, value=value,
                                  editable=False, span=True, index=ridx,
                                  sub_header=json.loads(sub_header or "[]"),
                                  sub_rows=json.loads(sub_rows or "[]")))
    return out


# ==================================================================
# 清单骨架渲染 (结构冻结: 展示 = 飞书第 5 章 17 节清单结构)
#
# 用户确认 (2026-08-17): 全过程只显示飞书规定的父子级字段及结构要求,
# 以 PEA-4139 模板检索结构为标准参照 — 每节按清单骨架渲染, 数据按标准名
# 填充, 清单外/缺失一律「无数据」, 不显示任何清单外结构.
# ==================================================================

# 每节清单骨架: [(kind, seq, label, note_type?), ...]
#   kind: sub(父级标题) / field(标签; 带 seq 者二级, 无 seq 者挂最近父级下) /
#         note(总结句槽位) / component(成分表) / subtable(生物限值子表)
# 结构参照 (用户确认): 定稿模板 PEA-4139 MSDS_CN 冠志 模板.docx 导入检索呈现的结构;
#   父子级对照模板原 word — 无冒号标题行=sub, 冒号结尾字段行=field(带 seq).
_SKELETON: dict[int, list[tuple]] = {
    0: [("sub", "0.1", "页眉"), ("field", "", "Version"), ("field", "", "产品名称"),
        ("sub", "0.2", "页脚"), ("field", "", "公司名称"), ("field", "", "产品型号"),
        ("field", "", "修订日期"), ("field", "", "页码")],
    1: [("sub", "1.1", "产品名称"), ("field", "", "中文名称"), ("field", "", "化学品分类"),
        ("field", "1.2", "产品使用建议和使用限制"), ("sub", "1.3", "供应商信息"),
        ("field", "", "供应商名称"), ("field", "", "供应商地址"),
        ("field", "", "电话"), ("field", "", "传真")],
    2: [("sub", "2.1", "紧急情况概述"), ("field", "2.2", "GHS危险性类别"),
        ("field", "2.3", "GHS标签要素"), ("field", "", "GHS象形图"),
        ("field", "2.4", "信号词"), ("field", "2.5", "危险性说明"),
        ("field", "2.5", "防范说明"), ("field", "2.6", "物理和化学危险"),
        ("field", "2.7", "健康危害"), ("field", "2.8", "环境危害"),
        ("field", "2.9", "其他危害")],
    3: [("field", "3.1", "产品类型"), ("sub", "3.2", "成分"), ("component", "", "成分表")],
    4: [("field", "4.1", "一般措施"), ("field", "4.2", "误服"), ("field", "4.3", "接触眼睛"),
        ("field", "4.4", "接触皮肤"), ("field", "4.5", "吸入")],
    5: [("field", "5.1", "合适的灭火剂"), ("field", "5.2", "不合适的灭火剂"),
        ("field", "5.3", "物质或混合物的特殊危害"), ("field", "5.4", "消防预防措施和保护设备")],
    6: [("field", "6.1", "个人预防措施、应急程序"), ("field", "6.2", "环境保护措施"),
        ("field", "6.3", "污染物收集和清除的方法")],
    7: [("field", "7.1", "安全操作防范"), ("field", "7.2", "安全储存条件")],
    8: [("sub", "8.1", "暴露控制"), ("field", "", "呼吸系统防护"), ("field", "", "手部防护"),
        ("field", "", "防护手套的合适材料"), ("field", "", "氟化橡胶 –FKM"),
        ("field", "", "丁基橡胶 –IIR"), ("field", "", "丁腈橡胶 – NBR"),
        ("field", "", "建议"), ("field", "", "眼睛防护"), ("field", "", "身体防护"),
        ("sub", "8.2", "生物限值"), ("subtable", "", "生物限值"),
        ("field", "8.3", "工程控制")],
    9: [("field", "9.1", "外观"), ("field", "9.2", "嗅觉阈值"), ("field", "9.3", "pH值"),
        ("field", "9.4", "离子性"), ("field", "9.5", "初沸点"), ("field", "9.6", "闪点"),
        ("field", "9.7", "蒸发速率"), ("field", "9.8", "可燃性（固态、气态）"),
        ("field", "9.9", "燃烧值"), ("field", "9.10", "饱和蒸气压"),
        ("field", "9.11", "相对蒸气密度"), ("field", "9.12", "密度"),
        ("field", "9.13", "水溶性"), ("field", "9.14", "表面张力"),
        ("field", "9.15", "辛醇/水分配系数对数值"), ("field", "9.16", "自燃温度"),
        ("field", "9.17", "引燃温度"), ("field", "9.18", "分解温度"),
        ("field", "9.19", "动力粘度"), ("field", "9.20", "爆炸特性"),
        ("field", "9.21", "粉尘爆炸级别"), ("field", "9.22", "固体含量"),
        # 14 项全量扩充字段 (9.23 ~ 9.36):
        ("field", "9.23", "有效成分"), ("field", "9.24", "玻璃化温度"),
        ("field", "9.25", "最低成膜温度"), ("field", "9.26", "NCO含量"),
        ("field", "9.27", "羟基含量"), ("field", "9.28", "熔点/凝固点"),
        ("field", "9.29", "酸值"), ("field", "9.30", "碘值"),
        ("field", "9.31", "倾点"), ("field", "9.32", "浊点"),
        ("field", "9.33", "分子量"), ("field", "9.34", "含量"),
        ("field", "9.35", "APHA值"), ("field", "9.36", "HLB值"),
        # 其他信息必须始终放在 Section 9 序号最末项 (9.37):
        ("field", "9.37", "其他信息")],
    10: [("field", "10.1", "化学稳定性"), ("field", "10.2", "危险分解产物"),
         ("field", "10.3", "可能的危害反应"), ("field", "10.4", "应避免的条件"),
         ("field", "10.5", "禁配物")],
    11: [("note", "", "", "产品说明"), ("note", "", "", "成分参考引导段"),
         ("field", "11.1", "急性毒性"), ("field", "11.2", "主要皮肤刺激性"),
         ("field", "11.3", "主要眼睛刺激性"), ("field", "11.4", "致敏性"),
         ("field", "11.5", "致突变性"), ("field", "11.6", "致癌性"),
         ("field", "11.7", "生殖毒性"),
         ("field", "11.8", "特异性靶器官系统毒性（一次接触/反复接触）"),
         ("field", "11.9", "吸入危险"), ("field", "11.10", "附加信息")],
    12: [("note", "", "", "产品说明"), ("note", "", "", "成分参考引导段"),
         ("field", "12.1", "生态毒性"), ("field", "12.2", "持久性和降解性"),
         ("field", "12.3", "其他不利的影响")],
    13: [("field", "", "处理方法")],
    14: [("field", "14.1", "公路和铁路运输"), ("field", "14.2", "海上运输"),
         ("field", "14.3", "空运"), ("field", "14.4", "用户特殊注意事项")],
    15: [("note", "", "", "指引段"), ("note", "", "其它的规定", ""),
         ("note", "", "符合下列法规要求", ""), ("note", "", "", "法规条目")],
    16: [("note", "", "", "免责声明")],
}

_ND = "无数据"   # 缺失值统一标注 (结构冻结: 宁可无数据不扩结构)

# 占位符与无效值正则 (用于 S9 仅有数据项输出 / 覆写写入项导出):
_PLACEHOLDER_RE = re.compile(
    r"^(无数据|无数据资料|不适用|无适用资料|未提供|无资料|无|N/A|n/a|None|null|"
    r"无数据。|不适用。|无。|Not applicable|not applicable|NA|—|--|-)$", re.IGNORECASE)


def is_meaningful_value(v: str) -> bool:
    """判断字段值是否为有意义的真实数据 (排除空值、无数据、不适用等占位符)."""
    t = (v or "").strip()
    return bool(t) and not bool(_PLACEHOLDER_RE.match(t))


# ------------------------------------------------------------------
# 结构写死 (结构冻结强制校验)
#
# 用户确认 (2026-08-17): 父子级结构及标签字段必须固定死, 防止其他线程/
# Agent 在批量洗数据时改动. 本指纹固化当前 17 节 schema 字段集 + 骨架
# (含 S9 方案2 全量扩充 14 字段 → S9 37 项, 全库 114 宽表列/116 字典项,
#  其他信息始终置于 9.37 序号末项), build_msds_db.py 每次入库前强制校验,
# 不一致即拒绝入库.
# ------------------------------------------------------------------

# v2.3 (2026-08-18): S13 骨架修正 — 移除脱离模板/骨架标准的 note 槽位
#   ('法规定义'/'欧盟废弃细则'), 仅保留事实存在的方法处理字段 '处理方法'.
#   模板 S13 仅 3 行 (节标题 + 首行说明 + 处理方法), 用户确认非法标签一律禁止.
# 旧固化值 e50fdef9f3cf9aa5 对应含非法 note 槽位的骨架, 已废弃.
STRUCT_FINGERPRINT = "a33596548469b67d"


def structure_fingerprint() -> str:
    """当前结构的指纹 (17 节 schema 字段集 + 骨架)."""
    import hashlib
    h = hashlib.sha256()
    for num in range(17):
        h.update(f"S{num}:".encode())
        for f in standard_fields(num):
            h.update(f"{f.name}|{f.kind}|{f.collapse}|{f.multi}|"
                     f"{','.join(f.aliases)};".encode())
        h.update("SKELETON:".encode())
        for it in _SKELETON.get(num, []):
            h.update("|".join(str(x) for x in it).encode())
    return h.hexdigest()[:16]


def validate_structure() -> str:
    """强制校验冻结结构; 不一致抛 RuntimeError (拒绝入库). 返回当前指纹."""
    cur = structure_fingerprint()
    if cur != STRUCT_FINGERPRINT:
        raise RuntimeError(
            "❌ 冻结结构校验失败: 当前 schema/_SKELETON 与固化版本不一致 "
            f"(期望 {STRUCT_FINGERPRINT}, 实际 {cur}).\n"
            "结构已被人为改动 — 禁止入库. 如需变更结构, 必须更新 "
            "STRUCT_FINGERPRINT 并经用户确认.")
    return cur


def _listed_rows_core(num: int, by_std: dict[str, list[str]],
                      notes: dict[str, list[str]], comps: list,
                      subtable, model_name: str = "",
                      s9_active_only: bool = False,
                      active_only: bool = False) -> list:
    """骨架渲染核心 (数据源无关): 按 _SKELETON 渲染一节 SectionRow 列表.

    供 数据库检索 (listed_section_rows) 与 docx 直读 (listed_rows_from_result)
    共用, 保证两类展示结构完全一致 (基于结构找内容).

    s9_active_only=True: Section 9 只输出有实际有效数据的项 (非 无数据/不适用),
    专为覆写模块写入 / S9 瘦身输出提供口子, 避免 37 个字段全量写入造成 docx 冗长.
    """
    from .structure import SectionRow
    out: list[SectionRow] = []
    # S1: 产品名称若为纯型号 → 与中文名称合并 (参照 PEA-4139 模板)
    if num == 1 and model_name:
        pname = (by_std.get("产品名称") or [""])[0].strip()
        if pname and pname == model_name:
            cname = (by_std.get("中文名称") or [""])[0].strip()
            if cname and model_name not in cname:
                by_std["中文名称"] = [f"{cname} {model_name}"]
            by_std["产品名称"] = [""]
    for item in _SKELETON.get(num, []):
        kind, seq, label = item[0], item[1], item[2]
        ntype = item[3] if len(item) > 3 else ""
        if kind == "sub":
            out.append(SectionRow(kind="sub", seq=seq, label=label,
                                  editable=False))
        elif kind == "field":
            vals = by_std.get(label) or []
            # 优先取有意义的真实数据 (排除 无数据/不适用 等占位符), 若全为占位符则取首个
            meaningful = [x for x in vals if is_meaningful_value(x)]
            if meaningful:
                v = "\n".join(meaningful).strip()
            else:
                v = "\n".join(x for x in vals if x and x.strip()).strip()
            filter_this = active_only or (s9_active_only and num == 9)
            if filter_this and not is_meaningful_value(v):
                continue
            out.append(SectionRow(kind="field", seq=seq, label=label,
                                  value=v or _ND, editable=True))
        elif kind == "note":
            if ntype == "法规条目":
                items = notes.get("法规条目") or []
                for i, v in enumerate(items, 1):
                    out.append(SectionRow(kind="note", label=f"法规条目{i}",
                                          value=v or _ND, editable=True,
                                          span=True))
                if not items:
                    out.append(SectionRow(kind="note", label="法规条目",
                                          value=_ND, editable=True, span=True))
            elif ntype:
                vals = notes.get(ntype) or []
                v = "\n".join(x for x in vals if x and x.strip()).strip()
                out.append(SectionRow(kind="note", label=ntype,
                                      value=v or _ND, editable=True, span=True))
            else:
                # 独立说明段 (S15 其它的规定/符合下列法规要求): 跨表格列字段判定,
                # 显示说明文字本身 (无内容时), 有内容则显示内容
                content = notes.get(label) or []
                v = "\n".join(x for x in content if x and x.strip()).strip()
                out.append(SectionRow(kind="note", label=label,
                                      value=v or label, editable=True,
                                      span=True))
        elif kind == "component":
            # 成分子表 (用户确认): 3 列 × N+1 行, 表头 [成分|CAS|含量], 每成分一行
            from .s3_component_std import standardize_component_name
            header = ["成分", "CAS", "含量"]
            rows: list[list[str]] = []
            for c in comps:
                cas = conc = ""
                for part in (c.value or "").split(" | "):
                    if part.startswith("CAS:"):
                        cas = part[4:].strip()
                    elif part.startswith("含量:"):
                        conc = part[3:].strip()
                std_c = standardize_component_name(c.label, cas)
                rows.append([std_c, cas, conc])
            out.append(SectionRow(
                kind="subtable", label="成分", value="", editable=False,
                span=True, sub_header=header, sub_rows=rows))
        elif kind == "subtable":
            if subtable is not None:
                out.append(subtable)
            else:
                # 生物限值子表: 无数据也显示表头 + 预留 1 行空行 (结构固定)
                out.append(SectionRow(
                    kind="subtable", label="生物限值", value="", editable=False,
                    span=True,
                    sub_header=["组分名称", "标准来源", "生物监测指标",
                                "生物限值", "采样时间"],
                    sub_rows=[["", "", "", "", ""]]))
    return out


def listed_section_rows(conn: sqlite3.Connection, model_id: int,
                        num: int, s9_active_only: bool = False,
                        active_only: bool = False) -> list:
    """数据库检索骨架渲染: 从明细收集 → _listed_rows_core (与 docx 直读同结构).

    s9_active_only=True: Section 9 仅输出有实际有效数据的项 (排除 无数据/不适用 占位).
    """
    from .structure import SectionRow
    rows = _section_rows(conn, model_id, {num})
    by_std: dict[str, list[str]] = {}
    notes: dict[str, list[str]] = {}
    comps: list[SectionRow] = []
    subtable: SectionRow | None = None
    for (sec, seq, label, std, value, kind, editable, ridx,
         note_type, sub_header, sub_rows) in rows:
        if kind == "field" and std:
            by_std.setdefault(std, []).append(value)
        elif kind == "note" and note_type:
            notes.setdefault(note_type, []).append(value)
        elif kind == "component":
            comps.append(SectionRow(kind="field", label=label, value=value,
                                    editable=bool(editable), index=ridx))
        elif kind == "subtable":
            subtable = SectionRow(kind="subtable", label=label, value=value,
                                  editable=False, span=True, index=ridx,
                                  sub_header=json.loads(sub_header or "[]"),
                                  sub_rows=json.loads(sub_rows or "[]"))
    model = conn.execute("SELECT model FROM msds_model WHERE model_id=?",
                         (model_id,)).fetchone()
    return _listed_rows_core(num, by_std, notes, comps, subtable,
                             model[0] if model else "",
                             s9_active_only=s9_active_only,
                             active_only=active_only)


def listed_rows_from_result(result, num: int, s9_active_only: bool = False,
                            active_only: bool = False) -> list:
    """docx 直读骨架渲染: 从 ParseResult 按清单骨架渲染一节.

    与数据库检索 (listed_section_rows) 完全同结构 — 用户确认: docx 直读
    展示/检索输出也必须符合骨架结构 (后续按输出做匹配覆写).

    s9_active_only=True: Section 9 仅输出有实际有效数据的项.
    """
    from .structure import SectionRow
    sec = result.sections.get(num)
    by_std: dict[str, list[str]] = {}
    notes: dict[str, list[str]] = {}
    comps: list[SectionRow] = []
    subtable: SectionRow | None = None
    if sec is not None:
        cur_major: dict[int, str] = {}
        for row in sec.iter_rows():
            if row.kind == "field":
                std = _db_std_name(num, row.label, cur_major)
                if std and std in _listed_field_names(num) \
                        and (row.value or "").strip():
                    by_std.setdefault(std, []).append(row.value)
            elif row.kind == "note":
                nt = classify_note(num, row.value)
                if nt:
                    notes.setdefault(nt, []).append(row.value)
            elif row.kind == "subtable":
                subtable = SectionRow(
                    kind="subtable", label=row.label,
                    value=_subtable_text(row.label, row.sub_header, row.sub_rows),
                    editable=False, span=True,
                    sub_header=list(row.sub_header),
                    sub_rows=[list(r) for r in row.sub_rows])
        if sec.is_component_table:
            for c in sec.components:
                comps.append(SectionRow(
                    kind="field", label=c.name,
                    value=f"CAS: {c.cas} | 含量: {c.conc}",
                    editable=c.editable))
    model_name = _model_of(result, result.file_name)
    return _listed_rows_core(num, by_std, notes, comps, subtable, model_name,
                             s9_active_only=s9_active_only,
                             active_only=active_only)


def listed_tree_nodes_from_result(result, sections: set[int] | None = None,
                                  s9_active_only: bool = False,
                                  active_only: bool = False) -> list:
    """docx 直读骨架树 (CLI --extract / GUI 目录): 结构与数据库检索完全一致."""
    from .extract import BigTitleNode, FieldNode, SectionNode
    nodes: list[SectionNode] = []
    for num in sorted(_SKELETON):
        if sections and num not in sections:
            continue
        sn = SectionNode(number=num, title=SEC_TITLES.get(num, f"第{num}节"),
                         full_title=sec_full_title(num))
        cur: BigTitleNode | None = None
        for row in listed_rows_from_result(result, num,
                                           s9_active_only=s9_active_only,
                                           active_only=active_only):
            if row.kind == "sub":
                cur = BigTitleNode(seq=row.seq, title=row.label, kind="sub")
                sn.big_titles.append(cur)
            elif row.kind == "field":
                if row.seq:
                    cur = BigTitleNode(seq=row.seq, title=row.label,
                                       value=row.value, kind="field")
                    sn.big_titles.append(cur)
                elif cur is not None:
                    cur.children.append(FieldNode(label=row.label,
                                                  value=row.value, kind="field"))
                else:
                    sn.direct_fields.append(FieldNode(label=row.label,
                                                      value=row.value, kind="field"))
            elif row.kind == "note":
                fn = FieldNode(label=row.label, value=row.value, kind="note")
                if cur is not None:
                    cur.children.append(fn)
                else:
                    sn.direct_fields.append(fn)
            elif row.kind == "subtable":
                fn = FieldNode(label=row.label, value=row.value, kind="subtable",
                               sub_header=row.sub_header, sub_rows=row.sub_rows)
                if cur is not None:
                    cur.children.append(fn)
                else:
                    sn.direct_fields.append(fn)
        nodes.append(sn)
    return nodes


def model_to_write_items(conn: sqlite3.Connection, model_id: int,
                         s9_active_only: bool = True) -> dict:
    """生成直接对接覆写模块 (msds_overwrite_engine.py) 的标准写入项契约 JSON.

    涵盖 0~16 全节标准数据:
      - S0: 页眉/页脚 (产品名称/产品型号)
      - S1~S16: 各节有效字段列表 (S9 默认仅输出有效值项, 避免覆写时生成冗余空行)
      - S3: 成分子表对象字典 {"产品类型": ..., "components": [...]}
      - S8: 包含 8.2 生物限值子表结构对象
    """
    detail = model_detail(conn, model_id)
    model = detail.get("model", "")

    sec_dict: dict[str, list | dict] = {}

    # S0
    sec_dict["0"] = [
        {"seq": "", "label": "产品名称", "value": model},
        {"seq": "", "label": "产品型号", "value": model}
    ]

    for num in range(1, 17):
        if num == 3:
            s3_type = "混合物"
            s3_comps = []
            for r in listed_section_rows(conn, model_id, 3):
                if r.kind == "field" and r.label == "产品类型" and is_meaningful_value(r.value):
                    s3_type = r.value
                elif r.kind == "subtable" and r.label == "成分":
                    for row in r.sub_rows:
                        if len(row) >= 3 and any(str(x).strip() for x in row):
                            s3_comps.append({"name": str(row[0]).strip(),
                                             "cas": str(row[1]).strip(),
                                             "conc": str(row[2]).strip()})
            sec_dict["3"] = {"产品类型": s3_type, "components": s3_comps}
        elif num == 8:
            s8_items = []
            for r in listed_section_rows(conn, model_id, 8):
                if r.kind == "field" and is_meaningful_value(r.value):
                    s8_items.append({"seq": r.seq or "", "label": r.label, "value": r.value})
                elif r.kind == "subtable" and r.label == "生物限值":
                    # 检查子表是否有真实数据
                    real_rows = [row for row in r.sub_rows if any(str(x).strip() and not _PLACEHOLDER_RE.match(str(x).strip()) for x in row)]
                    if real_rows:
                        s8_items.append({
                            "seq": "8.2",
                            "label": "生物限值",
                            "value": "",
                            "subtable": {
                                "header": list(r.sub_header),
                                "rows": real_rows
                            }
                        })
            sec_dict["8"] = s8_items
        elif num == 15:
            s15_items = []
            for r in listed_section_rows(conn, model_id, 15):
                if r.kind == "note" and is_meaningful_value(r.value):
                    s15_items.append({"seq": "", "label": r.label, "value": r.value})
                elif r.kind == "field" and is_meaningful_value(r.value):
                    s15_items.append({"seq": r.seq or "", "label": r.label, "value": r.value})
            sec_dict["15"] = s15_items
        else:
            items = []
            s9_clean = (num == 9 and s9_active_only)
            for r in listed_section_rows(conn, model_id, num, s9_active_only=s9_clean):
                if r.kind == "field" and is_meaningful_value(r.value):
                    items.append({"seq": r.seq or "", "label": r.label, "value": r.value})
                elif r.kind == "note" and is_meaningful_value(r.value):
                    items.append({"seq": "", "label": r.label, "value": r.value})
            sec_dict[str(num)] = items

    return {
        "sections": sec_dict,
        # 模板驱动保护节（除 S9 外全部）: 模板字段行全保留为骨架, 写入项只覆写值,
        # 模板多出字段保留; S9 保持瘦身驱动 (写入项为准, 过滤无数据占位行).
        # 修复 (2026-08-18): 旧值 [2] 仅保护 S2, 导致 S8 无生物限值数据时
        # 走通用 sync_section 且 keep_structure=False, 误删手部防护/手套材料/
        # 生物限值表头/工程控制行 — 违反飞书 v2.1 模板驱动规范.
        "keep_structure": [1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 16],
        "empty_policy": "warn",
        "missing_policy": "no_data",
        "missing_text": "无数据",
    }


if __name__ == "__main__":
    sys.exit("请通过 tools/build_msds_db.py 调用")
