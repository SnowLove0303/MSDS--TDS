# -*- coding: utf-8 -*-
"""生成 MSDS 入库总表 (透视结构). 供 GUI「导出 Excel 库表」与命令行 tools 复用.

透视结构:
  - 第一行: Section 节标题 (合并单元格, 每节跨其全部列)
  - 第二行: 该节下的小标题/标签 (sub 小标题 + field 字段标签, 逐列)
  - 第三行起: 每型号一行, 按 节→标签 一一对照填入内容
  - A1 / A2 留空, A 列为型号
  - 成分表 (S3) 展开为 名称/CAS/含量 每成分三列

围栏 (异常防护):
  - 自动过滤 Word 临时文件 (~$ 开头)
  - 单个文件读取失败 → 跳过并记录失败清单, 不中断整体导出
  - 目录无 docx / 全部失败 → 明确报错

命令行用法:
  python tools/build_pivot_table.py <入库目录> <输出xlsx>
"""
from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .docx_reader import read_msds
from .schema import standard_name

# ---- 样式 ----
FONT = "Microsoft YaHei"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")   # 深蓝: 节标题行
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")    # 浅蓝: sub 小标题
TAG_FILL = PatternFill("solid", fgColor="EDEDED")    # 浅灰: 字段标签
MODEL_FILL = PatternFill("solid", fgColor="F2F2F2")  # 型号列
NOTE_FILL = PatternFill("solid", fgColor="FFF7E6")   # 总结句
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _model_of(result, file_name: str) -> str:
    """型号: 优先取 S1 产品名称, 回退文件名前缀."""
    sec = result.sections.get(1)
    if sec:
        for row in sec.iter_rows():
            if row.kind == "field" and row.label == "产品名称" and row.value.strip():
                return row.value.strip()
    return re.split(r"\s+", Path(file_name).stem)[0]


# S9 理化特性 同义标签归一化 (消除不同模板/表述造成的重复概念列分裂):
#   - 复用 schema.standard_name 完整同义映射 + 单位/条件括号剥离:
#       粘度/25℃ 与 动力粘度 → 粘度    蒸发速率 与 百分比挥发性 → 蒸发速率
#       自燃温度 与 引燃温度/着火点 → 自燃温度    蒸汽压 → 饱和蒸气压
#       相对蒸气密度 → 蒸气密度    嗅觉阀值 → 嗅觉阈值
#   - 例外: standard_name 折叠到「其他信息」的别名 (分子量/APHA值/含量/HLB 等)
#     保留独立列, 不强行并入「其他信息」— 列多无妨, 但不能有重复概念.
_S9_PAREN = re.compile(r"[（(][^（）()]*[）)]$")


def _norm_s9(label: str) -> str:
    """S9 字段标签归一化: schema 同义归一化; 折叠别名保留独立列."""
    std = standard_name(9, label)
    if std != "其他信息":
        return std
    # 真「其他信息」自身也返回原名; 折叠别名 (分子量/APHA值/含量 等) 保留独立列
    return _S9_PAREN.sub("", label or "").strip()


def _norm_label(num: int, kind: str, label: str) -> str:
    """按节归一化字段标签 (仅 S9 field 做同义合并, 其余原样)."""
    if num == 9 and kind == "field":
        return _norm_s9(label)
    return label


def _collect_columns(results, max_comp: int) -> list[tuple[int, list[tuple[str, str, str]]]]:
    """收集每节列组 (保持首见顺序): (节号, [(kind, label, seq), ...]).

    kind: field / note / comp  (sub 分组标签不单独成列, 直接展示其子字段)
    规则:
      - sub 不单独成列 (页眉/页脚/成分/控制参数/暴露控制 等父级分组)
      - 值全空的 field 列剔除 (父级分组标签如 供应商信息/物质或混合物分类,
        装饰性标签如 物料安全数据表 长标题) — 其子字段信息不丢失
      - S9 field 标签归一化合并同义列 (闪点（℃）与闪点 归并为 闪点)
      - seq: 该列多数型号使用的序号 (同义列跨型号序号不一致时取众数, 供表头显示)
    """
    # 候选列: key=(num, kind, norm_label) → {labels, any_value, seqs:序号计数器}
    cand: dict[tuple[int, str, str], dict] = {}
    order: list[tuple[int, str, str]] = []
    # 永不输出的字段 (无实际用途, 如 S0 页码)
    _EXCLUDE_FIELDS = {(0, "页码")}

    def add(num: int, kind: str, label: str, has_value: bool, seq: str = ""):
        if (num, label) in _EXCLUDE_FIELDS:
            return
        nlabel = _norm_label(num, kind, label)
        key = (num, kind, nlabel)
        if key not in cand:
            cand[key] = {"labels": set(), "any_value": False, "seqs": {}}
            order.append(key)
        cand[key]["labels"].add(label)
        cand[key]["any_value"] |= has_value
        if seq:
            cand[key]["seqs"][seq] = cand[key]["seqs"].get(seq, 0) + 1

    for r in results:
        for num in sorted(r.sections):
            sec = r.sections[num]
            for row in sec.iter_rows():
                if row.kind == "section":
                    continue
                if row.kind == "sub":
                    continue                      # sub 分组不单独成列
                has_value = bool(row.value.strip())
                if row.kind == "field":
                    if row.span and not row.label.strip():
                        add(num, "note", "(总结句)", has_value)
                    else:
                        add(num, "field", row.label, has_value, row.seq)
                elif row.kind == "note":
                    add(num, "note", "(总结句)", has_value)
            # 成分表 → 展开为每成分三列 (多成分: 全部展开, 空型号留空)
            if sec.is_component_table:
                for i in range(max_comp):
                    add(num, "comp", f"成分{i+1}名称", True)
                    add(num, "comp", f"成分{i+1}CAS", True)
                    add(num, "comp", f"成分{i+1}含量", True)

    def _major_seq(seqs: dict[str, int]) -> str:
        """出现频率最高的序号 (众数), 无序号返回空."""
        if not seqs:
            return ""
        return max(seqs, key=lambda s: (seqs[s], -len(s))) if seqs else ""

    # 过滤: 剔除值全空的 field / note 列 (无任何型号有值 → 空列)
    sec_cols: dict[int, list[tuple[str, str, str]]] = {}
    for key in order:
        num, kind, nlabel = key
        if kind != "comp" and not cand[key]["any_value"]:
            continue
        # comp 列 (成分) 无序号; field/note 取众数序号
        seq = _major_seq(cand[key]["seqs"]) if kind != "comp" else ""
        sec_cols.setdefault(num, []).append((kind, nlabel, seq))
    return [(n, sec_cols[n]) for n in sorted(sec_cols)]


def _value_of(result, num: int, kind: str, label: str) -> str:
    sec = result.sections.get(num)
    if not sec:
        return ""
    if kind == "note":
        # 该节全部总结句合并 (保持原顺序, 换行连接)
        vals = [row.value for row in sec.iter_rows()
                if (row.kind in ("note", "field")) and row.span and not row.label.strip()]
        return "\n".join(v for v in vals if v.strip())
    if kind == "comp":
        m = re.match(r"成分(\d+)(名称|CAS|含量)", label)
        if not m:
            return ""
        idx = int(m.group(1)) - 1
        if idx >= len(sec.components):
            return ""
        c = sec.components[idx]
        return {"名称": c.name, "CAS": c.cas, "含量": c.conc}[m.group(2)]
    # field: 归一化匹配 (S9 同义列合并), 取首个非空值
    target = _norm_label(num, "field", label)
    for row in sec.iter_rows():
        if row.kind == "field" and row.label.strip() and row.value.strip():
            if _norm_label(num, "field", row.label) == target:
                return row.value
    return ""


def _s0_company_of(result) -> str:
    """从 S0 页脚合并长标题字段推断公司名称.

    部分模板的页脚把 公司名称+型号 合并为单个长标签且值为空, 如
      "广州冠志化工有限公司 2-苯氧基乙醇-MSDS"
      "广州冠志新材料科技有限公司OS-8501-MSDS"
    公司名称 = 标签中第一个以"公司(有限)"结尾的前缀; 匹配不到返回空.
    """
    sec = result.sections.get(0)
    if not sec:
        return ""
    for row in sec.iter_rows():
        if row.kind != "field" or not row.label:
            continue
        if "MSDS" not in row.label or "公司" not in row.label:
            continue
        head = re.split(r"-MSDS", row.label)[0].strip()
        m = re.search(r"(.*?有限公司|.*?公司)", head)
        if m:
            return m.group(1).strip()
    return ""


def _clean(v: str) -> str:
    return v.replace("\r\n", "\n").strip()


def build_pivot_table(in_dir: str | Path, out_path: str | Path) -> dict:
    """生成透视总表.

    Returns:
        {files, cols, rows, sections, failed, out}
        files: 成功读取并入库的文件数; failed: 读取失败文件名清单
    """
    in_dir = Path(in_dir)
    files = sorted(p for p in in_dir.glob("*.docx")
                   if not p.name.startswith("~$"))
    if not files:
        raise FileNotFoundError(f"目录中未找到 docx 文件: {in_dir}")

    # 围栏: 单个文件读取失败 → 跳过并记录, 不中断整体导出
    ok_files: list[Path] = []
    results = []
    failed: list[str] = []
    for p in files:
        try:
            results.append(read_msds(p))
            ok_files.append(p)
        except Exception as exc:
            failed.append(f"{p.name} ({exc.__class__.__name__})")
    if not ok_files:
        raise RuntimeError(f"目录下 {len(files)} 个 docx 全部读取失败:\n" + "\n".join(failed))

    max_comp = max((len(s.components)
                    for r in results for s in r.sections.values()
                    if s.is_component_table), default=0)
    return _write_pivot(results, ok_files, max_comp, out_path, failed)


def build_single_pivot(result, file_name: str, out_path: str | Path) -> dict:
    """生成单个 ParseResult 的透视总表 (标准格式, 与目录级 build_pivot_table 一致).

    供 GUI「导出当前 MSDS/模板」使用: 把当前显示源 (产品/模板) 的结构
    按 型号×节/标签 透视导出, 第一行 Section / 第二行 序号+标签 /
    A 列型号 / A1·A2 留空, 第三行起为该文件的数据行.
    """
    results = [result]
    ok_files = [Path(file_name) if not isinstance(file_name, Path) else file_name]
    max_comp = max((len(s.components)
                    for s in result.sections.values()
                    if s.is_component_table), default=0)
    return _write_pivot(results, ok_files, max_comp, out_path, failed=[])


def _write_pivot(results, ok_files, max_comp: int,
                 out_path: str | Path, failed: list[str]) -> dict:
    """核心: 用已读取的 results 生成标准透视总表并保存.

    标准格式:
      - 第一行: Section 节标题 (合并单元格)
      - 第二行: 序号 + 标签 (sub 分组不单独成列)
      - A 列: 型号; A1 / A2 留空
      - 第三行起: 每型号一行, 按 节→标签 对照填入
      - S9: 同义标签归一化, 空值标「无数据」
      - 成分表: 每成分 名称/CAS/含量 三列展开
    """
    columns = _collect_columns(results, max_comp)

    wb = Workbook()
    ws = wb.active
    ws.title = "入库总表"

    # ---- 表头两行 ----
    col = 2  # A 列留作型号
    span_ranges: list[tuple[int, int, str]] = []   # (start_col, end_col, 节标题)
    for num, cols in columns:
        start = col
        for kind, label, seq in cols:
            # 第二行: 序号 + 标签 (序号取该列多数型号的编号; 无序号只显示标签)
            head = f"{seq} {label}".strip() if seq else label
            c = ws.cell(2, col, head)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = TAG_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            col += 1
        end = col - 1
        span_ranges.append((start, end, num))

    # ---- 第一行: 节标题 ----
    for start, end, num in span_ranges:
        sec = results[0].sections.get(num)
        title = sec.full_title if sec else f"第{num}节"
        if num == 0:
            title = "0 页眉/页脚"
        ws.merge_cells(start_row=1, start_column=start,
                       end_row=1, end_column=end)
        c = ws.cell(1, start, title)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        for cc in range(start, end + 1):
            ws.cell(1, cc).border = BORDER
        ws.row_dimensions[1].height = 28

    # ---- A1 / A2 留空 ----
    ws.cell(1, 1, "")
    ws.cell(2, 1, "")
    for r_ in (1, 2):
        c = ws.cell(r_, 1)
        c.fill = HEAD_FILL if r_ == 1 else TAG_FILL

    # ---- 数据行 ----
    for ri, (p, r) in enumerate(zip(ok_files, results), start=3):
        model = _model_of(r, p.name)
        mc = ws.cell(ri, 1, model)
        mc.font = Font(name=FONT, size=10, bold=True)
        mc.fill = MODEL_FILL
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = BORDER

        col = 2
        for num, cols in columns:
            for kind, label, seq in cols:
                v = _clean(_value_of(r, num, kind, label))
                # S0 页脚公司名称缺失时, 从合并长标题字段推断 (如 "广州冠志化工有限公司 2-苯氧基乙醇-MSDS")
                if not v and num == 0 and kind == "field" and label == "公司名称":
                    v = _s0_company_of(r)
                # 空值统一标注 无数据: 区分"确实没有"与"读取遗漏" (S9 理化特性同此规则)
                if not v:
                    v = "无数据"
                c = ws.cell(ri, col, v)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = BORDER
                if kind == "note" and v:
                    c.fill = NOTE_FILL
                col += 1
        ws.row_dimensions[ri].height = 60

    # ---- 列宽 / 冻结 ----
    ws.column_dimensions["A"].width = 20
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "B3"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "files": len(ok_files),
        "cols": col - 1,
        "rows": len(ok_files),
        "sections": len(columns),
        "failed": failed,
        "out": str(out_path),
    }
