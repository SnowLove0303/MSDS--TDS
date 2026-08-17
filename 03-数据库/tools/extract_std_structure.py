# -*- coding: utf-8 -*-
"""从标准字段库 Excel 提取「父子级关系逻辑 + 全部标签」为规范 JSON.

输出结构 (outputs/标准字段结构.json):
{
  "source": Excel路径,
  "version": "1.0",
  "sections": [                       # 按节组织
    {
      "number": 0,
      "title": "Section0 页眉页脚",
      "groups": [                     # 父子级分组 (按 Excel 列序)
        {"parent": "0.1 页眉", "anchor": true,
         "fields": ["Version", "产品名称"]},      # anchor=true 表示该组父级自身是占位
        ...
      ]
    }
  ],
  "columns": [                        # 扁平列序 (矩阵输出用, 与 Excel 列序一致)
    {"sec": 0, "parent": "0.1 页眉", "label": "页眉", "type": "anchor"},
    {"sec": 0, "parent": "0.1 页眉", "label": "Version", "type": "field"},
    ...
  ]
}
"""
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

STD_XLSX = r'F:\正式项目与模块化内容\Word 覆写模块\数据库\正式库\标准字段数据库Excel\Excel.xlsx'
OUT_JSON = r'F:\正式项目与模块化内容\Word 覆写模块\结构读取\outputs\标准字段结构.json'

_SEQ_RE = re.compile(r'^\d+(?:\.\d+)*\s*')
_COMP_SLOT_RE = re.compile(r'^化学品名称（(\d+)）$')
_COMP_CAS_RE = re.compile(r'^CAS编号（(\d+)）$')
_COMP_CONC_RE = re.compile(r'^含量%（w/w）（(\d+)）$')


def bare(name):
    return _SEQ_RE.sub('', name or '').strip()


def extract():
    wb = load_workbook(STD_XLSX, data_only=True)
    ws = wb['Sheet1']
    ncol = ws.max_column

    # 1) 节边界: 行1 块标题 → 节号
    cur_sec = None
    sec_of_col = {}
    for ci in range(1, ncol + 1):
        r1 = str(ws.cell(1, ci).value or '').strip()
        m = re.match(r'Section(\d+)', r1)
        if m:
            cur_sec = int(m.group(1))
        elif r1:
            m2 = re.match(r'(\d+)\.', r1)
            if m2:
                cur_sec = int(m2.group(1))
        sec_of_col[ci] = cur_sec

    # 2) 逐列: (sec, parent, label, type)
    #    锚点判定: parent 去序号 == label 且 该列是该 (sec,parent) 组第一列
    #    (Excel 中父级锚点列总是组内首列; 8.2 生物限值 组内第二个
    #     "生物限值" 是子表列名, 非锚点)
    first_of_group = {}
    for ci in range(1, ncol + 1):
        parent = str(ws.cell(2, ci).value or '').strip()
        sec = sec_of_col.get(ci)
        if sec is None or not parent:
            continue
        first_of_group.setdefault((sec, parent), ci)
    columns = []
    for ci in range(1, ncol + 1):
        parent = str(ws.cell(2, ci).value or '').strip()
        label = str(ws.cell(3, ci).value or '').strip()
        sec = sec_of_col.get(ci)
        if sec is None or not label:
            continue
        if _COMP_SLOT_RE.match(label) or _COMP_CAS_RE.match(label) \
                or _COMP_CONC_RE.match(label):
            typ = 'comp'
        elif parent == '通栏说明':
            typ = 'note'
        elif (bare(parent) == label
                and first_of_group.get((sec, parent)) == ci):
            typ = 'anchor'
        else:
            typ = 'field'
        columns.append({'sec': sec, 'parent': parent, 'label': label,
                        'type': typ})

    # 3) 按节分组整理父子级
    sections = []
    sec_order = []
    for col in columns:
        sec = col['sec']
        if sec not in sec_order:
            sec_order.append(sec)
    # 标准节标题 (GB/T 16483)
    SEC_TITLES = {
        0: 'Section0 页眉页脚', 1: 'Section1 物料及供应商标识',
        2: 'Section2 危险性概述', 3: 'Section3 成分/组成资料',
        4: 'Section4 急救措施', 5: 'Section5 消防措施',
        6: 'Section6 意外泄漏措施', 7: 'Section7 操作和储存',
        8: 'Section8 接触控制/个人防护', 9: 'Section9 物理和化学特性',
        10: 'Section10 稳定性和反应性', 11: 'Section11 毒性资料',
        12: 'Section12 生态信息', 13: 'Section13 处理注意事项',
        14: 'Section14 运输信息', 15: 'Section15 法规信息',
        16: 'Section16 其他信息',
    }
    for sec in sec_order:
        title = SEC_TITLES.get(sec, f'Section{sec}')
        # 找该节行1 块标题 (节下字段组的标题)
        groups = []
        cur_parent = None
        cur_fields = []
        cur_anchor = False
        cur_note = False

        def flush():
            nonlocal cur_parent, cur_fields, cur_anchor, cur_note
            if cur_parent is not None:
                groups.append({
                    'parent': cur_parent,
                    'anchor': cur_anchor,
                    'note': cur_note,
                    'fields': cur_fields,
                })
            cur_parent = None
            cur_fields = []
            cur_anchor = False
            cur_note = False

        for col in columns:
            if col['sec'] != sec:
                continue
            parent = col['parent'] or f'(节下)'
            typ = col['type']
            if typ == 'anchor':
                flush()
                cur_parent = parent
                cur_anchor = True
                cur_fields = []
            elif cur_parent is None:
                flush()
                cur_parent = parent
                cur_anchor = False
                cur_note = (typ == 'note')
                cur_fields = [col['label']]
            elif parent == cur_parent:
                cur_fields.append(col['label'])
            else:
                flush()
                cur_parent = parent
                cur_anchor = False
                cur_note = (typ == 'note')
                cur_fields = [col['label']]
        flush()
        # 节标题: 优先用标准节标题 (SEC_TITLES); 行1 块标题仅当该节
        # 无标准标题定义时兜底 (当前 0..16 全有定义)
        sections.append({'number': sec, 'title': title,
                         'groups': groups})

    return {'source': STD_XLSX, 'version': '1.0',
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'sections': sections, 'columns': columns}


def main():
    data = extract()
    Path(OUT_JSON).write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    # 控制台摘要
    n_field = sum(1 for c in data['columns'] if c['type'] == 'field')
    n_anchor = sum(1 for c in data['columns'] if c['type'] == 'anchor')
    n_comp = sum(1 for c in data['columns'] if c['type'] == 'comp')
    n_note = sum(1 for c in data['columns'] if c['type'] == 'note')
    print(f'✅ 标准字段结构已提取 → {OUT_JSON}')
    print(f'   节 {len(data["sections"])} | 列 {len(data["columns"])}'
          f' (字段 {n_field} / 锚点 {n_anchor} / 成分 {n_comp} / 通栏 {n_note})')
    print()
    for sec in data['sections']:
        print(f'== {sec["title"]} ==')
        for g in sec['groups']:
            tag = '锚点' if g['anchor'] else ('通栏' if g['note'] else '分组')
            print(f'   [{tag}] {g["parent"]}: {g["fields"]}')


if __name__ == '__main__':
    main()
