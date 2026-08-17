# -*- coding: utf-8 -*-
"""批量填充标准字段库 Excel 数据区 (fill_std_library).

以标准库 Excel 表头 (行1 分组 / 行2 父级 / 行3 标签) 为唯一结构,
对每个 MSDS docx 逐列取值, 写入数据区一行.

规则:
  - 表头 (行1-3) 保留, 只重填数据区 (行4.., col2..)
  - 锚点列恒空; 无匹配 → 无数据 (黄底 FFF2CC)
  - S2 其他危险四子列: 关键词分类 (物理和化学危险/健康危害/环境危害/其他危害)
  - S9 物理和化学特性: 占位符值 (无数据/不适用等) 不加入检索结果 → 视为无数据
  - 标签尾 （N）/(N) → multi_idx (生态毒性(2) 扩列)
  - 扩列列 sec 以 base 标签所在节修正

用法:
  python fill_std_library.py <标准库xlsx> <docx或目录...> [--row-start 4] [--model-col 1]
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.docx_reader import read_msds
from core.std_output import tpl_values, cell_value, std_merged_counts, bare

ND_FILL = PatternFill('solid', fgColor='FFF2CC')
PLAIN = PatternFill(fill_type=None)

PLACEHOLDER_RE = re.compile(
    r'^(无数据|无数据资料|不适用|无适用资料|未提供|无资料|无|N/A|n/a|None|'
    r'无数据。|不适用。|无。|Not applicable|not applicable|NA|—|--)$')

_TRAIL_N_RE = re.compile(r'^(.*)[（(](\d+)[）)]$')


def has_data(v):
    v = (v or '').strip()
    return bool(v) and not PLACEHOLDER_RE.match(v)


# ---------------- S2 其他危险四子列分类 ----------------
S2_RULES = [
    ('物理和化学危险', ['物理和化学', '爆炸', '易燃', '氧化', '可燃', '高压']),
    ('健康危害', ['皮肤', '眼', '吸入', '食入', '刺激', '灼伤', '呼吸',
                 '过敏', '毒性', '致畸', '致癌']),
    ('环境危害', ['水生', '水体', '土壤', '环境', '生态']),
]
_EMPTY_LABEL_LINE_RE = re.compile(r'^[^：\n]+：\s*$')


def classify_other_hazards(values):
    out = {'物理和化学危险': [], '健康危害': [], '环境危害': [], '其他危害': []}
    for v in values:
        v = (v or '').strip()
        if not v or v == '无':
            continue
        lines = [ln.strip() for ln in v.split('\n') if ln.strip()]
        if lines and all(_EMPTY_LABEL_LINE_RE.match(ln) for ln in lines):
            continue
        hit = None
        for sub, kws in S2_RULES:
            if any(k in v for k in kws):
                hit = sub
                break
        out[hit or '其他危害'].append(v)
    return out


# ---------------- 列结构 ----------------
def load_columns(ws):
    ncol = ws.max_column
    cur_sec = None
    sec_of_col = {}
    for ci in range(1, ncol + 1):
        r1 = str(ws.cell(1, ci).value or '').strip()
        m = re.match(r'Section(\d+)', r1)
        if m:
            cur_sec = int(m.group(1))
        elif r1:
            m2 = re.match(r'(\d+)\.', r1)
            if m2:
                cur_sec = int(m2.group(1))
        sec_of_col[ci] = cur_sec

    # label → sec 映射 (扩列列 sec 修正用)
    label_sec = {}
    for ci in range(1, ncol + 1):
        sec = sec_of_col.get(ci)
        label = str(ws.cell(3, ci).value or '').strip()
        if ci == 1 or sec is None or not label:
            continue
        if not _TRAIL_N_RE.match(label):
            label_sec.setdefault(label, sec)

    cols = []
    for ci in range(1, ncol + 1):
        sec = sec_of_col.get(ci)
        label = str(ws.cell(3, ci).value or '').strip()
        if ci == 1 or sec is None or not label:
            continue
        m = _TRAIL_N_RE.match(label)
        if m and m.group(1) in label_sec:
            sec = label_sec[m.group(1)]
        parent = str(ws.cell(2, ci).value or '').strip()
        if re.match(r'^化学品名称（\d+）$', label) or \
           re.match(r'^CAS编号（\d+）$', label) or \
           re.match(r'^含量%（w/w）（\d+）$', label):
            typ = 'comp'
        elif parent == '通栏说明':
            typ = 'note'
        elif bare(parent) == label and parent:
            typ = 'anchor'
        else:
            typ = 'field'
        cols.append({'ci': ci, 'sec': sec, 'parent': parent, 'label': label,
                     'type': typ})
    return cols


def fill_row(ws, ri, result, cols, std_cnt):
    exact, merged, comps, notes = tpl_values(result)
    s2_sub = None
    vals2 = merged.get(2, {}).get('其他危险')
    if vals2:
        s2_sub = classify_other_hazards(vals2)
    for c in cols:
        label = c['label']
        if c['type'] == 'anchor':
            v = None
        elif c['sec'] == 2 and label in ('物理和化学危险', '健康危害',
                                         '环境危害', '其他危害'):
            lst = (s2_sub or {}).get(label, [])
            v = '\n'.join(lst) if lst else None
        elif c['type'] == 'comp':
            v = cell_value(result, c['sec'], label, c['type'], 1,
                           std_cnt, exact, merged, comps, notes)
        else:
            m = _TRAIL_N_RE.match(label)
            multi_idx, base = (int(m.group(2)), m.group(1)) if m else (1, label)
            v = cell_value(result, c['sec'], base, c['type'], multi_idx,
                           std_cnt, exact, merged, comps, notes)
            # S9: 占位符值不加入检索结果 → 视为无数据
            if c['sec'] == 9 and not has_data(v):
                v = None
        cell = ws.cell(ri, c['ci'])
        if v is not None and str(v).strip():
            cell.value = v
            cell.fill = PLAIN
        else:
            cell.value = '无数据'
            cell.fill = ND_FILL


# 目录扫描时排除的非数据文件 (模板/输出/正文/方案/PEA 模板等)
_EXCLUDE = ('模板', '输出', '正文', '方案', '~$')


def collect_files(inputs):
    files = []
    for raw in inputs:
        p = Path(raw)
        if p.is_dir():
            files.extend(sorted(f for f in p.glob('*.docx')
                                if not f.name.startswith('~$')
                                and not any(x in f.name for x in _EXCLUDE)))
        elif p.exists() and p.suffix.lower() == '.docx':
            files.append(p)
    return sorted(set(files), key=lambda x: str(x).lower())


def main():
    args = [a for a in sys.argv[1:]]
    if len(args) < 2:
        print(__doc__)
        sys.exit(1)
    xlsx, inputs = args[0], args[1:]

    wb = load_workbook(xlsx)
    ws = wb.active
    cols = load_columns(ws)
    std_cnt = std_merged_counts([c for c in cols if c['type'] == 'field'])
    files = collect_files(inputs)

    n_anchor = sum(1 for c in cols if c['type'] == 'anchor')
    print(f'列 {len(cols)} (锚点 {n_anchor}) | 文件 {len(files)}')

    failed = []
    for i, p in enumerate(files, start=4):
        try:
            r = read_msds(p)
            fill_row(ws, i, r, cols, std_cnt)
            print(f'  ✓ row{i} {p.name}')
        except Exception as exc:
            failed.append((p.name, str(exc)))
            print(f'  ✗ row{i} {p.name}: {exc}')

    wb.save(xlsx)
    print(f'\n✅ 填充完成 → {xlsx} | 失败 {len(failed)}')
    for name, err in failed:
        print(f'   - {name}: {err}')


if __name__ == '__main__':
    main()
